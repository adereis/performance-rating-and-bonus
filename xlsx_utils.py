"""
Utilities for parsing Workday XLSX exports.

This module provides functions for:
- Analyzing XLSX files (counting employees, detecting columns)
- Parsing employee data from Workday exports
- Creating Employee records from parsed data

Note on currency columns:
Workday exports include columns like "Bonus Target - Local Currency (XXX)" where
XXX is the manager's home currency (USD, AUD, EUR, etc.). For employees in the
manager's country, this column is empty. For international employees, it contains
the converted value. The code matches any 3-letter currency code pattern.
"""
import openpyxl
import re
from typing import Optional, Tuple, List, Dict, Any
from datetime import datetime


def parse_float(val) -> Optional[float]:
    """Safely parse a value to float."""
    try:
        return float(val) if val else None
    except (ValueError, TypeError):
        return None


def convert_decimal_to_percent(val: Optional[float], is_new_format: bool = True) -> Optional[float]:
    """
    Convert Workday decimal format to percentage.

    The new Workday format (2025+) stores percentages as decimals:
    - 0.05 = 5% bonus target
    - 0.10 = 10% bonus target
    - 1.20 = 120% bonus allocation (exceptional performer)

    The old format stored as actual percentages (5 = 5%).

    Args:
        val: The value from the spreadsheet
        is_new_format: If True, multiply by 100 to convert decimal to percent

    Returns:
        The percentage value (5 for 5%, 120 for 120%), or None
    """
    if val is None:
        return None
    if is_new_format:
        return val * 100
    return val


def parse_date(val) -> Optional[datetime]:
    """
    Parse a date value from Workday export.

    Handles both datetime objects (from Excel) and string formats.

    Args:
        val: The date value (datetime, string, or None)

    Returns:
        datetime object or None
    """
    if val is None:
        return None

    if isinstance(val, datetime):
        return val

    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        # Try common date formats
        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%dT%H:%M:%S']:
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue

    return None


def get_current_period_name() -> str:
    """
    Get the current period name in Workday format (e.g., "CY25 Q1").

    Uses calendar year (CY) and quarter based on current date.
    Q1 = Jan-Mar, Q2 = Apr-Jun, Q3 = Jul-Sep, Q4 = Oct-Dec
    """
    now = datetime.now()
    year_short = now.year % 100  # e.g., 2025 -> 25
    quarter = (now.month - 1) // 3 + 1  # 1-4
    return f"CY{year_short:02d} Q{quarter}"


def get_previous_period_name() -> str:
    """
    Get the previous period name in Workday format (e.g., "CY24 Q4").

    Returns the quarter immediately before the current one.
    Handles year rollover (Q1 → previous year's Q4).
    """
    now = datetime.now()
    quarter = (now.month - 1) // 3 + 1  # 1-4

    if quarter == 1:
        # Previous quarter is Q4 of the previous year
        prev_year_short = (now.year - 1) % 100
        return f"CY{prev_year_short:02d} Q4"
    else:
        year_short = now.year % 100
        return f"CY{year_short:02d} Q{quarter - 1}"


def detect_import_type(period_name: str, spreadsheet_type: str = 'bonus') -> Dict[str, Any]:
    """
    Suggest import type based on period from metadata and spreadsheet type.

    For bonus files: Compares the file's period to the current AND previous quarter
    to suggest whether this is likely a current period import or historical archive.
    The previous quarter is included because bonus processing typically happens in
    the quarter following the performance period (e.g., Q4 bonuses are processed in Q1).

    For talent files: Always suggests 'current' since talent calibration reports
    are for the current cycle and don't include period metadata.

    The user should confirm the suggested import type.

    Args:
        period_name: Period name from metadata (e.g., "CY25 Q3")
        spreadsheet_type: 'bonus' or 'talent'

    Returns:
        Dict with:
            - suggested_type: "current" or "historical"
            - period_id: Suggested period ID (e.g., "2025-Q3")
            - period_display: Period display name (e.g., "CY25 Q3")
            - current_period: The current period for reference
            - is_current_period: Whether file period matches current
            - is_talent_file: Whether this is a talent calibration file
    """
    current_period = get_current_period_name()

    # Talent files don't have period metadata - always import as current
    if spreadsheet_type == 'talent':
        return {
            'suggested_type': 'current',
            'period_id': None,
            'period_display': 'Current Cycle',
            'current_period': current_period,
            'is_current_period': True,
            'is_talent_file': True
        }

    # Parse period_name to generate period_id
    # Format: "CY25 Q3" or "CY25-Q3" → "2025-Q3"
    period_id = None
    if period_name:
        match = re.match(r'CY(\d{2})\s*[-]?\s*([QH]\d)', period_name)
        if match:
            year = 2000 + int(match.group(1))
            period_suffix = match.group(2)  # Q1, Q2, H1, etc.
            period_id = f"{year}-{period_suffix}"

    previous_period = get_previous_period_name()
    is_exact_current = period_name == current_period if period_name else False
    is_previous = period_name == previous_period if period_name else False

    # Treat previous quarter as "current" since bonus processing typically
    # happens the quarter after performance period (e.g., Q4 bonuses in Q1)
    is_current = is_exact_current or is_previous

    return {
        'suggested_type': 'current' if is_current else 'historical',
        'period_id': period_id,
        'period_display': period_name or 'Unknown',
        'current_period': current_period,
        'is_current_period': is_current,
        'is_talent_file': False
    }


def extract_workday_metadata(rows: List[tuple], header_idx: int) -> Dict[str, Any]:
    """
    Extract metadata from Workday export header rows.

    Supports two formats:

    NEW FORMAT (2025+):
    - Row 0: Report title (e.g., "RH Compensation Review Process - Bonus")
    - Row 2: Period info (e.g., "Compensation Review: Bonus - CY25 Q4")
    - Row 4: Manager org context
    - No budget row - total_pool calculated from sum of bonus targets

    OLD FORMAT (pre-2025):
    - Row 0: Report title with period (e.g., "Associate Awards:: ... Bonus - CY25 Q3")
    - Row 3: Budget summary (type, total_spend, "of", total_pool, %, style, currency)

    Args:
        rows: All rows from the spreadsheet
        header_idx: Index of the header row (metadata is above this)

    Returns:
        Dict with:
            - period_name: str or None (e.g., "CY25 Q3")
            - total_pool: float or None (budget amount, None for new format)
            - currency: str or None (e.g., "USD")
            - report_title: str or None (full title from row 0)
            - is_new_format: bool (True if new 2025+ format detected)
    """
    metadata = {
        'period_name': None,
        'total_pool': None,
        'currency': None,
        'report_title': None,
        'is_new_format': False,
    }

    if not rows or header_idx < 1:
        return metadata

    # Row 0: Extract report title
    if len(rows) > 0 and rows[0]:
        title = rows[0][0]
        if title and isinstance(title, str):
            metadata['report_title'] = title.strip()

    # Detect format based on row 2 or 3 content
    # New format has "Compensation Review: Bonus - CY25 Q4" pattern
    # In-progress reviews: period in row 2 (index 2)
    # Completed reviews: period in row 3 (index 3)
    for row_idx in [2, 3]:
        if len(rows) > row_idx and rows[row_idx]:
            for cell in rows[row_idx]:
                if cell and isinstance(cell, str) and 'Bonus - CY' in cell:
                    metadata['is_new_format'] = True
                    # Extract period from "Compensation Review: Bonus - CY25 Q4"
                    match = re.search(r'Bonus - (CY\d{2}\s*Q\d)', cell)
                    if match:
                        metadata['period_name'] = match.group(1).strip()
                    break
            if metadata['is_new_format']:
                break

    if metadata['is_new_format']:
        # New format: no budget row, total_pool will be calculated from data
        # Currency will be extracted from column headers (e.g., "Bonus Target (USD)")
        return metadata

    # Old format: try to extract period from row 0 title
    if not metadata['period_name'] and metadata['report_title']:
        title = metadata['report_title']
        # Extract period from patterns like "Bonus - CY25 Q3" or "Bonus CY25-H1"
        period_patterns = [
            r'[-–]\s*([A-Z]{2}\d{2}\s+[A-Z]\d)',      # "- CY25 Q3"
            r'[-–]\s*([A-Z]{2}\d{2}-[A-Z]\d)',        # "- CY25-Q3"
            r'[-–]\s*(\d{4}\s+[A-Z]\d)',              # "- 2025 Q3"
            r'[-–]\s*(\d{4}-[A-Z]\d)',                # "- 2025-Q3"
            r'[-–]\s*([A-Z]{2}\d{2}\s+H\d)',          # "- CY25 H1"
            r'[-–]\s*(\d{4}\s+H\d)',                  # "- 2025 H1"
        ]
        for pattern in period_patterns:
            match = re.search(pattern, title)
            if match:
                metadata['period_name'] = match.group(1).strip()
                break

    # Old format Row 4 (index 3): Budget summary row
    # Format: [type, total_spend, "of", total_pool, %, style, currency, ...]
    if len(rows) > 3 and rows[3]:
        budget_row = rows[3]

        # Check if this looks like the budget row (first cell is "Bonus" or similar)
        if budget_row[0] and isinstance(budget_row[0], str):
            first_cell = budget_row[0].lower().strip()
            if first_cell in ('bonus', 'merit', 'compensation'):
                # Extract total pool (column 4, index 3)
                if len(budget_row) > 3 and budget_row[3]:
                    pool_val = budget_row[3]
                    if isinstance(pool_val, (int, float)):
                        metadata['total_pool'] = float(pool_val)
                    elif isinstance(pool_val, str):
                        # Parse formatted number like "210,910.07"
                        try:
                            cleaned = pool_val.replace(',', '').strip()
                            metadata['total_pool'] = float(cleaned)
                        except (ValueError, TypeError):
                            pass

                # Extract currency (column 7, index 6)
                if len(budget_row) > 6 and budget_row[6]:
                    currency = budget_row[6]
                    if isinstance(currency, str) and len(currency) == 3 and currency.isalpha():
                        metadata['currency'] = currency.upper()

    return metadata


def validate_workday_format(rows: List[tuple], header_idx: Optional[int], headers: List[str], metadata: Dict[str, Any] = None) -> Tuple[bool, str]:
    """
    Validate that the file is a proper Workday export with required metadata.

    Checks:
    1. Header row was found
    2. Required columns exist (Associate, Associate ID)
    3. At least one data row exists
    4. Workday metadata present (period name, bonus pool) - new format required

    Args:
        rows: All rows from the spreadsheet
        header_idx: Index of the header row (or None if not found)
        headers: List of header strings (empty if header_idx is None)
        metadata: Optional pre-extracted metadata dict

    Returns:
        Tuple of (is_valid, error_message)
    """
    if header_idx is None:
        # Try to give a helpful error message
        sample_headers = []
        for row in rows[:5]:
            if row:
                non_empty = [str(c)[:30] for c in row[:5] if c]
                if non_empty:
                    sample_headers = non_empty
                    break

        return False, (
            "Could not find expected Workday columns.\n\n"
            "Expected columns: Associate, Associate ID, Supervisory Organization, "
            "Current Job Profile, Currency, Bonus Target - Local Currency\n\n"
            f"Found in file: {', '.join(sample_headers) if sample_headers else '(empty)'}\n\n"
            "Please export from Workday using the standard bonus allocation report."
        )

    # Check for required columns
    # Note: Talent reports use 'Worker' instead of 'Associate'
    normalized = [h.lower().strip() for h in headers if h]

    missing = []
    if 'associate' not in normalized and 'worker' not in normalized:
        missing.append('Associate (or Worker)')
    if 'associate id' not in normalized:
        missing.append('Associate ID')

    if missing:
        return False, f"Missing required columns: {', '.join(missing)}"

    # Check for data rows
    data_rows = 0
    for row in rows[header_idx + 1:]:
        if not _is_empty_row(row):
            data_rows += 1
            if data_rows >= 1:
                break

    if data_rows == 0:
        return False, "No employee data found after header row"

    # Check for required metadata (Workday format validation)
    # New format (2025+): total_pool will be calculated from data, not required in metadata
    # Old format: requires total_pool in metadata
    if metadata is not None:
        is_new_format = metadata.get('is_new_format', False)
        if not is_new_format and not metadata.get('total_pool'):
            return False, (
                "Missing bonus pool metadata.\n\n"
                "This file appears to be using an old export format that lacks required metadata.\n\n"
                "Please export a fresh file from Workday - the current export format includes:\n"
                "  • Row 1: Report title with period (e.g., 'Associate Awards:: ... Bonus - CY25 Q1')\n"
                "  • Row 4: Budget summary with total pool amount\n\n"
                "The bonus pool from Workday metadata is required for accurate calculations."
            )

    return True, ''


def _find_header_row(rows: List[tuple]) -> Optional[int]:
    """
    Find the header row by looking for required Workday column names.

    Scans rows until it finds one containing 'Associate ID' and either
    'Associate' or 'Worker' (case-insensitive). Bonus reports use 'Associate',
    while talent reports use 'Worker'.

    Args:
        rows: List of row tuples from the spreadsheet

    Returns:
        Index of the header row, or None if not found
    """
    for idx, row in enumerate(rows):
        if not row:
            continue
        # Normalize cell values for comparison
        normalized = {str(cell).lower().strip() for cell in row if cell}

        # Must have 'associate id' AND either 'associate' OR 'worker'
        has_id = 'associate id' in normalized
        has_name = 'associate' in normalized or 'worker' in normalized

        if has_id and has_name:
            return idx

    return None


def _is_empty_row(row: tuple) -> bool:
    """
    Check if a row is empty (all cells are None or whitespace-only strings).

    Args:
        row: Row tuple from the spreadsheet

    Returns:
        True if the row is empty, False otherwise
    """
    if not row:
        return True
    return all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row)


def analyze_xlsx(file_path: str) -> Dict[str, Any]:
    """
    Analyze an XLSX file and return metadata about its contents.

    Args:
        file_path: Path to the XLSX file

    Returns:
        Dict with:
            - success: bool
            - employee_count: int
            - has_bonus_column: bool (has 'Proposed Percent of Target Bonus')
            - notes_count: int (employees with Notes field)
            - partial_count: int (employees with bonus but no notes/rating)
            - columns: list of column headers
            - error: str (if success is False)
            - metadata: dict with period_name, total_pool, currency, report_title
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb.active

        rows = list(sheet.iter_rows(values_only=True))

        # Find the header row dynamically
        header_idx = _find_header_row(rows)
        headers = [str(h).strip() if h else '' for h in rows[header_idx]] if header_idx is not None else []

        # Detect spreadsheet type (talent vs bonus)
        spreadsheet_type = detect_spreadsheet_type(headers) if headers else 'bonus'

        # Extract metadata from header rows (needed for validation of bonus files)
        metadata = extract_workday_metadata(rows, header_idx) if header_idx is not None else {}

        # Validate the file format
        # For talent files, skip metadata validation (they don't have bonus pool)
        validation_metadata = None if spreadsheet_type == 'talent' else metadata
        is_valid, validation_error = validate_workday_format(rows, header_idx, headers, validation_metadata)
        if not is_valid:
            wb.close()
            return {
                'success': False,
                'error': validation_error
            }

        # Count employees (data rows start after header)
        employee_count = 0
        notes_count = 0
        allocation_count = 0  # Employees with proposed bonus allocation

        # Find column indices
        col_indices = find_column_indices(headers)

        for row in rows[header_idx + 1:]:
            # Skip empty rows
            if _is_empty_row(row):
                continue

            # Skip rows without an associate value
            associate_idx = col_indices.get('associate')
            if associate_idx is not None:
                associate_val = row[associate_idx] if associate_idx < len(row) else None
                if not associate_val or (isinstance(associate_val, str) and not associate_val.strip()):
                    continue

            employee_count += 1

            # Check for notes
            notes_idx = col_indices.get('notes')
            if notes_idx is not None and notes_idx < len(row) and row[notes_idx]:
                notes_count += 1

            # Check for bonus allocation
            bonus_idx = col_indices.get('proposed_percent_of_target')
            if bonus_idx is not None and bonus_idx < len(row) and row[bonus_idx]:
                allocation_count += 1

        wb.close()

        # Suggest import type based on period and spreadsheet type
        import_detection = detect_import_type(metadata.get('period_name'), spreadsheet_type)

        return {
            'success': True,
            'employee_count': employee_count,
            'spreadsheet_type': spreadsheet_type,  # 'talent' or 'bonus'
            'has_bonus_column': col_indices.get('proposed_percent_of_target') is not None,
            'notes_count': notes_count,
            'allocation_count': allocation_count,
            'columns': headers,
            'metadata': metadata,
            'import_detection': import_detection
        }

    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }


def find_column_indices(headers: List[str]) -> Dict[str, Optional[int]]:
    """
    Find column indices for known Workday export fields.

    Workday exports use the MANAGER'S home currency for conversion columns.
    For example, an Australian manager sees "(AUD)" columns, a US manager
    sees "(USD)" columns. This function uses regex to match any 3-letter
    currency code in parentheses.

    Args:
        headers: List of header strings

    Returns:
        Dict mapping field names to column indices (or None if not found)
    """
    # Normalize headers for matching
    normalized = [h.lower().strip() if h else '' for h in headers]

    indices = {}

    # Map of field name -> possible header variations
    # Variations with currency codes use regex patterns (marked with 're:' prefix)
    #
    # New Workday format (2025+):
    # - Uses 'Direct Manager' instead of 'Supervisory Organization'
    # - Uses 'Job Title' instead of 'Current Job Profile'
    # - Uses 'Error' (singular) instead of 'Errors'
    # - Uses 'Bonus Target (XXX)' pattern instead of 'Bonus Target - Local Currency (XXX)'
    # - Uses 'Base Pay All Countries (XXX)' pattern
    # - Includes new fields: Management Level, Country, Hire Date, Performance Review Name/Rating
    field_mappings = {
        'associate': ['associate'],
        'associate_id': ['associate id'],
        # New format uses 'Direct Manager', old format used 'Supervisory Organization'
        'supervisory_org': ['direct manager', 'supervisory organization'],
        # New format uses 'Job Title', old format used 'Current Job Profile'
        'job_profile': ['job title', 'current job profile'],
        'photo': ['photo'],
        # New format uses 'Error' (singular), old format used 'Errors'
        'errors': ['error', 'errors'],
        # Base pay columns - new format uses different naming pattern
        'base_pay': [
            'base pay all countries (local)',
            'current base pay - all countries',
            'current base pay all countries'
        ],
        # Match any 3-letter currency code: (USD), (AUD), (GBP), (EUR), etc.
        'base_pay_converted': [
            're:base pay all countries \\([a-z]{3}\\)',
            're:current base pay - all countries \\([a-z]{3}\\)',
            're:current base pay all countries \\([a-z]{3}\\)'
        ],
        'currency': ['currency'],
        'grade': ['grade'],
        'annual_bonus_target': ['annual bonus target %', 'annual bonus target percent'],
        # New format includes "(As of Report Run Date)" suffix
        'last_bonus_allocation': [
            're:last bonus allocation percent.*',
            'last bonus allocation %',
            'last bonus allocation percent'
        ],
        # New format uses 'Bonus Target (Local)' and 'Bonus Target (XXX)'
        'bonus_target_local': [
            'bonus target (local)',
            'bonus target - local currency',
            'bonus target local currency'
        ],
        'bonus_target_converted': [
            're:bonus target \\([a-z]{3}\\)',
            're:bonus target - local currency \\([a-z]{3}\\)',
            're:bonus target local currency \\([a-z]{3}\\)'
        ],
        # New format uses 'Proposed Bonus Amount (Local)' and 'Proposed Bonus Amount (XXX)'
        'proposed_bonus': [
            'proposed bonus amount (local)',
            'proposed bonus amount'
        ],
        'proposed_bonus_converted': ['re:proposed bonus amount \\([a-z]{3}\\)'],
        'proposed_percent_of_target': ['proposed % of target bonus', 'proposed percent of target bonus'],
        'notes': ['notes', 'single description'],
        'zero_bonus': ['zero bonus allocated'],
        # New fields in 2025 format
        'management_level': ['management level'],
        'country': ['country'],
        'hire_date': ['hire date'],
        'time_in_job_profile': ['time in job profile'],
        'perf_review_name': ['performance review name'],
        'perf_review_rating': ['re:overall performance rating.*'],
    }

    for field, variations in field_mappings.items():
        for var in variations:
            if var.startswith('re:'):
                # Regex pattern matching
                pattern = var[3:]  # Remove 're:' prefix
                for idx, header in enumerate(normalized):
                    if re.match(pattern, header):
                        indices[field] = idx
                        break
                if field in indices:
                    break
            else:
                # Exact string matching
                try:
                    idx = normalized.index(var)
                    indices[field] = idx
                    break
                except ValueError:
                    continue
        if field not in indices:
            indices[field] = None

    return indices


def parse_xlsx_employees(file_path: str) -> Tuple[bool, List[Dict[str, Any]], str, Dict[str, Any]]:
    """
    Parse all employee data from a Workday XLSX export.

    Handles both old and new (2025+) Workday formats:
    - New format: Uses decimal percentages (0.05 = 5%), different column names
    - Old format: Uses integer percentages (5 = 5%)

    Args:
        file_path: Path to the XLSX file

    Returns:
        Tuple of (success, employees_list, error_message, metadata)
        employees_list contains dicts with all parsed fields
        metadata contains period_name, total_pool, currency, is_new_format
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb.active

        rows = list(sheet.iter_rows(values_only=True))

        # Find the header row dynamically
        header_idx = _find_header_row(rows)
        headers = [str(h).strip() if h else '' for h in rows[header_idx]] if header_idx is not None else []

        # Extract metadata from header rows (needed for validation and format detection)
        metadata = extract_workday_metadata(rows, header_idx) if header_idx is not None else {}
        is_new_format = metadata.get('is_new_format', False)

        # Validate the file format (including metadata check)
        is_valid, validation_error = validate_workday_format(rows, header_idx, headers, metadata)
        if not is_valid:
            wb.close()
            return False, [], validation_error, metadata

        col_indices = find_column_indices(headers)

        employees = []

        for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
            # Skip empty rows
            if _is_empty_row(row):
                continue

            # Skip rows without an associate value
            associate_idx = col_indices.get('associate')
            if associate_idx is not None:
                associate_val = row[associate_idx] if associate_idx < len(row) else None
                if not associate_val or (isinstance(associate_val, str) and not associate_val.strip()):
                    continue

            # Get associate ID (required)
            assoc_id_idx = col_indices.get('associate_id')
            if assoc_id_idx is not None and assoc_id_idx < len(row) and row[assoc_id_idx]:
                associate_id = str(row[assoc_id_idx])
            else:
                associate_id = f"TEMP_{i}"

            # Parse percentage fields with format-aware conversion
            annual_bonus_pct = parse_float(_get_val(row, col_indices.get('annual_bonus_target')))
            last_bonus_pct = parse_float(_get_val(row, col_indices.get('last_bonus_allocation')))
            proposed_pct = parse_float(_get_val(row, col_indices.get('proposed_percent_of_target')))

            # New format uses decimal (0.05 = 5%), old format uses integer (5 = 5%)
            if is_new_format:
                annual_bonus_pct = convert_decimal_to_percent(annual_bonus_pct, is_new_format=True)
                last_bonus_pct = convert_decimal_to_percent(last_bonus_pct, is_new_format=True)
                proposed_pct = convert_decimal_to_percent(proposed_pct, is_new_format=True)

            # Build employee dict
            emp = {
                'associate_id': associate_id,
                'associate': str(row[associate_idx]) if associate_idx is not None and associate_idx < len(row) and row[associate_idx] else '',
                'supervisory_organization': _get_str(row, col_indices.get('supervisory_org')),
                'current_job_profile': _get_str(row, col_indices.get('job_profile')),
                'photo': _get_str(row, col_indices.get('photo')),
                'errors': _get_str(row, col_indices.get('errors')),
                'current_base_pay_all_countries': parse_float(_get_val(row, col_indices.get('base_pay'))),
                'current_base_pay_manager_currency': parse_float(_get_val(row, col_indices.get('base_pay_converted'))),
                'currency': _get_str(row, col_indices.get('currency')),
                'grade': _get_str(row, col_indices.get('grade')),
                'annual_bonus_target_percent': annual_bonus_pct,
                'last_bonus_allocation_percent': last_bonus_pct,
                'bonus_target_local_currency': parse_float(_get_val(row, col_indices.get('bonus_target_local'))),
                'bonus_target_manager_currency': parse_float(_get_val(row, col_indices.get('bonus_target_converted'))),
                'proposed_bonus_amount': parse_float(_get_val(row, col_indices.get('proposed_bonus'))),
                'proposed_bonus_amount_manager_currency': parse_float(_get_val(row, col_indices.get('proposed_bonus_converted'))),
                'proposed_percent_of_target_bonus': proposed_pct,
                'notes': _get_str(row, col_indices.get('notes')),
                'zero_bonus_allocated': _get_str(row, col_indices.get('zero_bonus')),
                # New fields in 2025 format
                'management_level': _get_str(row, col_indices.get('management_level')),
                'country': _get_str(row, col_indices.get('country')),
                'hire_date': parse_date(_get_val(row, col_indices.get('hire_date'))),
                'time_in_job_profile': _get_str(row, col_indices.get('time_in_job_profile')),
                'last_perf_review_name': _get_str(row, col_indices.get('perf_review_name')),
                'last_perf_review_rating': _get_str(row, col_indices.get('perf_review_rating')),
            }

            employees.append(emp)

        wb.close()

        # For new format, calculate total_pool from sum of bonus targets
        if is_new_format and not metadata.get('total_pool'):
            total_pool = sum(
                emp.get('bonus_target_manager_currency') or emp.get('bonus_target_local_currency') or 0
                for emp in employees
            )
            metadata['total_pool'] = total_pool

        return True, employees, '', metadata

    except Exception as e:
        return False, [], str(e), {}


def _get_val(row: tuple, idx: Optional[int]) -> Any:
    """Safely get a value from a row by index."""
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _get_str(row: tuple, idx: Optional[int]) -> str:
    """Safely get a string value from a row by index."""
    val = _get_val(row, idx)
    return str(val) if val else ''


# ═══════════════════════════════════════════════════════════════════════════════
# TALENT CALIBRATION IMPORT SUPPORT
# ═══════════════════════════════════════════════════════════════════════════════

# Markers used to detect spreadsheet type
TALENT_MARKERS = ['Performance: What', 'Performance: How', 'Future Talent', 'Movement Readiness']
# Updated for new format: includes both old and new column naming patterns
BONUS_MARKERS = [
    'Bonus Target',                 # Matches both "Bonus Target (USD)" and old "Bonus Target - Local Currency"
    'Annual Bonus Target Percent',  # Present in both formats
    'Base Pay All Countries',       # New format uses this pattern
    'Current Base Pay',             # Old format uses this pattern
    'Proposed Bonus Amount'         # Present in both formats
]


def detect_spreadsheet_type(headers: List[str]) -> str:
    """
    Detect whether a spreadsheet is a talent calibration or bonus report.

    Uses marker columns to determine the type. Talent reports have columns like
    "Performance: What" and "Movement Readiness". Bonus reports have columns
    like "Bonus Target" and "Proposed Bonus Amount".

    Args:
        headers: List of column header strings

    Returns:
        'talent' or 'bonus'
    """
    text = ' '.join(str(h) for h in headers if h)
    talent_score = sum(1 for m in TALENT_MARKERS if m in text)
    bonus_score = sum(1 for m in BONUS_MARKERS if m in text)
    return 'talent' if talent_score > bonus_score else 'bonus'


# Talent column mappings (Spec §5.2)
# Maps Workday column headers to database field names
TALENT_COLUMN_MAP = {
    # Identity (Worker and Associate both map to associate)
    'Associate ID': 'associate_id',
    'Worker': 'associate',
    'Associate': 'associate',

    # Standard context (shared with bonus)
    # Note: Talent files use "Associate's Manager" instead of "Supervisory Organization"
    'Supervisory Organization': 'supervisory_organization',
    "Associate's Manager": 'supervisory_organization',
    'Job Profile': 'current_job_profile',
    'Current Job Profile': 'current_job_profile',

    # Performance Assessment
    'Performance: What': 'talent_perf_what',
    'Performance: How': 'talent_perf_how',
    'Overall Performance Rating': 'talent_overall_perf',
    'Last Talent Assessment Cycle: Overall Performance Rating': 'talent_last_overall_perf',

    # Future Talent
    'Future Talent: Growth Agility': 'talent_growth_agility',
    'Future Talent: Change Agility': 'talent_change_agility',
    'Identified as Future Talent?': 'talent_identified_future',
    'Last Talent Assessment Cycle: Identified as Future Talent?': 'talent_last_identified_future',

    # Movement & Career
    'Movement Readiness': 'talent_movement_readiness',
    'Last Talent Assessment Cycle: Movement Readiness': 'talent_last_movement_readiness',
    'Proposed Talent Actions': 'talent_proposed_actions',

    # Promotion
    'Promotions: Proposed Job Profile & Code': 'talent_promo_job_profile',
    'Promotions: Business Need': 'talent_promo_business_need',
    'Promotions: Expanded Role Scope': 'talent_promo_role_scope',
    'Promotions: Associate Readiness': 'talent_promo_readiness',

    # Extended Identity Context
    'Time in Job Profile': 'time_in_job_profile',
    'Management Level': 'management_level',
    'Job Category': 'job_category',
    'Hire Date': 'hire_date',
    'Length of Service - Worker': 'length_of_service',
    'Region - Location Based': 'region',
    'Country': 'country',

    # Metadata
    'Calibration Status': 'talent_calibration_status',
}


def find_talent_column_indices(headers: List[str]) -> Dict[str, Optional[int]]:
    """
    Find column indices for talent calibration fields.

    Uses TALENT_COLUMN_MAP to map Workday column headers to field names.

    Args:
        headers: List of header strings

    Returns:
        Dict mapping field names to column indices (or None if not found)
    """
    indices = {}

    # Normalize headers for matching
    header_lower = [h.lower().strip() if h else '' for h in headers]
    header_exact = [h.strip() if h else '' for h in headers]

    for workday_col, field_name in TALENT_COLUMN_MAP.items():
        # Try exact match first (case-sensitive for Workday columns)
        if workday_col in header_exact:
            idx = header_exact.index(workday_col)
            # Don't overwrite if already found (first match wins)
            if field_name not in indices:
                indices[field_name] = idx
        else:
            # Fall back to case-insensitive match
            workday_col_lower = workday_col.lower()
            if workday_col_lower in header_lower:
                idx = header_lower.index(workday_col_lower)
                if field_name not in indices:
                    indices[field_name] = idx

    return indices


def parse_talent_xlsx_employees(file_path: str) -> Tuple[bool, List[Dict[str, Any]], str]:
    """
    Parse employee talent calibration data from a Workday XLSX export.

    Unlike bonus imports, talent imports:
    - Don't require bonus pool metadata
    - Map different columns (Performance What/How, Movement Readiness, etc.)
    - Can create new employees if associate_id not found in database

    Args:
        file_path: Path to the XLSX file

    Returns:
        Tuple of (success, employees_list, error_message)
        employees_list contains dicts with talent calibration fields
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb.active

        rows = list(sheet.iter_rows(values_only=True))

        # Find the header row dynamically
        header_idx = _find_header_row(rows)
        if header_idx is None:
            wb.close()
            return False, [], "Could not find header row with Associate and Associate ID columns"

        headers = [str(h).strip() if h else '' for h in rows[header_idx]]

        # Verify this is a talent spreadsheet
        spreadsheet_type = detect_spreadsheet_type(headers)
        if spreadsheet_type != 'talent':
            wb.close()
            return False, [], (
                "This appears to be a bonus spreadsheet, not a talent calibration report.\n\n"
                "Expected columns like: Performance: What, Performance: How, Movement Readiness\n\n"
                "Please import this file through the standard import process."
            )

        col_indices = find_talent_column_indices(headers)

        employees = []

        for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
            # Skip empty rows
            if _is_empty_row(row):
                continue

            # Get associate value (required)
            associate_idx = col_indices.get('associate')
            if associate_idx is not None:
                associate_val = row[associate_idx] if associate_idx < len(row) else None
                if not associate_val or (isinstance(associate_val, str) and not associate_val.strip()):
                    continue

            # Get associate ID (required)
            assoc_id_idx = col_indices.get('associate_id')
            if assoc_id_idx is not None and assoc_id_idx < len(row) and row[assoc_id_idx]:
                associate_id = str(row[assoc_id_idx])
            else:
                associate_id = f"TEMP_{i}"

            # Parse hire_date as datetime if present
            hire_date_val = _get_val(row, col_indices.get('hire_date'))
            hire_date = None
            if hire_date_val:
                if isinstance(hire_date_val, datetime):
                    hire_date = hire_date_val
                elif isinstance(hire_date_val, str):
                    try:
                        # Try common date formats
                        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']:
                            try:
                                hire_date = datetime.strptime(hire_date_val.strip(), fmt)
                                break
                            except ValueError:
                                continue
                    except Exception:
                        pass

            # Parse boolean for identified_future
            identified_future_val = _get_val(row, col_indices.get('talent_identified_future'))
            talent_identified_future = None
            if identified_future_val is not None:
                if isinstance(identified_future_val, bool):
                    talent_identified_future = identified_future_val
                elif isinstance(identified_future_val, str):
                    val_lower = identified_future_val.strip().lower()
                    if val_lower in ('yes', 'true', '1'):
                        talent_identified_future = True
                    elif val_lower in ('no', 'false', '0'):
                        talent_identified_future = False

            # Parse last_identified_future similarly
            last_identified_val = _get_val(row, col_indices.get('talent_last_identified_future'))
            talent_last_identified_future = None
            if last_identified_val is not None:
                if isinstance(last_identified_val, bool):
                    talent_last_identified_future = last_identified_val
                elif isinstance(last_identified_val, str):
                    val_lower = last_identified_val.strip().lower()
                    if val_lower in ('yes', 'true', '1'):
                        talent_last_identified_future = True
                    elif val_lower in ('no', 'false', '0'):
                        talent_last_identified_future = False

            # Build employee dict
            emp = {
                'associate_id': associate_id,
                'associate': str(row[associate_idx]) if associate_idx is not None and associate_idx < len(row) and row[associate_idx] else '',
                'supervisory_organization': _get_str(row, col_indices.get('supervisory_organization')),
                'current_job_profile': _get_str(row, col_indices.get('current_job_profile')),

                # Extended identity
                'management_level': _get_str(row, col_indices.get('management_level')),
                'job_category': _get_str(row, col_indices.get('job_category')),
                'hire_date': hire_date,
                'length_of_service': _get_str(row, col_indices.get('length_of_service')),
                'time_in_job_profile': _get_str(row, col_indices.get('time_in_job_profile')),
                'region': _get_str(row, col_indices.get('region')),
                'country': _get_str(row, col_indices.get('country')),

                # Performance Assessment
                'talent_perf_what': _get_str(row, col_indices.get('talent_perf_what')),
                'talent_perf_how': _get_str(row, col_indices.get('talent_perf_how')),
                'talent_overall_perf': _get_str(row, col_indices.get('talent_overall_perf')),
                'talent_last_overall_perf': _get_str(row, col_indices.get('talent_last_overall_perf')),

                # Future Talent
                'talent_growth_agility': _get_str(row, col_indices.get('talent_growth_agility')),
                'talent_change_agility': _get_str(row, col_indices.get('talent_change_agility')),
                'talent_identified_future': talent_identified_future,
                'talent_last_identified_future': talent_last_identified_future,

                # Movement & Career
                'talent_movement_readiness': _get_str(row, col_indices.get('talent_movement_readiness')),
                'talent_last_movement_readiness': _get_str(row, col_indices.get('talent_last_movement_readiness')),
                'talent_proposed_actions': _get_str(row, col_indices.get('talent_proposed_actions')),

                # Promotion
                'talent_promo_job_profile': _get_str(row, col_indices.get('talent_promo_job_profile')),
                'talent_promo_business_need': _get_str(row, col_indices.get('talent_promo_business_need')),
                'talent_promo_role_scope': _get_str(row, col_indices.get('talent_promo_role_scope')),
                'talent_promo_readiness': _get_str(row, col_indices.get('talent_promo_readiness')),

                # Metadata
                'talent_calibration_status': _get_str(row, col_indices.get('talent_calibration_status')),
            }

            employees.append(emp)

        wb.close()
        return True, employees, ''

    except Exception as e:
        return False, [], str(e)


def parse_proposed_actions_metadata(proposed_actions: str, tenets_config: dict) -> dict:
    """
    Parse tenets and mentor/mentees from Proposed Actions field.

    Supports two formats:

    1. New split format (from our tool's export):
        === WORKDAY CONTENT ===
        [Original text]

        === TOOL ADDITIONS ===
        Strengths: Tenet1; Tenet2
        Improvements: Tenet3
        Mentor: Name
        Mentees: Name1; Name2

        [MODIFIED] ← Update Workday with the content above

    2. Legacy format (bracket markers inline):
        [Strengths: Tenet1; Tenet2] [Improvements: Tenet3] [Mentor: Name]

    Args:
        proposed_actions: The raw Proposed Actions text
        tenets_config: Dict with 'tenets' list from tenets.json

    Returns:
        dict with keys:
            - clean_actions: Proposed Actions without metadata markers (Workday content only)
            - strength_ids: List of tenet IDs matching strengths
            - improvement_ids: List of tenet IDs matching improvements
            - mentor: Mentor name string (or None)
            - mentees: Mentees string (or None)
            - is_modified: True if [MODIFIED] marker present (new format only)
            - has_new_format: True if using new split format
    """
    import re

    if not proposed_actions:
        return {
            'clean_actions': '',
            'strength_ids': [],
            'improvement_ids': [],
            'mentor': None,
            'mentees': None,
            'is_modified': False,
            'has_new_format': False
        }

    # Build name -> id mapping
    name_to_id = {}
    if tenets_config and 'tenets' in tenets_config:
        for tenet in tenets_config['tenets']:
            name_to_id[tenet['name'].lower()] = tenet['id']

    strength_ids = []
    improvement_ids = []
    mentor = None
    mentees = None

    # Check for new split format
    if '=== WORKDAY CONTENT ===' in proposed_actions:
        is_modified = '[MODIFIED]' in proposed_actions

        # Extract WORKDAY CONTENT section
        workday_match = re.search(
            r'=== WORKDAY CONTENT ===\s*(.*?)(?:=== TOOL ADDITIONS ===|\[MODIFIED\]|$)',
            proposed_actions,
            re.DOTALL
        )
        workday_content = workday_match.group(1).strip() if workday_match else ''

        # Extract TOOL ADDITIONS section
        tool_match = re.search(
            r'=== TOOL ADDITIONS ===\s*(.*?)(?:\[MODIFIED\]|$)',
            proposed_actions,
            re.DOTALL
        )
        tool_additions = tool_match.group(1).strip() if tool_match else ''

        # Parse human-readable tool additions (line-based format)
        if tool_additions:
            for line in tool_additions.split('\n'):
                line = line.strip()
                if line.lower().startswith('strengths:'):
                    content = line[len('strengths:'):].strip()
                    names = [n.strip() for n in content.split(';')] if content else []
                    for name in names:
                        tenet_id = name_to_id.get(name.lower())
                        if tenet_id:
                            strength_ids.append(tenet_id)
                elif line.lower().startswith('improvements:'):
                    content = line[len('improvements:'):].strip()
                    names = [n.strip() for n in content.split(';')] if content else []
                    for name in names:
                        tenet_id = name_to_id.get(name.lower())
                        if tenet_id:
                            improvement_ids.append(tenet_id)
                elif line.lower().startswith('mentor:'):
                    mentor = line[len('mentor:'):].strip() or None
                elif line.lower().startswith('mentees:'):
                    mentees = line[len('mentees:'):].strip() or None

        return {
            'clean_actions': workday_content,
            'strength_ids': strength_ids,
            'improvement_ids': improvement_ids,
            'mentor': mentor,
            'mentees': mentees,
            'is_modified': is_modified,
            'has_new_format': True
        }

    # Legacy format: bracket markers inline
    clean_text = proposed_actions

    # Parse [Strengths: name1; name2] - uses semicolon since tenet names may contain commas
    strengths_match = re.search(r'\[Strengths:\s*([^\]]+)\]', proposed_actions, re.IGNORECASE)
    if strengths_match:
        # Split by semicolon first, fall back to comma for backwards compatibility
        content = strengths_match.group(1)
        names = [n.strip() for n in content.split(';')] if ';' in content else [n.strip() for n in content.split(',')]
        for name in names:
            tenet_id = name_to_id.get(name.lower())
            if tenet_id:
                strength_ids.append(tenet_id)
        clean_text = clean_text.replace(strengths_match.group(0), '')

    # Parse [Improvements: name1; name2]
    improvements_match = re.search(r'\[Improvements:\s*([^\]]+)\]', proposed_actions, re.IGNORECASE)
    if improvements_match:
        content = improvements_match.group(1)
        names = [n.strip() for n in content.split(';')] if ';' in content else [n.strip() for n in content.split(',')]
        for name in names:
            tenet_id = name_to_id.get(name.lower())
            if tenet_id:
                improvement_ids.append(tenet_id)
        clean_text = clean_text.replace(improvements_match.group(0), '')

    # Parse [Mentor: Name]
    mentor_match = re.search(r'\[Mentor:\s*([^\]]+)\]', clean_text, re.IGNORECASE)
    if mentor_match:
        mentor = mentor_match.group(1).strip()
        clean_text = clean_text.replace(mentor_match.group(0), '')

    # Parse [Mentees: Name1; Name2]
    mentees_match = re.search(r'\[Mentees:\s*([^\]]+)\]', clean_text, re.IGNORECASE)
    if mentees_match:
        mentees = mentees_match.group(1).strip()
        clean_text = clean_text.replace(mentees_match.group(0), '')

    # Clean up extra whitespace
    clean_text = re.sub(r'\n{3,}', '\n\n', clean_text.strip())

    return {
        'clean_actions': clean_text,
        'strength_ids': strength_ids,
        'improvement_ids': improvement_ids,
        'mentor': mentor if mentor else None,
        'mentees': mentees if mentees else None,
        'is_modified': False,
        'has_new_format': False
    }


def parse_proposed_actions_tenets(proposed_actions: str, tenets_config: dict) -> tuple:
    """
    Parse tenets from Proposed Actions field (backward compatibility wrapper).

    Looks for markers in format:
        [Strengths: Tenet Name 1, Tenet Name 2]
        [Improvements: Tenet Name 3]

    Also removes the tenet markers from the text, returning clean proposed actions.

    Args:
        proposed_actions: The raw Proposed Actions text
        tenets_config: Dict with 'tenets' list from tenets.json

    Returns:
        tuple: (clean_actions, strength_ids, improvement_ids)
            - clean_actions: Proposed Actions without tenet markers
            - strength_ids: List of tenet IDs matching strengths
            - improvement_ids: List of tenet IDs matching improvements

    Note: For full metadata parsing including mentor/mentees, use parse_proposed_actions_metadata().
    """
    result = parse_proposed_actions_metadata(proposed_actions, tenets_config)
    return result['clean_actions'], result['strength_ids'], result['improvement_ids']


def parse_modified_text_field(text: str) -> dict:
    """
    Parse a text field that may have [MODIFIED] marker from our tool's export.

    Used for promotion fields (Business Need, Role Scope, Readiness) which
    are plain text with optional modification marker.

    Format:
        [Field content here]

        [MODIFIED] ← Update Workday

    Args:
        text: The field text, possibly with [MODIFIED] marker

    Returns:
        dict with keys:
            - content: The field content without the marker
            - is_modified: True if [MODIFIED] marker was present
    """
    import re

    if not text:
        return {'content': '', 'is_modified': False}

    is_modified = '[MODIFIED]' in text

    # Remove the marker and any trailing instruction text
    content = re.sub(r'\n*\[MODIFIED\].*$', '', text, flags=re.DOTALL).strip()

    return {'content': content, 'is_modified': is_modified}
