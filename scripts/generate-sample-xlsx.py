#!/usr/bin/env python3
"""
Generate sample XLSX files for the Performance Rating System.

Creates fictitious Workday employee data for testing bonus and talent workflows.

Usage:
    # Bonus spreadsheets (Workday compensation export format)
    python3 scripts/generate-sample-xlsx.py                    # Small team (12 employees)
    python3 scripts/generate-sample-xlsx.py --large            # Large org (55 employees)
    python3 scripts/generate-sample-xlsx.py --calibrated       # With ratings pre-filled
    python3 scripts/generate-sample-xlsx.py --historical       # 6 quarterly files

    # Talent spreadsheets (Workday talent calibration export format)
    python3 scripts/generate-sample-xlsx.py --talent           # Small team talent
    python3 scripts/generate-sample-xlsx.py --talent --large   # Large org talent

Output files:
    Bonus:   sample-data-small.xlsx, sample-data-large.xlsx
             sample-data-calibrated-small.xlsx, sample-data-calibrated-large.xlsx
             samples/sample-historical-*.xlsx
    Talent:  sample-data-talent-small.xlsx, sample-data-talent-large.xlsx
"""
import openpyxl
from openpyxl import Workbook
import random
import sys
import os
import argparse
from datetime import datetime, timedelta

# Add parent directory to path for imports when running as standalone script
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from notes_parser import format_notes_field
from xlsx_utils import get_current_period_name
from models import derive_overall_performance, derive_future_talent as _derive_future_talent


# =============================================================================
# SHARED EMPLOYEE DATA
# =============================================================================

def get_small_team_data():
    """
    Small team: 12 employees under single manager (Della Gate).
    Perfect for testing with a manageable dataset.

    New format (2025+) includes additional fields:
    - Direct Manager (replaces Supervisory Organization)
    - Time in Job Profile
    - Hire Date
    - Country
    - Management Level
    - Performance Review Name/Rating
    """
    manager = "Della Gate"

    # (name, job, salary, grade, bonus_pct, management_level, years_in_role)
    employees = [
        ('Paige Duty', 'Principal SRE', 180000, 'IC4', 3.75, 'IC 4', 2),
        ('Lee Latency', 'Senior Software Developer', 150000, 'IC3', 3.0, 'IC 3', 3),
        ('Mona Torr', 'Senior SRE', 145000, 'IC3', 3.0, 'IC 3', 1),
        ('Robin Rollback', 'Software Developer', 120000, 'IC2', 2.5, 'IC 2', 2),
        ('Kenny Canary', 'Software Developer', 115000, 'IC2', 2.5, 'IC 2', 1),
        ('Tracey Loggins', 'Senior SRE', 155000, 'IC3', 3.0, 'IC 3', 4),
        ('Sue Q. Ell', 'Senior Software Developer', 148000, 'IC3', 3.0, 'IC 3', 2),
        ('Jason Blob', 'Software Developer', 118000, 'IC2', 2.5, 'IC 2', 1),
        ('Al Ert', 'Principal SRE', 175000, 'IC4', 3.75, 'IC 4', 3),
        ('Addie Min', 'Senior Software Developer', 152000, 'IC3', 3.0, 'IC 3', 2),
        ('Tim Out', 'Software Developer', 110000, 'IC2', 2.5, 'IC 2', 0),
        ('Barbie Que', 'Senior SRE', 149000, 'IC3', 3.0, 'IC 3', 1),
    ]

    # Sample performance ratings from last talent cycle
    perf_ratings = [
        'High Impact Performer', 'Successful Performer', 'Successful Performer',
        'Evolving Performer', 'Successful Performer', 'Successful Performer',
        'High Impact Performer', 'Successful Performer', 'Successful Performer',
        'Successful Performer', 'Successful Performer', 'Successful Performer',
    ]

    result = []
    base_date = datetime.now()
    for i, (name, job, salary, grade, bonus_pct, mgmt_level, years_in_role) in enumerate(employees):
        total_years = years_in_role + random.randint(0, 3)
        hire_date = base_date - timedelta(days=total_years * 365 + random.randint(0, 365))
        time_in_role = f'{years_in_role} year(s), {random.randint(0, 11)} month(s)'

        result.append({
            'associate': name,
            'direct_manager': manager,
            'job_profile': job,
            'salary': salary,
            'salary_local': salary,
            'currency': 'USD',
            'grade': grade,
            'bonus_pct': bonus_pct,
            'associate_id': f'EMP{1000 + i}',
            'management_level': mgmt_level,
            'country': 'United States',
            'hire_date': hire_date,
            'time_in_job_profile': time_in_role,
            'perf_review_name': '2025-Q2 Talent Assessment & Calibration',
            'perf_review_rating': perf_ratings[i],
        })

    return result


def get_large_org_data():
    """
    Large org: 55 employees across 5 managers (50 ICs + 5 managers).
    Tests multi-manager/multi-org scenario with international employees.

    The 5 managers (Della Gate, Rhoda Map, Kay P. Eye, Agie Enda, Mai Stone)
    report to a VP (not in database).
    """
    manager_names = ['Della Gate', 'Rhoda Map', 'Kay P. Eye', 'Agie Enda', 'Mai Stone']
    director_name = 'Sam Director'

    names = [
        'Paige Duty', 'Lee Latency', 'Mona Torr', 'Robin Rollback',
        'Kenny Canary', 'Tracey Loggins', 'Sue Q. Ell', 'Jason Blob',
        'Al Ert', 'Addie Min', 'Tim Out', 'Barbie Que',
        'Terry Byte', 'Nole Pointer', 'Marge Conflict', 'Bridget Branch',
        'Cody Ryder', 'Cy Ferr', 'Phil Wall', 'Lana Wan',
        'Artie Ficial', 'Ruth Cause', 'Matt Rick', 'Cassie Cache',
        'Sue Do', 'Pat Ch', 'Devin Null', 'Justin Time',
        'Annie O\'Maly', 'Sam Box', 'Val Idation', 'Bill Ding',
        'Ty Po', 'Mike Roservices', 'Lou Pe', 'Connie Tainer',
        'Noah Node', 'Sara Ver', 'Exa M. Elle', 'Dee Ploi',
        'Ray D. O\'Button', 'Cam Elcase', 'Hashim Map', 'Ben Chmark',
        'Grace Full', 'Shel Script', 'Sal T. Hash', 'Reba Boot',
        'Stan Dup', 'Kay Eight'
    ]

    # Format: (job, salary, grade, bonus_pct) for USD
    #         (job, salary_usd, grade, bonus_pct, currency, salary_local, country) for international
    team_configs = {
        'Della Gate': [
            ('Principal Software Developer', 220000, 'IC5', 5),
            ('Staff Software Developer', 180000, 'IC4', 3.75),
            ('Staff Software Developer', 175000, 'IC4', 3.75),
            ('Senior Software Developer', 150000, 'IC3', 3),
            ('Senior Software Developer', 145000, 'IC3', 3),
            ('Software Developer', 120000, 'IC2', 2.5),
            ('Software Developer', 115000, 'IC2', 2.5),
            ('Software Developer', 118000, 'IC2', 2.5),
            ('Software Developer', 112000, 'IC2', 2.5),
            ('Senior Software Developer', 152000, 'IC3', 3),
        ],
        'Rhoda Map': [
            ('Staff Software Developer', 175000, 'IC4', 3.75),
            ('Staff Software Developer', 172000, 'IC4', 3.75),
            ('Senior Software Developer', 155000, 'IC3', 3),
            ('Senior Software Developer', 149000, 'IC3', 3),
            ('Software Developer', 110000, 'IC2', 2.5),
            ('Software Developer', 125000, 'IC2', 2.5),
            ('Senior Software Developer', 147000, 'IC3', 3),
            ('Software Developer', 122000, 'IC2', 2.5),
            ('Software Developer', 119000, 'IC2', 2.5),
            ('Software Developer', 116000, 'IC2', 2.5),
        ],
        'Kay P. Eye': [
            ('Principal Software Developer', 215000, 'IC5', 5),
            ('Staff Software Developer', 182000, 'IC4', 3.75),
            ('Staff Software Developer', 178000, 'IC4', 3.75),
            ('Senior Software Developer', 158000, 'IC3', 3),
            ('Senior Software Developer', 152000, 'IC3', 3),
            ('Software Developer', 128000, 'IC2', 2.5),
            ('Software Developer', 122000, 'IC2', 2.5),
            ('Software Developer', 118000, 'IC2', 2.5),
            ('Senior Software Developer', 155000, 'IC3', 3),
            ('Software Developer', 125000, 'IC2', 2.5),
        ],
        'Agie Enda': [
            ('Senior SRE', 132911, 'IC3', 3, 'GBP', 105000, 'United Kingdom'),
            ('SRE', 98734, 'IC2', 2.5, 'GBP', 78000, 'United Kingdom'),
            ('Principal SRE', 185000, 'IC4', 3.75),
            ('Principal SRE', 183000, 'IC4', 3.75),
            ('Senior SRE', 155000, 'IC3', 3),
            ('Senior SRE', 152000, 'IC3', 3),
            ('SRE', 125000, 'IC2', 2.5),
            ('SRE', 122000, 'IC2', 2.5),
            ('Senior SRE', 148000, 'IC3', 3),
            ('SRE', 130000, 'IC2', 2.5),
        ],
        'Mai Stone': [
            ('Principal SRE', 188000, 'IC4', 3.75),
            ('Principal SRE', 186000, 'IC4', 3.75),
            ('Senior SRE', 160000, 'IC3', 3),
            ('Senior SRE', 156000, 'IC3', 3),
            ('Senior SRE', 153000, 'IC3', 3),
            ('SRE', 128000, 'IC2', 2.5),
            ('SRE', 124000, 'IC2', 2.5),
            ('SRE', 120000, 'IC2', 2.5),
            ('SRE', 118000, 'IC2', 2.5),
            ('SRE', 115000, 'IC2', 2.5),
        ],
    }

    perf_ratings_pool = [
        'High Impact Performer', 'High Impact Performer',
        'Successful Performer', 'Successful Performer', 'Successful Performer',
        'Successful Performer', 'Successful Performer', 'Successful Performer',
        'Evolving Performer', 'Evolving Performer',
    ]

    grade_to_level = {
        'IC2': 'IC 2', 'IC3': 'IC 3', 'IC4': 'IC 4', 'IC5': 'IC 5',
        'M3': 'Manager'
    }

    result = []
    name_idx = 0
    base_date = datetime.now()

    # Add 5 managers
    manager_salaries = [210000, 205000, 215000, 208000, 212000]
    for idx, manager_name in enumerate(manager_names):
        years_tenure = random.randint(3, 8)
        hire_date = base_date - timedelta(days=years_tenure * 365 + random.randint(0, 365))
        time_in_role = f'{random.randint(1, 3)} year(s), {random.randint(0, 11)} month(s)'

        result.append({
            'associate': manager_name,
            'direct_manager': director_name,
            'job_profile': 'Engineering Manager',
            'salary': manager_salaries[idx],
            'salary_local': manager_salaries[idx],
            'currency': 'USD',
            'grade': 'M3',
            'bonus_pct': 4.5,
            'associate_id': f'MGR{100 + idx}',
            'management_level': 'Manager',
            'country': 'United States',
            'hire_date': hire_date,
            'time_in_job_profile': time_in_role,
            'perf_review_name': '2025-Q2 Talent Assessment & Calibration',
            'perf_review_rating': random.choice(['High Impact Performer', 'Successful Performer']),
        })

    # Add ICs
    for manager_name in manager_names:
        configs = team_configs[manager_name]

        for config in configs:
            if len(config) == 7:
                job, salary_usd, grade, bonus_pct, currency, salary_local, country = config
            else:
                job, salary_usd, grade, bonus_pct = config
                currency = 'USD'
                salary_local = salary_usd
                country = 'United States'

            years_in_role = random.randint(0, 4)
            total_years = years_in_role + random.randint(0, 3)
            hire_date = base_date - timedelta(days=total_years * 365 + random.randint(0, 365))
            time_in_role = f'{years_in_role} year(s), {random.randint(0, 11)} month(s)'

            result.append({
                'associate': names[name_idx],
                'direct_manager': manager_name,
                'job_profile': job,
                'salary': salary_usd,
                'salary_local': salary_local,
                'currency': currency,
                'grade': grade,
                'bonus_pct': bonus_pct,
                'associate_id': f'EMP{1000 + name_idx}',
                'management_level': grade_to_level.get(grade, 'IC 2'),
                'country': country,
                'hire_date': hire_date,
                'time_in_job_profile': time_in_role,
                'perf_review_name': '2025-Q2 Talent Assessment & Calibration',
                'perf_review_rating': random.choice(perf_ratings_pool),
            })

            name_idx += 1

    return result


# =============================================================================
# BONUS SPREADSHEET GENERATION
# =============================================================================

def create_bonus_headers(sheet, period_name=None, manager_currency='USD'):
    """Add Workday extended export headers with metadata rows (NEW FORMAT 2025+)."""
    if period_name is None:
        period_name = get_current_period_name()

    sheet.append(['RH Compensation Review Process - Bonus'])
    sheet.append(['Effective as of Date', datetime.now().strftime('%Y-%m-%d')])
    sheet.append(['Compensation Review Process - In Progress',
                  f'Compensation Review: Bonus - {period_name}'])
    sheet.append(['Compensation Review Process - Completed'])
    sheet.append(['Supervisory Organization', 'Supervisory Organization (Sample Manager)'])
    sheet.append(['Include Subordinate Organizations', 'Yes'])
    sheet.append(['Bonus Cycle Review', '', '', ''])

    headers = [
        'Associate ID',
        'Associate',
        'Job Title',
        'Time in Job Profile',
        'Hire Date',
        'Base Pay All Countries (Local)',
        f'Base Pay All Countries ({manager_currency})',
        'Grade',
        'Annual Bonus Target Percent',
        'Currency',
        'Bonus Target (Local)',
        f'Bonus Target ({manager_currency})',
        f'Last Bonus Allocation Percent (As of Report Run Date)',
        'Proposed Percent of Target Bonus',
        'Proposed Bonus Amount (Local)',
        f'Proposed Bonus Amount ({manager_currency})',
        'Direct Manager',
        'Notes',
        'Error',
        'Country',
        'Management Level',
        'Performance Review Name',
        'Overall Performance Rating (Note: From Most Recent Talent Cycle)',
    ]

    sheet.append(headers)


def write_bonus_employee_data(sheet, employees, include_notes=False):
    """Write employee data to worksheet in NEW Workday format (2025+)."""
    for emp in employees:
        if emp['currency'] == 'USD':
            base_pay_local = emp['salary']
            base_pay_converted = None
            bonus_target_local = emp['salary'] * (emp['bonus_pct'] / 100)
            bonus_target_converted = None
        else:
            base_pay_local = emp['salary_local']
            base_pay_converted = emp['salary']
            bonus_target_local = emp['salary_local'] * (emp['bonus_pct'] / 100)
            bonus_target_converted = emp['salary'] * (emp['bonus_pct'] / 100)

        last_bonus_choices = [None, None, None, 0.85, 0.90, 0.95, 1.00, 1.05, 1.10, 1.15]
        last_bonus_pct = random.choice(last_bonus_choices)
        bonus_pct_decimal = emp['bonus_pct'] / 100

        # Generate notes if calibrated
        notes = ''
        if include_notes:
            rating, justification = generate_rating_for_employee(emp)
            notes = format_notes_field(
                performance_rating=float(rating),
                justification=justification
            )

        row = [
            emp['associate_id'],
            emp['associate'],
            emp['job_profile'],
            emp.get('time_in_job_profile', ''),
            emp.get('hire_date'),
            base_pay_local,
            base_pay_converted,
            emp['grade'],
            bonus_pct_decimal,
            emp['currency'],
            bonus_target_local,
            bonus_target_converted,
            last_bonus_pct,
            None,
            None,
            None,
            emp.get('direct_manager', ''),
            notes,
            '',
            emp.get('country', 'United States'),
            emp.get('management_level', ''),
            emp.get('perf_review_name', ''),
            emp.get('perf_review_rating', ''),
        ]

        sheet.append(row)


def generate_rating_for_employee(emp):
    """Generate a performance rating and justification for an employee."""
    job = emp.get('job_profile', '')
    if 'Senior' in job or 'Lead' in job:
        rating = random.choice([95, 100, 105, 110, 115, 120, 125, 130])
    elif 'Manager' in job or 'Director' in job:
        rating = random.choice([90, 95, 100, 105, 110, 115, 120, 125, 130, 135])
    else:
        rating = random.choice([85, 90, 95, 100, 100, 105, 105, 110, 115, 120, 125, 130, 135])

    if rating >= 130:
        justifications = [
            "Exceptional performer who consistently exceeds expectations.",
            "Outstanding contributions across multiple high-impact projects.",
            "Top performer with significant business impact this cycle.",
        ]
    elif rating >= 110:
        justifications = [
            "Strong performer who regularly exceeds expectations.",
            "Solid contributions with notable achievements this cycle.",
            "High quality work with positive team impact.",
        ]
    elif rating >= 90:
        justifications = [
            "Solid performer meeting all expectations.",
            "Reliable contributor with consistent delivery.",
            "Good work quality and team collaboration.",
        ]
    else:
        justifications = [
            "Developing performer working toward expectations.",
            "Growth opportunity identified; coaching in progress.",
            "Building skills with support from team.",
        ]

    return rating, random.choice(justifications)


def create_bonus_xlsx(large=False, calibrated=False):
    """Create sample bonus XLSX file."""
    if large:
        employees = get_large_org_data()
        filename = 'sample-data-calibrated-large.xlsx' if calibrated else 'sample-data-large.xlsx'
        description = "55 employees (5 managers + 50 ICs)"
    else:
        employees = get_small_team_data()
        filename = 'sample-data-calibrated-small.xlsx' if calibrated else 'sample-data-small.xlsx'
        description = "12 employees under 1 manager (Della Gate)"

    wb = Workbook()
    sheet = wb.active
    create_bonus_headers(sheet)
    write_bonus_employee_data(sheet, employees, include_notes=calibrated)
    wb.save(filename)

    print(f"✓ Created {filename}")
    print(f"  - {description}")
    if large:
        print(f"  - Managers: Della Gate, Rhoda Map, Kay P. Eye, Agie Enda, Mai Stone")
    print(f"  - Total employees: {len(employees)}")

    us_count = sum(1 for e in employees if e['currency'] == 'USD')
    intl_count = len(employees) - us_count
    if intl_count > 0:
        print(f"  - {us_count} US-based (USD), {intl_count} international (GBP)")

    if calibrated:
        ratings = [generate_rating_for_employee(emp)[0] for emp in employees]
        high_count = sum(1 for r in ratings if r >= 120)
        solid_count = sum(1 for r in ratings if 90 <= r < 120)
        below_count = sum(1 for r in ratings if r < 90)
        print(f"  - Ratings: {high_count} high, {solid_count} solid, {below_count} below")


# =============================================================================
# HISTORICAL SPREADSHEET GENERATION
# =============================================================================

HISTORICAL_TENETS = [
    'Delete More Than You Add',
    'Leave the Campfire Cleaner',
    'Tests or It\'s a Hallucination',
    'Comments are Apologies',
    'Ship It to Learn It',
    'YAGNI (You Ain\'t Gonna Need It)',
    'Fail Fast, Fix Faster',
    'Sleep is a Feature',
    'Automate Yourself Out of a Job',
    'Treat Servers Like Cattle, Not Pets',
    'Be a Rubber Duck',
    'Blame the Process, Not the Person',
    'Strong Opinions, Loosely Held',
]

JUSTIFICATIONS = {
    'exceptional': [
        "Consistently delivered outstanding results. Led major initiatives and mentored junior team members effectively.",
        "Exceptional technical leadership and cross-team collaboration. Drove significant improvements in system reliability.",
        "Outstanding performance across all dimensions. Key contributor to critical project deliveries.",
    ],
    'exceeds': [
        "Strong performance with notable achievements in code quality and team collaboration.",
        "Exceeded expectations on key deliverables. Good balance of individual contribution and team support.",
        "Solid technical growth and increasing impact on team success.",
    ],
    'meets': [
        "Met all expectations. Reliable contributor with consistent delivery.",
        "Solid performance. Completed assigned work on time with good quality.",
        "Dependable team member who delivers consistently.",
    ],
    'needs_improvement': [
        "Needs improvement in time management and delivery consistency.",
        "Technical skills developing but requires more focus on quality.",
        "Performance below expectations. Clear development plan established.",
    ],
}


def get_historical_employee_timelines():
    """Define employee timelines for historical data generation."""
    return {
        'EMP1005': {'left_after_quarter': '2024-Q1', 'performance_pattern': 'declining'},
        'EMP1015': {'left_after_quarter': '2023-Q4', 'performance_pattern': 'steady'},
        'MGR102': {'left_after_quarter': '2024-Q2', 'performance_pattern': 'steady'},
        'EMP1048': {'joined_quarter': '2024-Q2', 'performance_pattern': 'improving'},
        'EMP1049': {'joined_quarter': '2024-Q3', 'performance_pattern': 'steady'},
        'EMP1042': {'joined_quarter': '2024-Q1', 'performance_pattern': 'variable'},
        'EMP1000': {
            'pre_promotion': ('Staff Software Developer', 'IC4', -30000),
            'promotions': {'2024-Q1': ('Principal Software Developer', 'IC5', 30000)},
            'performance_pattern': 'improving',
        },
        'EMP1003': {
            'pre_promotion': ('Senior Software Developer', 'IC3', -25000),
            'promotions': {'2024-Q3': ('Staff Software Developer', 'IC4', 25000)},
            'performance_pattern': 'steady',
        },
        'EMP1020': {
            'pre_promotion': ('Software Developer', 'IC2', -20000),
            'promotions': {'2023-Q4': ('Senior Software Developer', 'IC3', 20000)},
            'performance_pattern': 'improving',
        },
        'EMP1030': {
            'pre_promotion': ('SRE', 'IC2', -43000),
            'promotions': {
                '2023-Q4': ('Senior SRE', 'IC3', 18000),
                '2024-Q3': ('Principal SRE', 'IC4', 25000),
            },
            'performance_pattern': 'improving',
        },
        'EMP1010': {'performance_pattern': 'variable'},
        'EMP1025': {'performance_pattern': 'declining'},
    }


def quarter_to_index(quarter):
    """Convert quarter string to index (0-5)."""
    quarters = ['2023-Q3', '2023-Q4', '2024-Q1', '2024-Q2', '2024-Q3', '2024-Q4']
    return quarters.index(quarter)


def is_employee_active_in_quarter(emp_id, quarter, timelines):
    """Check if employee was active during a given quarter."""
    q_idx = quarter_to_index(quarter)
    timeline = timelines.get(emp_id, {})

    joined = timeline.get('joined_quarter')
    if joined and quarter_to_index(joined) > q_idx:
        return False

    left = timeline.get('left_after_quarter')
    if left and quarter_to_index(left) < q_idx:
        return False

    return True


def get_employee_job_for_quarter(emp, quarter, timelines):
    """Get the job title, grade, and salary for an employee in a specific quarter."""
    q_idx = quarter_to_index(quarter)
    emp_id = emp['associate_id']
    timeline = timelines.get(emp_id, {})
    promotions = timeline.get('promotions', {})
    pre_promotion = timeline.get('pre_promotion')

    job = emp['job_profile']
    grade = emp['grade']
    salary = emp['salary']
    salary_local = emp.get('salary_local', salary)

    if pre_promotion and promotions:
        pre_job, pre_grade, salary_delta = pre_promotion
        job = pre_job
        grade = pre_grade
        salary = emp['salary'] + salary_delta
        salary_local = emp.get('salary_local', emp['salary']) + salary_delta

    quarters = ['2023-Q3', '2023-Q4', '2024-Q1', '2024-Q2', '2024-Q3', '2024-Q4']
    for promo_q in quarters[:q_idx + 1]:
        if promo_q in promotions:
            new_job, new_grade, promo_salary_delta = promotions[promo_q]
            job = new_job
            grade = new_grade
            salary += promo_salary_delta
            salary_local += promo_salary_delta

    return job, grade, salary, salary_local


def get_bonus_pct_for_grade(grade):
    """Get bonus percentage for a grade."""
    return {'IC2': 2.5, 'IC3': 3.0, 'IC4': 3.75, 'IC5': 5.0, 'M3': 4.5}.get(grade, 3.0)


def generate_rating_for_pattern(pattern, quarter_index):
    """Generate a performance rating based on pattern and quarter index."""
    base_ratings = {
        'steady': [100, 100, 100, 100, 100, 100],
        'improving': [85, 90, 100, 110, 115, 125],
        'declining': [120, 115, 105, 95, 85, 75],
        'variable': [110, 85, 120, 95, 130, 100],
    }
    base = base_ratings.get(pattern, base_ratings['steady'])[quarter_index]
    return max(50, min(185, base + random.randint(-10, 10)))


def generate_justification(rating):
    """Generate appropriate justification based on rating."""
    if rating >= 130:
        return random.choice(JUSTIFICATIONS['exceptional'])
    elif rating >= 110:
        return random.choice(JUSTIFICATIONS['exceeds'])
    elif rating >= 90:
        return random.choice(JUSTIFICATIONS['meets'])
    else:
        return random.choice(JUSTIFICATIONS['needs_improvement'])


def generate_historical_notes(rating, include_full_details=True):
    """Generate the Notes field content for historical import."""
    if not include_full_details:
        choice = random.choice(['rating_only', 'partial', 'empty'])
        if choice == 'empty':
            return ''
        elif choice == 'rating_only':
            return f"[Performance Rating: {rating}%]"
        else:
            return format_notes_field(
                performance_rating=rating,
                justification=generate_justification(rating),
            )

    strengths = random.sample(HISTORICAL_TENETS, 3)
    improvements = random.sample([t for t in HISTORICAL_TENETS if t not in strengths], random.choice([2, 3]))

    mentors = ['Della Gate', 'Rhoda Map', 'Kay P. Eye', 'Agie Enda', 'Mai Stone',
               'Paige Duty', 'Lee Latency', 'Al Ert', 'Terry Byte']
    mentor = random.choice(mentors) if random.random() > 0.3 else None

    mentees_list = ['Tim Out', 'Robin Rollback', 'Kenny Canary', 'Jason Blob',
                   'Ty Po', 'Lou Pe', 'Sam Box', 'Pat Ch']
    mentees = ', '.join(random.sample(mentees_list, random.randint(0, 2))) or None

    return format_notes_field(
        performance_rating=rating,
        justification=generate_justification(rating),
        mentor=mentor,
        mentees=mentees,
        tenets_strengths='; '.join(strengths),
        tenets_improvements='; '.join(improvements),
    )


def write_historical_employee_data(sheet, employees, quarter, timelines):
    """Write employee data for a specific historical quarter."""
    q_idx = quarter_to_index(quarter)
    grade_to_level = {'IC2': 'IC 2', 'IC3': 'IC 3', 'IC4': 'IC 4', 'IC5': 'IC 5', 'M3': 'Manager'}

    for emp in employees:
        emp_id = emp['associate_id']

        if not is_employee_active_in_quarter(emp_id, quarter, timelines):
            continue

        job, grade, salary, salary_local = get_employee_job_for_quarter(emp, quarter, timelines)
        bonus_pct = get_bonus_pct_for_grade(grade)
        timeline = timelines.get(emp_id, {})
        pattern = timeline.get('performance_pattern', 'steady')
        rating = generate_rating_for_pattern(pattern, q_idx)
        include_full_details = random.random() > 0.15
        notes = generate_historical_notes(rating, include_full_details)
        bonus_allocation = (rating / 100) * random.uniform(0.95, 1.05) if notes else None

        last_bonus_choices = [None, None, 0.85, 0.90, 0.95, 1.00, 1.05, 1.10, 1.15]
        last_bonus_pct = random.choice(last_bonus_choices)

        currency = emp.get('currency', 'USD')
        country = emp.get('country', 'United States')
        if currency == 'USD':
            base_pay_local = salary
            base_pay_converted = None
            bonus_target_local = salary * (bonus_pct / 100)
            bonus_target_converted = None
        else:
            base_pay_local = salary_local
            base_pay_converted = salary
            bonus_target_local = salary_local * (bonus_pct / 100)
            bonus_target_converted = salary * (bonus_pct / 100)

        bonus_pct_decimal = bonus_pct / 100

        row = [
            emp_id,
            emp['associate'],
            job,
            emp.get('time_in_job_profile', ''),
            emp.get('hire_date'),
            base_pay_local,
            base_pay_converted,
            grade,
            bonus_pct_decimal,
            currency,
            bonus_target_local,
            bonus_target_converted,
            last_bonus_pct,
            round(bonus_allocation, 4) if bonus_allocation else None,
            None,
            None,
            emp.get('direct_manager', ''),
            notes,
            '',
            country,
            grade_to_level.get(grade, 'IC 2'),
            emp.get('perf_review_name', ''),
            emp.get('perf_review_rating', ''),
        ]

        sheet.append(row)


def create_historical_xlsx():
    """Create 6 quarterly historical spreadsheets for testing history feature."""
    quarters = ['2023-Q3', '2023-Q4', '2024-Q1', '2024-Q2', '2024-Q3', '2024-Q4']
    employees = get_large_org_data()
    timelines = get_historical_employee_timelines()
    period_names = {
        '2023-Q3': 'CY23 Q3', '2023-Q4': 'CY23 Q4',
        '2024-Q1': 'CY24 Q1', '2024-Q2': 'CY24 Q2',
        '2024-Q3': 'CY24 Q3', '2024-Q4': 'CY24 Q4',
    }

    print("Creating historical quarterly data...")
    print("=" * 60)

    for quarter in quarters:
        active_employees = [emp for emp in employees
                          if is_employee_active_in_quarter(emp['associate_id'], quarter, timelines)]

        wb = Workbook()
        sheet = wb.active
        create_bonus_headers(sheet, period_name=period_names[quarter])
        write_historical_employee_data(sheet, employees, quarter, timelines)

        filename = f'samples/sample-historical-{quarter}.xlsx'
        wb.save(filename)
        print(f"✓ Created {filename} ({len(active_employees)} employees)")

    print("=" * 60)
    print()
    print("Corner cases included:")
    print("  - EMP1005 (Robin Rollback): Left after 2024-Q1")
    print("  - EMP1015 (Bridget Branch): Left after 2023-Q4")
    print("  - MGR102 (Kay P. Eye): Manager who left after 2024-Q2")
    print("  - EMP1048 (Stan Dup): Joined in 2024-Q2")
    print("  - EMP1049 (Kay Eight): Joined in 2024-Q3")
    print("  - EMP1042 (Cam Elcase): Joined in 2024-Q1")
    print("  - EMP1000 (Paige Duty): Promoted IC4→IC5 in 2024-Q1")
    print("  - EMP1003 (Lee Latency): Promoted IC3→IC4 in 2024-Q3")
    print("  - EMP1020 (Artie Ficial): Promoted IC2→IC3 in 2023-Q4")
    print("  - EMP1030 (Noah Node): Promoted IC2→IC3 (2023-Q4) and IC3→IC4 (2024-Q3)")
    print("  - ~15% of records have incomplete Notes data")


# =============================================================================
# TALENT SPREADSHEET GENERATION
# =============================================================================

PERF_WHAT_OPTIONS = ['Surpasses Expectations', 'Meets Expectations', 'Meets Some Expectations']
PERF_HOW_OPTIONS = ['Surpasses Expectations', 'Meets Expectations', 'Meets Some Expectations', 'Does Not Meet Expectations']
AGILITY_OPTIONS = ['Always/Most of the Time', 'Sometimes']
MOVEMENT_READINESS_OPTIONS = [
    'Continue growing in current role',
    'Ready Now to be promoted in current role',
    'Ready for lateral move',
    'Ready to be promoted outside of current role',
    'Not well placed'
]


def derive_future_talent(growth: str, change: str) -> str:
    """Wrapper returning 'Yes'/'No' string for XLSX output."""
    return 'Yes' if _derive_future_talent(growth, change) else 'No'


def generate_perf_what_how():
    """Generate Performance What/How based on distribution from Spec §11.3."""
    distribution = [
        ('Surpasses Expectations', 'Surpasses Expectations', 0.10),
        ('Surpasses Expectations', 'Meets Expectations', 0.15),
        ('Surpasses Expectations', 'Meets Some Expectations', 0.02),
        ('Meets Expectations', 'Surpasses Expectations', 0.15),
        ('Meets Expectations', 'Meets Expectations', 0.40),
        ('Meets Expectations', 'Meets Some Expectations', 0.10),
        ('Meets Expectations', 'Does Not Meet Expectations', 0.02),
        ('Meets Some Expectations', 'Surpasses Expectations', 0.02),
        ('Meets Some Expectations', 'Meets Expectations', 0.03),
        ('Meets Some Expectations', 'Meets Some Expectations', 0.01),
    ]

    r = random.random()
    cumulative = 0
    for what, how, prob in distribution:
        cumulative += prob
        if r < cumulative:
            return what, how

    return 'Meets Expectations', 'Meets Expectations'


def generate_movement_readiness():
    """Generate movement readiness based on Spec §11.3 distribution."""
    r = random.random()
    if r < 0.75:
        return 'Continue growing in current role'
    elif r < 0.95:
        return 'Ready Now to be promoted in current role'
    else:
        return 'Ready for lateral move'


def generate_agility():
    """Generate growth/change agility ratings (~40% Always)."""
    return 'Always/Most of the Time' if random.random() < 0.40 else 'Sometimes'


def create_talent_headers(sheet):
    """Add Workday talent export headers."""
    for _ in range(5):
        sheet.append([])

    headers = [
        'Associate ID', 'Worker', 'Supervisory Organization', 'Job Profile',
        'Management Level', 'Job Category', 'Hire Date', 'Length of Service - Worker',
        'Time in Job Profile', 'Region - Location Based', 'Country',
        'Performance: What', 'Performance: How', 'Overall Performance Rating',
        'Last Talent Assessment Cycle: Overall Performance Rating',
        'Future Talent: Growth Agility', 'Future Talent: Change Agility',
        'Identified as Future Talent?', 'Last Talent Assessment Cycle: Identified as Future Talent?',
        'Movement Readiness', 'Last Talent Assessment Cycle: Movement Readiness',
        'Proposed Talent Actions', 'Promotions: Proposed Job Profile & Code',
        'Promotions: Business Need', 'Promotions: Expanded Role Scope',
        'Promotions: Associate Readiness', 'Calibration Status'
    ]
    sheet.append(headers)


def generate_length_of_service(hire_date):
    """Generate length of service string based on hire date."""
    if isinstance(hire_date, str):
        hire_date = datetime.strptime(hire_date, '%Y-%m-%d')
    delta = datetime.now() - hire_date
    years = delta.days // 365
    months = (delta.days % 365) // 30
    if years > 0:
        return f"{years} year{'s' if years > 1 else ''}, {months} month{'s' if months != 1 else ''}"
    else:
        return f"{months} month{'s' if months != 1 else ''}"


def write_talent_employee_data(sheet, employees):
    """Write talent calibration data to worksheet."""
    region_countries = {
        'Americas': ['United States', 'Canada'],
        'EMEA': ['United Kingdom', 'Germany'],
        'APAC': ['Australia', 'Japan']
    }

    for emp in employees:
        perf_what, perf_how = generate_perf_what_how()
        growth = generate_agility()
        change = generate_agility()
        movement = generate_movement_readiness()

        overall_perf = derive_overall_performance(perf_what, perf_how)
        future_talent = derive_future_talent(growth, change)

        # Last cycle data
        last_perf_what, last_perf_how = generate_perf_what_how()
        last_overall = derive_overall_performance(last_perf_what, last_perf_how)
        last_future_talent = derive_future_talent(generate_agility(), generate_agility())
        last_movement = generate_movement_readiness()

        # Region/country
        country = emp.get('country', 'United States')
        if country == 'United Kingdom':
            region = 'EMEA'
        elif country in ['Australia', 'Japan']:
            region = 'APAC'
        else:
            region = 'Americas'

        # Supervisory org format for talent files
        if emp.get('direct_manager'):
            sup_org = f"Supervisory Organization ({emp['direct_manager']})"
        else:
            sup_org = 'Supervisory Organization (Sample Manager)'

        hire_date = emp.get('hire_date')
        if isinstance(hire_date, datetime):
            hire_date_str = hire_date.strftime('%Y-%m-%d')
            length_of_service = generate_length_of_service(hire_date)
        else:
            hire_date_str = hire_date or ''
            length_of_service = emp.get('time_in_job_profile', '')

        row = [
            emp['associate_id'],
            emp['associate'],
            sup_org,
            emp['job_profile'],
            emp.get('management_level', 'Individual Contributor'),
            'Engineering',
            hire_date_str,
            length_of_service,
            emp.get('time_in_job_profile', ''),
            region,
            country,
            perf_what,
            perf_how,
            overall_perf,
            last_overall,
            growth,
            change,
            future_talent,
            last_future_talent,
            movement,
            last_movement,
            '',  # Proposed Talent Actions
            '',  # Promo job profile
            '',  # Promo business need
            '',  # Promo role scope
            '',  # Promo readiness
            'In Progress',  # Calibration status
        ]
        sheet.append(row)


def create_talent_xlsx(large=False):
    """Create sample talent calibration XLSX file."""
    if large:
        employees = get_large_org_data()
        filename = 'sample-data-talent-large.xlsx'
        description = "55 employees (5 managers + 50 ICs)"
    else:
        employees = get_small_team_data()
        filename = 'sample-data-talent-small.xlsx'
        description = "12 employees"

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Talent Calibration"

    create_talent_headers(sheet)
    write_talent_employee_data(sheet, employees)

    wb.save(filename)

    # Calculate distribution stats
    perf_counts = {}
    movement_counts = {}
    future_talent_count = 0

    for _ in employees:
        perf_what, perf_how = generate_perf_what_how()
        overall = derive_overall_performance(perf_what, perf_how)
        perf_counts[overall] = perf_counts.get(overall, 0) + 1

        movement = generate_movement_readiness()
        movement_counts[movement] = movement_counts.get(movement, 0) + 1

        growth = generate_agility()
        change = generate_agility()
        if derive_future_talent(growth, change) == 'Yes':
            future_talent_count += 1

    print(f"✓ Created {filename}")
    print(f"  - {description}")
    print(f"\nDistribution stats:")
    print(f"  Overall Performance:")
    for perf, count in sorted(perf_counts.items()):
        print(f"    {perf}: {count} ({100*count/len(employees):.1f}%)")
    print(f"  Future Talent: ~{100*future_talent_count/len(employees):.0f}%")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main entry point with argument parsing."""
    parser = argparse.ArgumentParser(
        description='Generate sample XLSX files for the Performance Rating System.',
        epilog='''
Examples:
  %(prog)s                      # Small bonus spreadsheet (12 employees)
  %(prog)s --large              # Large bonus spreadsheet (55 employees)
  %(prog)s --calibrated         # Bonus with ratings pre-filled
  %(prog)s --historical         # 6 quarterly historical files
  %(prog)s --talent             # Small talent spreadsheet
  %(prog)s --talent --large     # Large talent spreadsheet
        ''',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument('--large', action='store_true',
                        help='Generate large org (55 employees) instead of small team (12)')
    parser.add_argument('--talent', action='store_true',
                        help='Generate talent calibration spreadsheet instead of bonus')
    parser.add_argument('--historical', action='store_true',
                        help='Generate 6 quarterly historical bonus spreadsheets')
    parser.add_argument('--calibrated', action='store_true',
                        help='Pre-fill ratings in Notes column (bonus only)')

    args = parser.parse_args()

    if args.talent:
        if args.historical or args.calibrated:
            parser.error('--talent cannot be combined with --historical or --calibrated')
        create_talent_xlsx(large=args.large)
    elif args.historical:
        if args.calibrated:
            parser.error('--historical cannot be combined with --calibrated')
        create_historical_xlsx()
    else:
        create_bonus_xlsx(large=args.large, calibrated=args.calibrated)


if __name__ == '__main__':
    main()
