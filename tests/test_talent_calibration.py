"""
Tests for talent calibration features.

Covers:
- Derivation functions (derive_overall_performance, derive_future_talent)
- Spreadsheet type detection
- Talent column mapping and parsing
"""
import pytest
from datetime import datetime
import tempfile
import os

from models import derive_overall_performance, derive_future_talent
from xlsx_utils import (
    detect_spreadsheet_type,
    find_talent_column_indices,
    TALENT_COLUMN_MAP,
    TALENT_MARKERS,
    BONUS_MARKERS,
)


class TestDeriveOverallPerformance:
    """Tests for derive_overall_performance() per Spec §4.1."""

    def test_surpasses_surpasses_yields_high_impact(self):
        """Surpasses + Surpasses = High Impact Performer."""
        assert derive_overall_performance(
            "Surpasses Expectations", "Surpasses Expectations"
        ) == "High Impact Performer"

    def test_surpasses_meets_yields_successful(self):
        """Surpasses + Meets = Successful Performer."""
        assert derive_overall_performance(
            "Surpasses Expectations", "Meets Expectations"
        ) == "Successful Performer"

    def test_meets_meets_yields_successful(self):
        """Meets + Meets = Successful Performer."""
        assert derive_overall_performance(
            "Meets Expectations", "Meets Expectations"
        ) == "Successful Performer"

    def test_meets_surpasses_yields_successful(self):
        """Meets + Surpasses = Successful Performer."""
        assert derive_overall_performance(
            "Meets Expectations", "Surpasses Expectations"
        ) == "Successful Performer"

    def test_surpasses_some_yields_successful(self):
        """Surpasses + Meets Some = Successful Performer."""
        assert derive_overall_performance(
            "Surpasses Expectations", "Meets Some Expectations"
        ) == "Successful Performer"

    def test_meets_some_yields_evolving(self):
        """Meets + Meets Some = Evolving Performer."""
        assert derive_overall_performance(
            "Meets Expectations", "Meets Some Expectations"
        ) == "Evolving Performer"

    def test_some_meets_yields_evolving(self):
        """Meets Some + Meets = Evolving Performer."""
        assert derive_overall_performance(
            "Meets Some Expectations", "Meets Expectations"
        ) == "Evolving Performer"

    def test_some_surpasses_yields_successful(self):
        """Meets Some + Surpasses = Successful Performer (Surpasses carries)."""
        assert derive_overall_performance(
            "Meets Some Expectations", "Surpasses Expectations"
        ) == "Successful Performer"

    def test_some_some_yields_evolving(self):
        """Meets Some + Meets Some = Evolving Performer."""
        assert derive_overall_performance(
            "Meets Some Expectations", "Meets Some Expectations"
        ) == "Evolving Performer"

    def test_does_not_meet_with_low_rating_yields_low(self):
        """Does Not Meet + Meets Some or worse = Low Performer."""
        # Does Not Meet + Does Not Meet = Low
        assert derive_overall_performance(
            "Does Not Meet Expectations", "Does Not Meet Expectations"
        ) == "Low Performer"
        # Does Not Meet + Meets Some = Low (either order)
        assert derive_overall_performance(
            "Meets Some Expectations", "Does Not Meet Expectations"
        ) == "Low Performer"
        assert derive_overall_performance(
            "Does Not Meet Expectations", "Meets Some Expectations"
        ) == "Low Performer"

    def test_does_not_meet_with_good_rating_yields_evolving(self):
        """Does Not Meet + Meets or Surpasses = Evolving Performer."""
        # Does Not Meet + Meets = Evolving (either order)
        assert derive_overall_performance(
            "Meets Expectations", "Does Not Meet Expectations"
        ) == "Evolving Performer"
        assert derive_overall_performance(
            "Does Not Meet Expectations", "Meets Expectations"
        ) == "Evolving Performer"
        # Does Not Meet + Surpasses = Evolving (either order)
        assert derive_overall_performance(
            "Surpasses Expectations", "Does Not Meet Expectations"
        ) == "Evolving Performer"
        assert derive_overall_performance(
            "Does Not Meet Expectations", "Surpasses Expectations"
        ) == "Evolving Performer"

    def test_null_what_returns_none(self):
        """None or empty What returns None."""
        assert derive_overall_performance(None, "Meets Expectations") is None
        assert derive_overall_performance("", "Meets Expectations") is None

    def test_null_how_returns_none(self):
        """None or empty How returns None."""
        assert derive_overall_performance("Meets Expectations", None) is None
        assert derive_overall_performance("Meets Expectations", "") is None

    def test_both_null_returns_none(self):
        """Both None/empty returns None."""
        assert derive_overall_performance(None, None) is None
        assert derive_overall_performance("", "") is None

    def test_case_insensitive(self):
        """Function handles case variations."""
        assert derive_overall_performance(
            "SURPASSES EXPECTATIONS", "SURPASSES EXPECTATIONS"
        ) == "High Impact Performer"
        assert derive_overall_performance(
            "meets expectations", "meets some expectations"
        ) == "Evolving Performer"


class TestDeriveFutureTalent:
    """Tests for derive_future_talent() per Spec §4.2."""

    def test_both_always_returns_true(self):
        """Both agility = Always/Most of the Time returns True."""
        assert derive_future_talent(
            "Always/Most of the Time", "Always/Most of the Time"
        ) is True

    def test_growth_always_change_sometimes_returns_false(self):
        """Growth Always + Change Sometimes returns False."""
        assert derive_future_talent(
            "Always/Most of the Time", "Sometimes"
        ) is False

    def test_growth_sometimes_change_always_returns_false(self):
        """Growth Sometimes + Change Always returns False."""
        assert derive_future_talent(
            "Sometimes", "Always/Most of the Time"
        ) is False

    def test_both_sometimes_returns_false(self):
        """Both Sometimes returns False."""
        assert derive_future_talent("Sometimes", "Sometimes") is False

    def test_null_growth_returns_false(self):
        """None/empty growth returns False."""
        assert derive_future_talent(None, "Always/Most of the Time") is False
        assert derive_future_talent("", "Always/Most of the Time") is False

    def test_null_change_returns_false(self):
        """None/empty change returns False."""
        assert derive_future_talent("Always/Most of the Time", None) is False
        assert derive_future_talent("Always/Most of the Time", "") is False

    def test_both_null_returns_false(self):
        """Both None/empty returns False (not None)."""
        assert derive_future_talent(None, None) is False
        assert derive_future_talent("", "") is False

    def test_case_insensitive(self):
        """Function handles case variations."""
        assert derive_future_talent(
            "ALWAYS/MOST OF THE TIME", "always/most of the time"
        ) is True


class TestDetectSpreadsheetType:
    """Tests for detect_spreadsheet_type()."""

    def test_bonus_headers_detected_as_bonus(self):
        """Headers with bonus markers detected as bonus."""
        headers = [
            "Associate", "Associate ID", "Bonus Target - Local Currency",
            "Annual Bonus Target Percent", "Current Base Pay", "Proposed Bonus Amount"
        ]
        assert detect_spreadsheet_type(headers) == "bonus"

    def test_talent_headers_detected_as_talent(self):
        """Headers with talent markers detected as talent."""
        headers = [
            "Associate", "Associate ID", "Performance: What", "Performance: How",
            "Future Talent: Growth Agility", "Movement Readiness"
        ]
        assert detect_spreadsheet_type(headers) == "talent"

    def test_mixed_headers_uses_higher_score(self):
        """When both markers present, uses higher score."""
        # More talent markers
        headers_talent = [
            "Associate", "Associate ID", "Performance: What", "Performance: How",
            "Future Talent: Growth Agility", "Movement Readiness", "Bonus Target"
        ]
        assert detect_spreadsheet_type(headers_talent) == "talent"

        # More bonus markers
        headers_bonus = [
            "Associate", "Associate ID", "Performance: What",
            "Bonus Target", "Annual Bonus Target Percent", "Current Base Pay"
        ]
        assert detect_spreadsheet_type(headers_bonus) == "bonus"

    def test_empty_headers_defaults_to_bonus(self):
        """Empty/no markers defaults to bonus."""
        assert detect_spreadsheet_type([]) == "bonus"
        assert detect_spreadsheet_type(["Associate", "Associate ID"]) == "bonus"


class TestTalentColumnMapping:
    """Tests for TALENT_COLUMN_MAP and find_talent_column_indices()."""

    def test_column_map_has_required_fields(self):
        """TALENT_COLUMN_MAP includes all required talent fields."""
        # Required fields per spec
        required_fields = [
            'associate_id', 'associate',
            'talent_perf_what', 'talent_perf_how',
            'talent_growth_agility', 'talent_change_agility',
            'talent_movement_readiness',
        ]
        mapped_fields = set(TALENT_COLUMN_MAP.values())
        for field in required_fields:
            assert field in mapped_fields, f"Missing required field: {field}"

    def test_find_indices_exact_match(self):
        """find_talent_column_indices matches exact column names."""
        headers = [
            "Associate ID", "Worker", "Performance: What", "Performance: How",
            "Movement Readiness", "Future Talent: Growth Agility"
        ]
        indices = find_talent_column_indices(headers)

        assert indices['associate_id'] == 0
        assert indices['associate'] == 1
        assert indices['talent_perf_what'] == 2
        assert indices['talent_perf_how'] == 3
        assert indices['talent_movement_readiness'] == 4
        assert indices['talent_growth_agility'] == 5

    def test_find_indices_case_insensitive_fallback(self):
        """find_talent_column_indices falls back to case-insensitive match."""
        headers = [
            "associate id", "WORKER", "performance: what", "PERFORMANCE: HOW"
        ]
        indices = find_talent_column_indices(headers)

        assert indices['associate_id'] == 0
        assert indices['associate'] == 1
        assert indices['talent_perf_what'] == 2
        assert indices['talent_perf_how'] == 3

    def test_find_indices_missing_columns_are_none(self):
        """Missing columns return None, not raise error."""
        headers = ["Associate ID", "Worker"]
        indices = find_talent_column_indices(headers)

        assert indices['associate_id'] == 0
        assert indices['associate'] == 1
        assert indices.get('talent_perf_what') is None
        assert indices.get('talent_movement_readiness') is None

    def test_worker_and_associate_both_map_to_associate(self):
        """Both 'Worker' and 'Associate' columns map to 'associate' field."""
        # Worker column
        headers_worker = ["Associate ID", "Worker"]
        indices_worker = find_talent_column_indices(headers_worker)
        assert indices_worker['associate'] == 1

        # Associate column
        headers_assoc = ["Associate ID", "Associate"]
        indices_assoc = find_talent_column_indices(headers_assoc)
        assert indices_assoc['associate'] == 1


class TestParseProposedActionsTenets:
    """Tests for parsing tenets from Proposed Actions field."""

    def test_parse_strengths_only(self):
        """Parse strengths from Proposed Actions."""
        from xlsx_utils import parse_proposed_actions_tenets

        tenets_config = {
            'tenets': [
                {'id': 'tenet_1', 'name': 'Delete More Than You Add'},
                {'id': 'tenet_2', 'name': 'Ship It to Learn It'},
            ]
        }
        # Uses semicolon separator (tenet names may contain commas)
        text = "Focus on collaboration.\n\n[Strengths: Delete More Than You Add; Ship It to Learn It]"

        clean, strengths, improvements = parse_proposed_actions_tenets(text, tenets_config)

        assert clean == "Focus on collaboration."
        assert strengths == ['tenet_1', 'tenet_2']
        assert improvements == []

    def test_parse_improvements_only(self):
        """Parse improvements from Proposed Actions."""
        from xlsx_utils import parse_proposed_actions_tenets

        tenets_config = {
            'tenets': [
                {'id': 'tenet_1', 'name': 'Sleep is a Feature'},
            ]
        }
        text = "[Improvements: Sleep is a Feature]"

        clean, strengths, improvements = parse_proposed_actions_tenets(text, tenets_config)

        assert clean == ""
        assert strengths == []
        assert improvements == ['tenet_1']

    def test_parse_both_strengths_and_improvements(self):
        """Parse both strengths and improvements."""
        from xlsx_utils import parse_proposed_actions_tenets

        tenets_config = {
            'tenets': [
                {'id': 'delete_more', 'name': 'Delete More Than You Add'},
                {'id': 'sleep_feature', 'name': 'Sleep is a Feature'},
            ]
        }
        # Uses semicolon separator as produced by export
        text = "Good work!\n\n[Strengths: Delete More Than You Add] [Improvements: Sleep is a Feature]"

        clean, strengths, improvements = parse_proposed_actions_tenets(text, tenets_config)

        assert clean == "Good work!"
        assert strengths == ['delete_more']
        assert improvements == ['sleep_feature']

    def test_empty_input_returns_empty(self):
        """Empty input returns empty results."""
        from xlsx_utils import parse_proposed_actions_tenets

        clean, strengths, improvements = parse_proposed_actions_tenets('', {})

        assert clean == ''
        assert strengths == []
        assert improvements == []

    def test_no_markers_returns_original_text(self):
        """Text without markers is returned unchanged."""
        from xlsx_utils import parse_proposed_actions_tenets

        text = "Just regular proposed actions text."
        clean, strengths, improvements = parse_proposed_actions_tenets(text, {'tenets': []})

        assert clean == text
        assert strengths == []
        assert improvements == []


class TestTalentMarkers:
    """Tests for marker constants."""

    def test_talent_markers_match_spec(self):
        """TALENT_MARKERS match Spec §5.1 markers."""
        expected = ['Performance: What', 'Performance: How', 'Future Talent', 'Movement Readiness']
        assert TALENT_MARKERS == expected

    def test_bonus_markers_match_spec(self):
        """BONUS_MARKERS match Spec §5.1 markers."""
        expected = ['Bonus Target', 'Annual Bonus Target Percent', 'Current Base Pay', 'Proposed Bonus Amount']
        assert BONUS_MARKERS == expected


class TestCrossCycleAlignment:
    """Tests for cross-cycle alignment per Spec §7.4."""

    def test_high_impact_aligned(self):
        """High Impact Performer with 120-200% is aligned."""
        from models import get_cross_cycle_alignment
        assert get_cross_cycle_alignment(120, "High Impact Performer") == "aligned"
        assert get_cross_cycle_alignment(150, "High Impact Performer") == "aligned"
        assert get_cross_cycle_alignment(200, "High Impact Performer") == "aligned"

    def test_high_impact_review(self):
        """High Impact Performer with <120% needs review."""
        from models import get_cross_cycle_alignment
        assert get_cross_cycle_alignment(119, "High Impact Performer") == "review"
        assert get_cross_cycle_alignment(100, "High Impact Performer") == "review"

    def test_successful_aligned(self):
        """Successful Performer with 90-119% is aligned."""
        from models import get_cross_cycle_alignment
        assert get_cross_cycle_alignment(90, "Successful Performer") == "aligned"
        assert get_cross_cycle_alignment(100, "Successful Performer") == "aligned"
        assert get_cross_cycle_alignment(119, "Successful Performer") == "aligned"

    def test_successful_review(self):
        """Successful Performer outside 90-119% needs review."""
        from models import get_cross_cycle_alignment
        assert get_cross_cycle_alignment(89, "Successful Performer") == "review"
        assert get_cross_cycle_alignment(120, "Successful Performer") == "review"

    def test_evolving_aligned(self):
        """Evolving Performer with 70-89% is aligned."""
        from models import get_cross_cycle_alignment
        assert get_cross_cycle_alignment(70, "Evolving Performer") == "aligned"
        assert get_cross_cycle_alignment(80, "Evolving Performer") == "aligned"
        assert get_cross_cycle_alignment(89, "Evolving Performer") == "aligned"

    def test_low_aligned(self):
        """Low Performer with 0-69% is aligned."""
        from models import get_cross_cycle_alignment
        assert get_cross_cycle_alignment(0, "Low Performer") == "aligned"
        assert get_cross_cycle_alignment(50, "Low Performer") == "aligned"
        assert get_cross_cycle_alignment(69, "Low Performer") == "aligned"

    def test_incomplete_null_bonus(self):
        """None bonus returns incomplete."""
        from models import get_cross_cycle_alignment
        assert get_cross_cycle_alignment(None, "Successful Performer") == "incomplete"

    def test_incomplete_null_talent(self):
        """None talent returns incomplete."""
        from models import get_cross_cycle_alignment
        assert get_cross_cycle_alignment(100, None) == "incomplete"

    def test_incomplete_both_null(self):
        """Both None returns incomplete."""
        from models import get_cross_cycle_alignment
        assert get_cross_cycle_alignment(None, None) == "incomplete"


class TestTalentImportEndpoints:
    """Integration tests for talent file import via REST endpoints."""

    def test_analyze_import_detects_talent_file(self, client, talent_xlsx_file):
        """Analyze endpoint correctly detects talent spreadsheet type."""
        with open(talent_xlsx_file, 'rb') as f:
            response = client.post(
                '/api/import/analyze',
                data={'file': (f, 'talent-report.xlsx')},
                content_type='multipart/form-data'
            )

        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['spreadsheet_type'] == 'talent'

    def test_analyze_talent_file_suggests_current_import(self, client, talent_xlsx_file):
        """Talent files should suggest 'current' import (not historical)."""
        with open(talent_xlsx_file, 'rb') as f:
            response = client.post(
                '/api/import/analyze',
                data={'file': (f, 'talent-report.xlsx')},
                content_type='multipart/form-data'
            )

        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        # Talent files always suggest current import (no period metadata)
        detection = data['import_detection']
        assert detection['suggested_type'] == 'current'
        assert detection['is_talent_file'] is True
        assert detection['is_current_period'] is True
        assert detection['period_display'] == 'Current Cycle'

    def test_import_talent_file_success(self, client, talent_xlsx_file):
        """Import endpoint successfully imports talent file."""
        with open(talent_xlsx_file, 'rb') as f:
            response = client.post(
                '/api/import/current',
                data={'file': (f, 'talent-report.xlsx')},
                content_type='multipart/form-data'
            )

        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['imported'] == 2  # Fixture has 2 employees

    def test_import_talent_file_populates_talent_fields(self, client, talent_xlsx_file, db_session):
        """Import correctly populates talent-specific fields."""
        from models import Employee

        with open(talent_xlsx_file, 'rb') as f:
            response = client.post(
                '/api/import/current',
                data={'file': (f, 'talent-report.xlsx')},
                content_type='multipart/form-data'
            )

        assert response.status_code == 200

        # Verify talent fields were populated
        employee = db_session.query(Employee).first()
        assert employee is not None
        # At least one talent field should be set
        assert any([
            employee.talent_perf_what,
            employee.talent_perf_how,
            employee.talent_growth_agility,
            employee.talent_change_agility,
        ])

    def test_manager_talent_inputs_preserved_on_reimport(self, client, talent_xlsx_file, db_session):
        """Manager-entered talent fields are preserved on re-import (Spec §5.3)."""
        from models import Employee

        # First import creates employees
        with open(talent_xlsx_file, 'rb') as f:
            response = client.post(
                '/api/import/current',
                data={'file': (f, 'talent-report.xlsx')},
                content_type='multipart/form-data'
            )
        assert response.status_code == 200

        # Get the imported employee and modify manager-input fields
        employee = db_session.query(Employee).filter_by(associate_id='T001').first()
        assert employee is not None

        # Simulate manager edits (these should be preserved on re-import)
        employee.talent_perf_what = 'Meets Expectations'  # Changed from fixture value
        employee.talent_perf_how = 'Surpasses Expectations'  # Changed from fixture value
        employee.talent_growth_agility = 'Sometimes'  # Changed from fixture value
        employee.talent_movement_readiness = 'Ready in 1-2 Years'  # Changed
        employee.talent_proposed_actions = 'Promote to Senior'  # New value
        employee.talent_tenets_strengths = 'Customer Obsession'  # New value
        db_session.commit()

        # Re-import the same file
        with open(talent_xlsx_file, 'rb') as f:
            response = client.post(
                '/api/import/current',
                data={'file': (f, 'talent-report.xlsx')},
                content_type='multipart/form-data'
            )
        assert response.status_code == 200
        data = response.get_json()
        assert data['updated'] == 2  # Both employees updated, not imported

        # Refresh and verify manager fields were PRESERVED
        db_session.expire_all()
        employee = db_session.query(Employee).filter_by(associate_id='T001').first()

        # Manager-entered fields should be preserved
        assert employee.talent_perf_what == 'Meets Expectations'
        assert employee.talent_perf_how == 'Surpasses Expectations'
        assert employee.talent_growth_agility == 'Sometimes'
        assert employee.talent_movement_readiness == 'Ready in 1-2 Years'
        assert employee.talent_proposed_actions == 'Promote to Senior'
        assert employee.talent_tenets_strengths == 'Customer Obsession'

    def test_import_detects_derivation_mismatch(self, client, tmp_path):
        """Import detects when Overall Performance doesn't match What/How derivation."""
        from openpyxl import Workbook

        # Create a talent file with intentional mismatch
        wb = Workbook()
        ws = wb.active
        ws.title = "Talent Calibration"

        # Headers at row 5 (matching real Workday format)
        ws['A1'] = 'Talent Report'
        headers = [
            'Associate ID', 'Worker', 'Supervisory Organization',
            'Current Job Profile', 'Performance: What', 'Performance: How',
            'Overall Performance Rating'  # This is what we'll mismatch
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=header)

        # Employee with MISMATCH: Surpasses + Meets = Successful (not High Impact)
        # We'll put "High Impact Performer" which is WRONG per our logic
        ws.cell(row=6, column=1, value='MISMATCH01')
        ws.cell(row=6, column=2, value='Mismatch Employee')
        ws.cell(row=6, column=3, value='Engineering (Manager)')
        ws.cell(row=6, column=4, value='Engineer')
        ws.cell(row=6, column=5, value='Surpasses Expectations')  # What
        ws.cell(row=6, column=6, value='Meets Expectations')       # How
        ws.cell(row=6, column=7, value='High Impact Performer')    # WRONG: should be Successful

        # Employee with MATCH: Surpasses + Surpasses = High Impact (correct)
        ws.cell(row=7, column=1, value='MATCH01')
        ws.cell(row=7, column=2, value='Match Employee')
        ws.cell(row=7, column=3, value='Engineering (Manager)')
        ws.cell(row=7, column=4, value='Senior Engineer')
        ws.cell(row=7, column=5, value='Surpasses Expectations')   # What
        ws.cell(row=7, column=6, value='Surpasses Expectations')   # How
        ws.cell(row=7, column=7, value='High Impact Performer')    # CORRECT

        # Save to temp file
        test_file = tmp_path / 'mismatch_test.xlsx'
        wb.save(test_file)

        # Import the file
        with open(test_file, 'rb') as f:
            response = client.post(
                '/api/import/current',
                data={'file': (f, 'mismatch_test.xlsx')},
                content_type='multipart/form-data'
            )

        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['imported'] == 2

        # Should detect 1 mismatch
        assert 'derivation_mismatch_count' in data
        assert data['derivation_mismatch_count'] == 1
        assert len(data['derivation_mismatches']) == 1

        # Verify mismatch details
        mismatch = data['derivation_mismatches'][0]
        assert mismatch['associate_id'] == 'MISMATCH01'
        assert mismatch['what'] == 'Surpasses Expectations'
        assert mismatch['how'] == 'Meets Expectations'
        assert mismatch['imported'] == 'High Impact Performer'
        assert mismatch['expected'] == 'Successful Performer'

    def test_import_no_mismatch_when_all_match(self, client, talent_xlsx_file):
        """Import returns no mismatch data when all derivations match."""
        # The standard talent_xlsx_file fixture doesn't have Overall Performance column
        # so there should be no mismatches (nothing to compare)
        with open(talent_xlsx_file, 'rb') as f:
            response = client.post(
                '/api/import/current',
                data={'file': (f, 'talent-report.xlsx')},
                content_type='multipart/form-data'
            )

        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        # No mismatch fields when there are no mismatches
        assert 'derivation_mismatch_count' not in data
        assert 'derivation_mismatches' not in data


class TestCalibrateAPIValidation:
    """Tests for /api/calibrate endpoint validation."""

    def test_enum_validation_case_insensitive(self, client, db_session):
        """API accepts enum values with different casing and normalizes them."""
        from models import Employee

        # Create test employee
        emp = Employee(
            associate_id='CASE001',
            associate='Case Test Employee',
            supervisory_organization='Engineering',
        )
        db_session.add(emp)
        db_session.commit()

        # Submit with different casing than canonical values
        import json
        response = client.post('/api/calibrate', json={
            'associate_id': 'CASE001',
            'talent_perf_what': 'MEETS EXPECTATIONS',  # Should be "Meets Expectations"
            'talent_perf_how': 'surpasses expectations',  # Should be "Surpasses Expectations"
            'talent_growth_agility': 'ALWAYS/MOST OF THE TIME',  # Should be "Always/Most of the Time"
        })

        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True

        # Verify values are normalized to canonical casing
        db_session.expire_all()
        employee = db_session.query(Employee).filter_by(associate_id='CASE001').first()
        assert employee.talent_perf_what == 'Meets Expectations'
        assert employee.talent_perf_how == 'Surpasses Expectations'
        assert employee.talent_growth_agility == 'Always/Most of the Time'

    def test_enum_validation_rejects_invalid_values(self, client, db_session):
        """API rejects values that don't match any valid option."""
        from models import Employee

        # Create test employee
        emp = Employee(
            associate_id='INVALID001',
            associate='Invalid Value Test',
            supervisory_organization='Engineering',
        )
        db_session.add(emp)
        db_session.commit()

        # Submit with invalid value
        response = client.post('/api/calibrate', json={
            'associate_id': 'INVALID001',
            'talent_perf_what': 'Not A Valid Option',
        })

        assert response.status_code == 400
        data = response.get_json()
        assert data['success'] is False
        assert 'Invalid value' in data['error']


class TestTenetsRoundTrip:
    """Integration test for tenets export → re-import round-trip.

    Tests that tenets embedded in the Proposed Actions field during export
    are correctly parsed back when the file is re-imported.
    """

    def test_tenets_round_trip_export_import(self, client, db_session, tmp_path, sample_tenets):
        """Export talent file with tenets, re-import, verify tenets preserved."""
        import json
        import io
        from openpyxl import load_workbook
        from models import Employee

        # Create employees with tenets
        emp = Employee(
            associate_id='RT001',
            associate='Round Trip Test',
            supervisory_organization='Engineering',
            current_job_profile='Senior Engineer',
            talent_perf_what='Meets Expectations',
            talent_perf_how='Meets Expectations',
            talent_growth_agility='Always/Most of the Time',
            talent_change_agility='Sometimes',
            talent_movement_readiness='Continue growing in current role',
            talent_proposed_actions='Focus on technical leadership',
            talent_tenets_strengths=json.dumps(['delete_more', 'campfire_cleaner']),
            talent_tenets_improvements=json.dumps(['ship_to_learn']),
        )
        db_session.add(emp)
        db_session.commit()

        # Export talent data
        response = client.get('/export/talent')
        assert response.status_code == 200
        assert 'spreadsheet' in response.content_type or 'excel' in response.content_type

        # Save exported file
        export_path = tmp_path / 'talent_export.xlsx'
        with open(export_path, 'wb') as f:
            f.write(response.data)

        # Verify exported file has tenets embedded in Proposed Actions
        wb = load_workbook(export_path)
        ws = wb.active

        # Find Proposed Talent Actions column
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        proposed_actions_idx = None
        for idx, header in enumerate(header_row):
            if header and 'Proposed' in str(header):
                proposed_actions_idx = idx
                break

        assert proposed_actions_idx is not None, "Proposed Talent Actions column not found"

        # Check data row
        data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        proposed_actions_value = data_row[proposed_actions_idx]

        # Verify tenets are embedded (format: [Strengths: ...] [Improvements: ...])
        assert proposed_actions_value is not None
        assert '[Strengths:' in proposed_actions_value
        assert '[Improvements:' in proposed_actions_value
        assert 'Focus on technical leadership' in proposed_actions_value

        # Clear existing data
        db_session.query(Employee).delete()
        db_session.commit()

        # Re-import the exported file
        with open(export_path, 'rb') as f:
            response = client.post(
                '/api/import/current',
                data={'file': (f, 'talent_export.xlsx')},
                content_type='multipart/form-data'
            )

        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True

        # Verify tenets were parsed from Proposed Actions and restored
        db_session.expire_all()
        employee = db_session.query(Employee).filter_by(associate_id='RT001').first()
        assert employee is not None

        # Proposed actions should be clean (tenets stripped out)
        assert employee.talent_proposed_actions == 'Focus on technical leadership'

        # Tenets should be restored as JSON arrays
        assert employee.talent_tenets_strengths is not None
        strengths = json.loads(employee.talent_tenets_strengths)
        assert 'delete_more' in strengths
        assert 'campfire_cleaner' in strengths

        assert employee.talent_tenets_improvements is not None
        improvements = json.loads(employee.talent_tenets_improvements)
        assert 'ship_to_learn' in improvements
