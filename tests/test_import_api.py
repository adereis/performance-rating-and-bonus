"""
Tests for the import API endpoints.
"""
import pytest
import os
import io
import tempfile
from openpyxl import Workbook
from models import Employee, Period, RatingSnapshot


def create_test_xlsx(employees_data, include_headers=True, manager_currency='USD',
                     period_name='CY25 Q3', total_pool=100000.0):
    """
    Create a test XLSX file with Workday extended format.

    The extended format includes metadata rows before the actual data:
    - Row 1: Report title with period
    - Row 2: Subtitle
    - Row 3: Budget headers
    - Row 4: Budget data (type, spend, "of", pool, %, style, currency)
    - Rows 5-8: Section headers
    - Row 9: Column headers
    - Row 10+: Employee data

    Args:
        employees_data: List of dicts with employee data
        include_headers: Whether to include header rows
        manager_currency: Manager's home currency (USD, AUD, EUR, etc.)
        period_name: Period identifier (e.g., "CY25 Q3")
        total_pool: Total bonus pool amount

    Returns:
        BytesIO object containing the XLSX file
    """
    wb = Workbook()
    ws = wb.active

    # Row 1: Report title with period
    ws.append([f'Associate Awards:: Compensation Review: Bonus - {period_name}'])

    # Row 2: Subtitle
    ws.append(['My Current Organizations Budget and Spend'])

    # Row 3: Budget headers
    ws.append(['Name', 'Total Spend Text Value', 'of', 'Total Pool Text Value',
               '% Pool Spent', 'Data Viz Color [Singular]'])

    # Row 4: Budget data
    ws.append(['Bonus', '0.00', 'of', str(total_pool), 0.0, 'style1', manager_currency, 0.0])

    # Rows 5-8: Section headers (as in real Workday export)
    ws.append(['Compensation Planning Header', 'Compensation Planning'])
    ws.append(['Process Preferences'])
    ws.append(['Organization Issues'])
    ws.append(['Associate', '', '', '', '', 'Bonus'])

    # Row 9: Column headers (matching Workday export)
    headers = [
        'Associate',
        'Supervisory Organization',
        'Current Job Profile',
        'Photo',
        'Errors',
        'Associate ID',
        'Current Base Pay All Countries',
        f'Current Base Pay All Countries ({manager_currency})',
        'Currency',
        'Grade',
        'Annual Bonus Target %',
        'Last Bonus Allocation %',
        'Bonus Target - Local Currency',
        f'Bonus Target - Local Currency ({manager_currency})',
        'Proposed Bonus Amount',
        f'Proposed Bonus Amount ({manager_currency})',
        'Proposed % of Target Bonus',
        'Notes',
        'Zero Bonus Allocated'
    ]
    ws.append(headers)

    # Row 10+: Employee data rows
    for emp in employees_data:
        row = [
            emp.get('associate', ''),
            emp.get('supervisory_organization', ''),
            emp.get('current_job_profile', ''),
            emp.get('photo', ''),
            emp.get('errors', ''),
            emp.get('associate_id', ''),
            emp.get('current_base_pay_all_countries', ''),
            emp.get('current_base_pay_manager_currency', ''),
            emp.get('currency', ''),
            emp.get('grade', ''),
            emp.get('annual_bonus_target_percent', ''),
            emp.get('last_bonus_allocation_percent', ''),
            emp.get('bonus_target_local_currency', ''),
            emp.get('bonus_target_manager_currency', ''),
            emp.get('proposed_bonus_amount', ''),
            emp.get('proposed_bonus_amount_manager_currency', ''),
            emp.get('proposed_percent_of_target_bonus', ''),
            emp.get('notes', ''),
            emp.get('zero_bonus_allocated', '')
        ]
        ws.append(row)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


class TestImportPage:
    """Tests for the import page route."""

    def test_import_page_loads(self, client, db_session):
        """Test that import page loads successfully."""
        response = client.get('/import')
        assert response.status_code == 200
        assert b'Import Workday Data' in response.data

    def test_import_page_has_upload_zone(self, client, db_session):
        """Test that import page has file upload zone."""
        response = client.get('/import')
        assert b'upload-zone' in response.data
        assert b'Drop your XLSX file here' in response.data


class TestImportAnalyze:
    """Tests for the /api/import/analyze endpoint."""

    def test_analyze_no_file(self, client, db_session):
        """Test analyze without file returns error."""
        response = client.post('/api/import/analyze')
        assert response.status_code == 400
        data = response.get_json()
        assert data['success'] is False
        assert 'No file' in data['error']

    def test_analyze_invalid_extension(self, client, db_session):
        """Test analyze with non-Excel file returns error."""
        data = {
            'file': (io.BytesIO(b'not an excel file'), 'test.txt')
        }
        response = client.post('/api/import/analyze', data=data, content_type='multipart/form-data')
        assert response.status_code == 400
        data = response.get_json()
        assert data['success'] is False
        assert 'Excel file' in data['error']

    def test_analyze_valid_xlsx(self, client, db_session):
        """Test analyze with valid XLSX returns metadata."""
        employees = [
            {
                'associate_id': 'EMP001',
                'associate': 'John Doe',
                'supervisory_organization': 'Engineering',
                'current_job_profile': 'Senior Engineer',
                'notes': 'Performance Rating: 125%\nJustification: Great work',
                'proposed_percent_of_target_bonus': 118.5
            },
            {
                'associate_id': 'EMP002',
                'associate': 'Jane Smith',
                'supervisory_organization': 'Product',
                'current_job_profile': 'Product Manager',
                'notes': '',
                'proposed_percent_of_target_bonus': 100.0
            }
        ]

        xlsx_file = create_test_xlsx(employees)

        data = {
            'file': (xlsx_file, 'test.xlsx'),
            'import_type': 'current'
        }
        response = client.post('/api/import/analyze', data=data, content_type='multipart/form-data')

        assert response.status_code == 200
        result = response.get_json()
        assert result['success'] is True
        assert result['employee_count'] == 2
        assert result['has_bonus_column'] is True
        assert result['notes_count'] == 1  # Only John has notes

    def test_analyze_historical_period_exists(self, client, db_session):
        """Test analyze detects existing period for historical import."""
        # Create existing period
        period = Period(id='2024-H1', name='First Half 2024')
        db_session.add(period)
        db_session.commit()

        employees = [
            {
                'associate_id': 'EMP001',
                'associate': 'John Doe',
                'notes': 'Performance Rating: 125%'
            }
        ]

        xlsx_file = create_test_xlsx(employees)

        data = {
            'file': (xlsx_file, 'test.xlsx'),
            'import_type': 'historical',
            'period_id': '2024-H1'
        }
        response = client.post('/api/import/analyze', data=data, content_type='multipart/form-data')

        assert response.status_code == 200
        result = response.get_json()
        assert result['success'] is True
        assert result['period_exists'] is True
        assert result['period_id'] == '2024-H1'


class TestImportCurrent:
    """Tests for the /api/import/current endpoint."""

    def test_import_current_no_file(self, client, db_session):
        """Test import current without file returns error."""
        response = client.post('/api/import/current')
        assert response.status_code == 400
        data = response.get_json()
        assert data['success'] is False

    def test_import_current_new_employees(self, client, db_session):
        """Test importing new employees."""
        employees = [
            {
                'associate_id': 'NEW001',
                'associate': 'New Person',
                'supervisory_organization': 'Engineering',
                'current_job_profile': 'Software Engineer',
                'currency': 'USD',
                'current_base_pay_manager_currency': 100000
            },
            {
                'associate_id': 'NEW002',
                'associate': 'Another Person',
                'supervisory_organization': 'Product',
                'current_job_profile': 'Product Manager',
                'currency': 'USD',
                'current_base_pay_manager_currency': 120000
            }
        ]

        xlsx_file = create_test_xlsx(employees)

        data = {
            'file': (xlsx_file, 'test.xlsx')
        }
        response = client.post('/api/import/current', data=data, content_type='multipart/form-data')

        assert response.status_code == 200
        result = response.get_json()
        assert result['success'] is True
        assert result['imported'] == 2
        assert result['updated'] == 0

        # Verify in database
        emp1 = db_session.query(Employee).filter(Employee.associate_id == 'NEW001').first()
        assert emp1 is not None
        assert emp1.associate == 'New Person'
        assert emp1.performance_rating_percent is None  # Manager fields initialized empty

    def test_import_current_updates_existing(self, client, db_session):
        """Test importing updates existing employees but preserves ratings."""
        # Create existing employee with rating
        existing = Employee(
            associate_id='EMP001',
            associate='Old Name',
            supervisory_organization='Old Org',
            performance_rating_percent=125.0,
            justification='Previous rating'
        )
        db_session.add(existing)
        db_session.commit()

        # Import with updated info
        employees = [
            {
                'associate_id': 'EMP001',
                'associate': 'New Name',
                'supervisory_organization': 'New Org',
                'current_job_profile': 'New Job'
            }
        ]

        xlsx_file = create_test_xlsx(employees)

        data = {
            'file': (xlsx_file, 'test.xlsx')
        }
        response = client.post('/api/import/current', data=data, content_type='multipart/form-data')

        assert response.status_code == 200
        result = response.get_json()
        assert result['success'] is True
        assert result['imported'] == 0
        assert result['updated'] == 1

        # Verify update
        emp = db_session.query(Employee).filter(Employee.associate_id == 'EMP001').first()
        assert emp.associate == 'New Name'
        assert emp.supervisory_organization == 'New Org'
        # Rating should be preserved
        assert emp.performance_rating_percent == 125.0
        assert emp.justification == 'Previous rating'

    def test_import_current_with_clear_existing(self, client, db_session):
        """Test importing with clear_existing removes old data."""
        # Create existing employees with ratings
        for i in range(3):
            emp = Employee(
                associate_id=f'OLD{i}',
                associate=f'Old Employee {i}',
                performance_rating_percent=100.0 + i * 10
            )
            db_session.add(emp)
        db_session.commit()

        assert db_session.query(Employee).count() == 3

        # Import new employees with clear_existing
        employees = [
            {
                'associate_id': 'NEW001',
                'associate': 'New Person',
                'supervisory_organization': 'Engineering'
            }
        ]

        xlsx_file = create_test_xlsx(employees)

        data = {
            'file': (xlsx_file, 'test.xlsx'),
            'clear_existing': 'true'
        }
        response = client.post('/api/import/current', data=data, content_type='multipart/form-data')

        assert response.status_code == 200
        result = response.get_json()
        assert result['success'] is True
        assert result['cleared'] == 3
        assert result['imported'] == 1
        assert result['updated'] == 0

        # Verify old employees are gone
        assert db_session.query(Employee).count() == 1
        assert db_session.query(Employee).filter(Employee.associate_id == 'OLD0').first() is None

        # Verify new employee exists
        new_emp = db_session.query(Employee).filter(Employee.associate_id == 'NEW001').first()
        assert new_emp is not None
        assert new_emp.associate == 'New Person'


class TestImportHistorical:
    """Tests for the /api/import/historical endpoint."""

    def test_import_historical_no_file(self, client, db_session):
        """Test import historical without file returns error."""
        response = client.post('/api/import/historical')
        assert response.status_code == 400

    def test_import_historical_missing_period_id(self, client, db_session):
        """Test import historical without period_id returns error."""
        employees = [{'associate_id': 'EMP001', 'associate': 'John'}]
        xlsx_file = create_test_xlsx(employees)

        data = {
            'file': (xlsx_file, 'test.xlsx'),
            'period_name': 'First Half 2024'
        }
        response = client.post('/api/import/historical', data=data, content_type='multipart/form-data')

        assert response.status_code == 400
        result = response.get_json()
        assert 'Period ID' in result['error']

    def test_import_historical_creates_period_and_snapshots(self, client, db_session):
        """Test importing historical data creates period and snapshots."""
        employees = [
            {
                'associate_id': 'EMP001',
                'associate': 'John Doe',
                'supervisory_organization': 'Engineering',
                'current_job_profile': 'Senior Engineer',
                'bonus_target_manager_currency': 15000,
                'proposed_percent_of_target_bonus': 118.5,
                'notes': 'Performance Rating: 125%\nJustification: Excellent work\nMentor: Alice\nStrengths: Leadership'
            },
            {
                'associate_id': 'EMP002',
                'associate': 'Jane Smith',
                'supervisory_organization': 'Product',
                'current_job_profile': 'Product Manager',
                'bonus_target_manager_currency': 12000,
                'proposed_percent_of_target_bonus': 105.0,
                'notes': ''  # No notes - partial data
            }
        ]

        xlsx_file = create_test_xlsx(employees)

        data = {
            'file': (xlsx_file, 'test.xlsx'),
            'period_id': '2024-H1',
            'period_name': 'First Half 2024'
        }
        response = client.post('/api/import/historical', data=data, content_type='multipart/form-data')

        assert response.status_code == 200
        result = response.get_json()
        assert result['success'] is True
        assert result['imported'] == 2
        assert result['updated'] == 0
        assert result['full_details'] == 1  # Only John has full details from notes

        # Verify period was created
        period = db_session.query(Period).filter(Period.id == '2024-H1').first()
        assert period is not None
        assert period.name == 'First Half 2024'
        assert period.archived_at is not None

        # Verify snapshots
        snapshots = db_session.query(RatingSnapshot).filter(
            RatingSnapshot.period_id == '2024-H1'
        ).all()
        assert len(snapshots) == 2

        # Check John's snapshot (full details)
        john_snap = db_session.query(RatingSnapshot).filter(
            RatingSnapshot.period_id == '2024-H1',
            RatingSnapshot.associate_id == 'EMP001'
        ).first()
        assert john_snap.performance_rating == 125.0
        assert john_snap.bonus_allocation == 118.5
        assert john_snap.justification == 'Excellent work'
        assert john_snap.mentors == 'Alice'
        assert john_snap.tenets_strengths == 'Leadership'
        assert john_snap.has_full_details is True
        assert john_snap.snapshot_name == 'John Doe'
        assert john_snap.snapshot_org == 'Engineering'

        # Check Jane's snapshot (partial)
        jane_snap = db_session.query(RatingSnapshot).filter(
            RatingSnapshot.period_id == '2024-H1',
            RatingSnapshot.associate_id == 'EMP002'
        ).first()
        assert jane_snap.performance_rating is None  # No rating in notes
        assert jane_snap.bonus_allocation == 105.0  # From Workday column
        assert jane_snap.has_full_details is False

    def test_import_historical_updates_existing_snapshots(self, client, db_session):
        """Test re-importing historical data updates existing snapshots."""
        # Create existing period and snapshot
        period = Period(id='2024-H1', name='Old Name')
        db_session.add(period)

        snapshot = RatingSnapshot(
            period_id='2024-H1',
            associate_id='EMP001',
            performance_rating=100.0,
            bonus_allocation=100.0,
            snapshot_name='Old Name',
            has_full_details=False
        )
        db_session.add(snapshot)
        db_session.commit()

        # Re-import with updated data
        employees = [
            {
                'associate_id': 'EMP001',
                'associate': 'New Name',
                'supervisory_organization': 'New Org',
                'proposed_percent_of_target_bonus': 125.0,
                'notes': 'Performance Rating: 130%\nJustification: Updated review'
            }
        ]

        xlsx_file = create_test_xlsx(employees)

        data = {
            'file': (xlsx_file, 'test.xlsx'),
            'period_id': '2024-H1',
            'period_name': 'Updated Name'
        }
        response = client.post('/api/import/historical', data=data, content_type='multipart/form-data')

        assert response.status_code == 200
        result = response.get_json()
        assert result['success'] is True
        assert result['imported'] == 0
        assert result['updated'] == 1

        # Verify period was updated
        period = db_session.query(Period).filter(Period.id == '2024-H1').first()
        assert period.name == 'Updated Name'

        # Verify snapshot was updated
        snapshot = db_session.query(RatingSnapshot).filter(
            RatingSnapshot.period_id == '2024-H1',
            RatingSnapshot.associate_id == 'EMP001'
        ).first()
        assert snapshot.performance_rating == 130.0
        assert snapshot.bonus_allocation == 125.0
        assert snapshot.justification == 'Updated review'
        assert snapshot.snapshot_name == 'New Name'
        assert snapshot.has_full_details is True


class TestXlsxUtils:
    """Tests for the xlsx_utils module."""

    def test_analyze_xlsx_counts_employees(self, db_session):
        """Test analyze_xlsx correctly counts employees."""
        from xlsx_utils import analyze_xlsx

        employees = [
            {'associate_id': f'EMP{i}', 'associate': f'Employee {i}'}
            for i in range(10)
        ]

        xlsx_file = create_test_xlsx(employees)

        # Save to temp file
        temp_dir = os.path.expanduser('~/tmp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, 'test_analyze.xlsx')

        with open(temp_path, 'wb') as f:
            f.write(xlsx_file.read())

        try:
            result = analyze_xlsx(temp_path)
            assert result['success'] is True
            assert result['employee_count'] == 10
        finally:
            os.remove(temp_path)

    def test_analyze_xlsx_detects_bonus_column(self, db_session):
        """Test analyze_xlsx detects bonus column presence."""
        from xlsx_utils import analyze_xlsx

        employees = [
            {
                'associate_id': 'EMP001',
                'associate': 'John',
                'proposed_percent_of_target_bonus': 115.0
            }
        ]

        xlsx_file = create_test_xlsx(employees)

        temp_dir = os.path.expanduser('~/tmp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, 'test_bonus.xlsx')

        with open(temp_path, 'wb') as f:
            f.write(xlsx_file.read())

        try:
            result = analyze_xlsx(temp_path)
            assert result['success'] is True
            assert result['has_bonus_column'] is True
        finally:
            os.remove(temp_path)

    def test_parse_xlsx_employees(self, db_session):
        """Test parse_xlsx_employees extracts all fields."""
        from xlsx_utils import parse_xlsx_employees

        employees = [
            {
                'associate_id': 'EMP001',
                'associate': 'John Doe',
                'supervisory_organization': 'Engineering',
                'current_job_profile': 'Senior Engineer',
                'currency': 'USD',
                'current_base_pay_manager_currency': 150000,
                'bonus_target_manager_currency': 22500,
                'proposed_percent_of_target_bonus': 118.5,
                'notes': 'Performance Rating: 125%'
            }
        ]

        xlsx_file = create_test_xlsx(employees)

        temp_dir = os.path.expanduser('~/tmp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, 'test_parse.xlsx')

        with open(temp_path, 'wb') as f:
            f.write(xlsx_file.read())

        try:
            success, parsed, error = parse_xlsx_employees(temp_path)

            assert success is True
            assert len(parsed) == 1

            emp = parsed[0]
            assert emp['associate_id'] == 'EMP001'
            assert emp['associate'] == 'John Doe'
            assert emp['supervisory_organization'] == 'Engineering'
            assert emp['current_job_profile'] == 'Senior Engineer'
            assert emp['currency'] == 'USD'
            assert emp['current_base_pay_manager_currency'] == 150000
            assert emp['bonus_target_manager_currency'] == 22500
            assert emp['proposed_percent_of_target_bonus'] == 118.5
            assert 'Performance Rating: 125%' in emp['notes']
        finally:
            os.remove(temp_path)


class TestInternationalManagerCurrency:
    """Tests for non-USD manager currency support (e.g., Australian managers)."""

    def test_parse_xlsx_with_aud_columns(self, client, db_session):
        """Test that XLSX with AUD columns parses correctly."""
        from xlsx_utils import parse_xlsx_employees

        # Australian manager with mixed team (AUD and NZD employees)
        employees_data = [
            {
                'associate': 'Alice Melbourne',
                'associate_id': 'AU001',
                'supervisory_organization': 'APAC Engineering',
                'current_job_profile': 'Senior Engineer',
                'current_base_pay_all_countries': 150000,  # AUD
                'current_base_pay_manager_currency': None,  # Empty for local employees
                'currency': 'AUD',
                'grade': 'IC3',
                'annual_bonus_target_percent': 15,
                'bonus_target_local_currency': 22500,  # AUD
                'bonus_target_manager_currency': None,  # Empty for AUD employees
            },
            {
                'associate': 'Bob Auckland',
                'associate_id': 'NZ001',
                'supervisory_organization': 'APAC Engineering',
                'current_job_profile': 'Software Developer',
                'current_base_pay_all_countries': 120000,  # NZD
                'current_base_pay_manager_currency': 108000,  # Converted to AUD
                'currency': 'NZD',
                'grade': 'IC2',
                'annual_bonus_target_percent': 10,
                'bonus_target_local_currency': 12000,  # NZD
                'bonus_target_manager_currency': 10800,  # Converted to AUD
            },
        ]

        # Create XLSX with AUD as manager currency
        xlsx_file = create_test_xlsx(employees_data, manager_currency='AUD')

        # Save to temp file for parsing
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
            f.write(xlsx_file.read())

        try:
            success, parsed, error = parse_xlsx_employees(temp_path)

            assert success is True, f"Parse failed: {error}"
            assert len(parsed) == 2

            # Check AUD employee (local to manager)
            aud_emp = next(e for e in parsed if e['associate_id'] == 'AU001')
            assert aud_emp['currency'] == 'AUD'
            assert aud_emp['bonus_target_local_currency'] == 22500
            assert aud_emp['bonus_target_manager_currency'] is None  # Empty for local

            # Check NZD employee (international)
            nzd_emp = next(e for e in parsed if e['associate_id'] == 'NZ001')
            assert nzd_emp['currency'] == 'NZD'
            assert nzd_emp['bonus_target_local_currency'] == 12000  # NZD
            assert nzd_emp['bonus_target_manager_currency'] == 10800  # AUD conversion
        finally:
            os.remove(temp_path)

    def test_import_with_aud_columns(self, client, db_session):
        """Test that import works correctly with AUD columns."""
        employees_data = [
            {
                'associate': 'Charlie Sydney',
                'associate_id': 'AU002',
                'supervisory_organization': 'APAC Sales',
                'current_job_profile': 'Account Executive',
                'current_base_pay_all_countries': 130000,
                'currency': 'AUD',
                'annual_bonus_target_percent': 20,
                'bonus_target_local_currency': 26000,
                'bonus_target_manager_currency': None,
            },
            {
                'associate': 'Diana Mumbai',
                'associate_id': 'IN001',
                'supervisory_organization': 'APAC Sales',
                'current_job_profile': 'Sales Rep',
                'current_base_pay_all_countries': 2500000,  # INR
                'current_base_pay_manager_currency': 45000,  # AUD
                'currency': 'INR',
                'annual_bonus_target_percent': 10,
                'bonus_target_local_currency': 250000,  # INR
                'bonus_target_manager_currency': 4500,  # AUD
            },
        ]

        xlsx_file = create_test_xlsx(employees_data, manager_currency='AUD')

        response = client.post('/api/import/current',
                              data={'file': (xlsx_file, 'au-team.xlsx')},
                              content_type='multipart/form-data')

        assert response.status_code == 200
        result = response.get_json()
        assert result['success'] is True
        assert result['imported'] == 2

        # Verify employees were imported correctly
        aud_emp = db_session.query(Employee).filter_by(associate_id='AU002').first()
        assert aud_emp is not None
        assert aud_emp.currency == 'AUD'
        assert aud_emp.bonus_target_local_currency == 26000
        assert aud_emp.bonus_target_manager_currency is None

        inr_emp = db_session.query(Employee).filter_by(associate_id='IN001').first()
        assert inr_emp is not None
        assert inr_emp.currency == 'INR'
        assert inr_emp.bonus_target_local_currency == 250000
        assert inr_emp.bonus_target_manager_currency == 4500

    def test_bonus_calculation_with_aud_manager(self, client, db_session):
        """Test that bonus calculation uses correct fallback for AUD manager."""
        # Create employees with AUD as manager currency
        # AUD employee: use bonus_target_local_currency (already in manager's currency)
        # INR employee: use bonus_target_manager_currency (converted to AUD)
        aud_emp = Employee(
            associate_id='AU003',
            associate='Emma Perth',
            supervisory_organization='APAC Ops',
            current_job_profile='Operations Manager',
            currency='AUD',
            bonus_target_local_currency=20000,  # AUD
            bonus_target_manager_currency=None,  # Empty for local
            performance_rating_percent=110,
        )
        inr_emp = Employee(
            associate_id='IN002',
            associate='Frank Delhi',
            supervisory_organization='APAC Ops',
            current_job_profile='Operations Analyst',
            currency='INR',
            bonus_target_local_currency=400000,  # INR
            bonus_target_manager_currency=7200,  # AUD conversion
            performance_rating_percent=100,
        )
        db_session.add(aud_emp)
        db_session.add(inr_emp)
        db_session.commit()

        # Calculate total bonus pool using fallback logic
        total_pool = 0
        for emp in [aud_emp, inr_emp]:
            # This is the fallback logic used in calculations
            bonus_target = emp.bonus_target_manager_currency or emp.bonus_target_local_currency
            total_pool += bonus_target

        # Expected: 20000 (AUD) + 7200 (AUD conversion) = 27200 AUD
        assert total_pool == 27200


class TestMetadataExtraction:
    """Tests for Workday metadata extraction from extended format."""

    def test_extract_period_name_from_title(self, db_session):
        """Test that period name is extracted from report title."""
        from xlsx_utils import analyze_xlsx

        employees = [{'associate_id': 'EMP001', 'associate': 'John Doe'}]
        xlsx_file = create_test_xlsx(employees, period_name='CY25 Q3', total_pool=50000)

        temp_dir = os.path.expanduser('~/tmp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, 'test_metadata.xlsx')

        with open(temp_path, 'wb') as f:
            f.write(xlsx_file.read())

        try:
            result = analyze_xlsx(temp_path)
            assert result['success'] is True
            assert 'metadata' in result
            assert result['metadata']['period_name'] == 'CY25 Q3'
        finally:
            os.remove(temp_path)

    def test_extract_total_pool(self, db_session):
        """Test that total pool is extracted from budget row."""
        from xlsx_utils import analyze_xlsx

        employees = [{'associate_id': 'EMP001', 'associate': 'John Doe'}]
        xlsx_file = create_test_xlsx(employees, total_pool=150000.50)

        temp_dir = os.path.expanduser('~/tmp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, 'test_pool.xlsx')

        with open(temp_path, 'wb') as f:
            f.write(xlsx_file.read())

        try:
            result = analyze_xlsx(temp_path)
            assert result['success'] is True
            assert result['metadata']['total_pool'] == 150000.50
        finally:
            os.remove(temp_path)

    def test_extract_currency(self, db_session):
        """Test that currency is extracted from budget row."""
        from xlsx_utils import analyze_xlsx

        employees = [{'associate_id': 'EMP001', 'associate': 'John Doe'}]
        xlsx_file = create_test_xlsx(employees, manager_currency='AUD')

        temp_dir = os.path.expanduser('~/tmp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, 'test_currency.xlsx')

        with open(temp_path, 'wb') as f:
            f.write(xlsx_file.read())

        try:
            result = analyze_xlsx(temp_path)
            assert result['success'] is True
            assert result['metadata']['currency'] == 'AUD'
        finally:
            os.remove(temp_path)


class TestValidation:
    """Tests for Workday format validation."""

    def test_rejects_wrong_format_file(self, db_session):
        """Test that files with wrong column format are rejected."""
        from xlsx_utils import analyze_xlsx

        # Create file with wrong headers
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Email', 'Department'])  # Wrong headers
        ws.append(['John Doe', 'john@example.com', 'Engineering'])

        temp_dir = os.path.expanduser('~/tmp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, 'test_wrong_format.xlsx')
        wb.save(temp_path)

        try:
            result = analyze_xlsx(temp_path)
            assert result['success'] is False
            assert 'Could not find expected Workday columns' in result['error']
        finally:
            os.remove(temp_path)

    def test_parse_rejects_wrong_format(self, db_session):
        """Test that parse_xlsx_employees rejects wrong format files."""
        from xlsx_utils import parse_xlsx_employees

        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Full Name', 'Role'])  # Wrong headers
        ws.append(['001', 'John Doe', 'Engineer'])

        temp_dir = os.path.expanduser('~/tmp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, 'test_wrong_parse.xlsx')
        wb.save(temp_path)

        try:
            success, employees, error = parse_xlsx_employees(temp_path)
            assert success is False
            assert len(employees) == 0
            assert 'Could not find expected Workday columns' in error
        finally:
            os.remove(temp_path)

    def test_shows_found_headers_in_error(self, db_session):
        """Test that error message shows what headers were found."""
        from xlsx_utils import analyze_xlsx

        wb = Workbook()
        ws = wb.active
        ws.append(['CustomField1', 'CustomField2', 'CustomField3'])
        ws.append(['Value1', 'Value2', 'Value3'])

        temp_dir = os.path.expanduser('~/tmp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, 'test_custom_headers.xlsx')
        wb.save(temp_path)

        try:
            result = analyze_xlsx(temp_path)
            assert result['success'] is False
            # Error should show what was actually found
            assert 'Found in file' in result['error']
        finally:
            os.remove(temp_path)

    def test_rejects_old_format_without_metadata(self, db_session):
        """Test that files with correct columns but no metadata are rejected."""
        from xlsx_utils import analyze_xlsx

        # Create file with correct Workday headers but no metadata rows
        wb = Workbook()
        ws = wb.active
        # Headers without metadata rows (old format)
        ws.append([
            'Associate', 'Associate ID', 'Supervisory Organization',
            'Current Job Profile', 'Currency', 'Bonus Target - Local Currency'
        ])
        ws.append([
            'John Doe', 'EMP001', 'Engineering (Manager)', 'Developer', 'USD', 10000
        ])

        temp_dir = os.path.expanduser('~/tmp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, 'test_old_format.xlsx')
        wb.save(temp_path)

        try:
            result = analyze_xlsx(temp_path)
            assert result['success'] is False
            assert 'Missing bonus pool metadata' in result['error']
            assert 'old export format' in result['error']
        finally:
            os.remove(temp_path)


class TestPeriodDetection:
    """Tests for the period detection logic in xlsx_utils."""

    def test_detect_import_type_current_quarter(self):
        """Test that current quarter is detected as current."""
        from xlsx_utils import detect_import_type, get_current_period_name

        current = get_current_period_name()
        result = detect_import_type(current, 'bonus')

        assert result['suggested_type'] == 'current'
        assert result['is_current_period'] is True
        assert result['is_talent_file'] is False

    def test_detect_import_type_previous_quarter(self):
        """Test that previous quarter is also detected as current.

        This reflects business reality: bonus processing for Q4 happens in Q1,
        so CY25 Q4 files imported in CY26 Q1 should suggest 'current'.
        """
        from xlsx_utils import detect_import_type, get_previous_period_name

        previous = get_previous_period_name()
        result = detect_import_type(previous, 'bonus')

        assert result['suggested_type'] == 'current'
        assert result['is_current_period'] is True
        assert result['is_talent_file'] is False

    def test_detect_import_type_historical_quarter(self):
        """Test that quarters older than previous are detected as historical."""
        from xlsx_utils import detect_import_type

        # CY20 Q1 should definitely be historical
        result = detect_import_type('CY20 Q1', 'bonus')

        assert result['suggested_type'] == 'historical'
        assert result['is_current_period'] is False
        assert result['period_id'] == '2020-Q1'

    def test_get_previous_period_name_q1_rollover(self):
        """Test that previous period correctly handles year rollover from Q1."""
        from xlsx_utils import get_previous_period_name
        from datetime import datetime
        from unittest.mock import patch

        # Mock: January 2026 (Q1) → previous should be Q4 2025
        with patch('xlsx_utils.datetime') as mock_datetime:
            mock_datetime.now.return_value = datetime(2026, 1, 15)
            result = get_previous_period_name()
            assert result == 'CY25 Q4'

    def test_get_previous_period_name_mid_year(self):
        """Test previous period calculation for mid-year quarters."""
        from xlsx_utils import get_previous_period_name
        from datetime import datetime
        from unittest.mock import patch

        # Mock: July 2025 (Q3) → previous should be Q2 2025
        with patch('xlsx_utils.datetime') as mock_datetime:
            mock_datetime.now.return_value = datetime(2025, 7, 15)
            result = get_previous_period_name()
            assert result == 'CY25 Q2'

    def test_detect_import_type_talent_file(self):
        """Test that talent files always suggest current import."""
        from xlsx_utils import detect_import_type

        # Even with an old period, talent files should suggest current
        result = detect_import_type('CY20 Q1', 'talent')

        assert result['suggested_type'] == 'current'
        assert result['is_current_period'] is True
        assert result['is_talent_file'] is True

    def test_detect_import_type_unknown_period(self):
        """Test handling of unknown/missing period."""
        from xlsx_utils import detect_import_type

        result = detect_import_type(None, 'bonus')

        assert result['suggested_type'] == 'historical'
        assert result['is_current_period'] is False
        assert result['period_display'] == 'Unknown'
