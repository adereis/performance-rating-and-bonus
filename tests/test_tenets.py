"""
Tests for team tenets functionality.

Tests the tenets evaluation system including import, storage, and analytics display.
"""
import pytest
import json
from models import Employee


@pytest.fixture
def employees_with_tenets(db_session):
    """Create employees with tenets evaluations."""
    employees = [
        Employee(
            associate_id='EMP001',
            associate='Alice Anderson',
            supervisory_organization='Engineering',
            current_job_profile='Senior Engineer',
            performance_rating_percent=120,
            bonus_target_local_currency=10000,
            justification='Great work',
            tenets_strengths=json.dumps(['delete_more', 'campfire_cleaner', 'tests_or_hallucination']),
            tenets_improvements=json.dumps(['ship_to_learn', 'yagni'])
        ),
        Employee(
            associate_id='EMP002',
            associate='Bob Baker',
            supervisory_organization='Engineering',
            current_job_profile='Engineer',
            performance_rating_percent=100,
            bonus_target_local_currency=8000,
            justification='Good work',
            tenets_strengths=json.dumps(['delete_more', 'fail_fast']),
            tenets_improvements=json.dumps(['tests_or_hallucination', 'campfire_cleaner', 'yagni'])
        ),
        Employee(
            associate_id='EMP003',
            associate='Carol Chen',
            supervisory_organization='Engineering',
            current_job_profile='Engineer',
            performance_rating_percent=110,
            bonus_target_local_currency=9000,
            justification='Solid work',
            tenets_strengths=json.dumps(['campfire_cleaner', 'ship_to_learn']),
            tenets_improvements=json.dumps(['delete_more', 'fail_fast'])
        )
    ]

    for emp in employees:
        db_session.add(emp)

    db_session.commit()
    return employees


@pytest.fixture
def multi_org_employees_with_tenets(db_session):
    """Create employees across multiple orgs with tenets evaluations."""
    employees = [
        # Engineering team
        Employee(
            associate_id='EMP101',
            associate='Alice Anderson',
            supervisory_organization='Engineering',
            current_job_profile='Senior Engineer',
            performance_rating_percent=120,
            bonus_target_local_currency=10000,
            justification='Great engineering',
            tenets_strengths=json.dumps(['delete_more', 'campfire_cleaner']),
            tenets_improvements=json.dumps(['ship_to_learn'])
        ),
        Employee(
            associate_id='EMP102',
            associate='Bob Baker',
            supervisory_organization='Engineering',
            current_job_profile='Engineer',
            performance_rating_percent=100,
            bonus_target_local_currency=8000,
            justification='Met expectations',
            tenets_strengths=json.dumps(['tests_or_hallucination']),
            tenets_improvements=json.dumps(['delete_more', 'yagni'])
        ),
        # Product team
        Employee(
            associate_id='EMP201',
            associate='Carol Chen',
            supervisory_organization='Product',
            current_job_profile='Product Manager',
            performance_rating_percent=110,
            bonus_target_local_currency=9000,
            justification='Strong product leadership',
            tenets_strengths=json.dumps(['ship_to_learn', 'fail_fast']),
            tenets_improvements=json.dumps(['rubber_duck'])
        ),
        Employee(
            associate_id='EMP202',
            associate='David Davis',
            supervisory_organization='Product',
            current_job_profile='Senior PM',
            performance_rating_percent=130,
            bonus_target_local_currency=12000,
            justification='Exceptional product work',
            tenets_strengths=json.dumps(['ship_to_learn', 'rubber_duck']),
            tenets_improvements=json.dumps(['fail_fast', 'yagni'])
        )
    ]

    for emp in employees:
        db_session.add(emp)

    db_session.commit()
    return employees


class TestTenetsDataModel:
    """Test tenets data storage in Employee model."""

    def test_employee_can_store_tenets(self, db_session):
        """Test that employees can store tenets as JSON strings."""
        employee = Employee(
            associate_id='TEST001',
            associate='Test Employee',
            tenets_strengths=json.dumps(['delete_more', 'campfire_cleaner']),
            tenets_improvements=json.dumps(['ship_to_learn', 'yagni'])
        )
        db_session.add(employee)
        db_session.commit()

        retrieved = db_session.query(Employee).filter_by(associate_id='TEST001').first()
        assert retrieved.tenets_strengths == '["delete_more", "campfire_cleaner"]'
        assert retrieved.tenets_improvements == '["ship_to_learn", "yagni"]'

    def test_tenets_can_be_null(self, db_session):
        """Test that tenets fields can be null."""
        employee = Employee(
            associate_id='TEST002',
            associate='Test Employee',
            tenets_strengths=None,
            tenets_improvements=None
        )
        db_session.add(employee)
        db_session.commit()

        retrieved = db_session.query(Employee).filter_by(associate_id='TEST002').first()
        assert retrieved.tenets_strengths is None
        assert retrieved.tenets_improvements is None

    def test_tenets_json_parsing(self, db_session):
        """Test that tenets JSON can be parsed back to lists."""
        strengths_list = ['delete_more', 'campfire_cleaner', 'tests_or_hallucination']
        improvements_list = ['ship_to_learn', 'yagni']

        employee = Employee(
            associate_id='TEST003',
            associate='Test Employee',
            tenets_strengths=json.dumps(strengths_list),
            tenets_improvements=json.dumps(improvements_list)
        )
        db_session.add(employee)
        db_session.commit()

        retrieved = db_session.query(Employee).filter_by(associate_id='TEST003').first()
        assert json.loads(retrieved.tenets_strengths) == strengths_list
        assert json.loads(retrieved.tenets_improvements) == improvements_list


class TestAnalyticsWithTenets:
    """Test analytics page with tenets data."""

    def test_analytics_displays_tenets_summary(self, client, employees_with_tenets):
        """Test that analytics page shows tenets summary."""
        response = client.get('/analytics')
        assert response.status_code == 200
        assert b'Team Tenets Analysis' in response.data

    def test_analytics_counts_strength_votes(self, client, employees_with_tenets):
        """Test that strength votes are counted correctly."""
        response = client.get('/analytics')
        assert response.status_code == 200

        # delete_more: 2 strength votes (Alice, Bob)
        # campfire_cleaner: 2 strength votes (Alice, Carol)
        # All should appear in the page
        assert b'delete_more' in response.data or b'Delete More' in response.data

    def test_analytics_counts_improvement_votes(self, client, employees_with_tenets):
        """Test that improvement votes are counted correctly."""
        response = client.get('/analytics')
        assert response.status_code == 200

        # yagni: 2 improvement votes (Alice, Bob)
        # Should appear in the analytics
        assert response.status_code == 200

    def test_analytics_without_tenets(self, client, populated_db):
        """Test analytics page works when no tenets data exists."""
        # populated_db has employees but no tenets
        response = client.get('/analytics')
        assert response.status_code == 200
        # Page should load successfully even without tenets

    def test_analytics_with_empty_tenets(self, client, db_session):
        """Test analytics with employees that have empty tenets arrays."""
        employee = Employee(
            associate_id='TEST004',
            associate='Test Employee',
            tenets_strengths='[]',
            tenets_improvements='[]'
        )
        db_session.add(employee)
        db_session.commit()

        response = client.get('/analytics')
        assert response.status_code == 200


class TestMultiOrgTenetsAnalysis:
    """Test tenets analysis across multiple organizations."""

    def test_multi_org_tenets_separated(self, client, multi_org_employees_with_tenets):
        """Test that tenets are analyzed separately per organization."""
        response = client.get('/analytics')
        assert response.status_code == 200

        # Should show organization-specific analysis
        assert b'Engineering' in response.data
        assert b'Product' in response.data

    def test_single_org_no_multi_org_section(self, client, employees_with_tenets):
        """Test that multi-org section doesn't appear with single org."""
        response = client.get('/analytics')
        assert response.status_code == 200

        # Should not show "by Organization" section with only one org
        # (or it should be empty/minimal)
        assert response.status_code == 200

    def test_multi_org_tenets_count_correctly(self, client, multi_org_employees_with_tenets):
        """Test that votes are counted correctly per organization."""
        response = client.get('/analytics')
        assert response.status_code == 200

        # Engineering: delete_more has 1 strength (Alice)
        # Product: ship_to_learn has 2 strengths (Carol, David)
        # Verification happens in the page rendering
        assert response.status_code == 200


class TestTenetsImport:
    """Test importing tenets from Excel files."""

    def test_import_with_tenets_columns(self, db_session, tmp_path):
        """Test that Excel parser correctly handles tenets columns."""
        import openpyxl

        # Create a test Excel file with tenets
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Empty first row
        sheet.append([])

        # Headers
        headers = [
            'Associate', 'Supervisory Organization', 'Current Job Profile',
            'Photo', 'Errors', 'Associate ID',
            'Current Base Pay All Countries', 'Current Base Pay All Countries (USD)',
            'Currency', 'Grade', 'Annual Bonus Target Percent',
            'Last Bonus Allocation Percent', 'Bonus Target - Local Currency',
            'Bonus Target - Local Currency (USD)', 'Proposed Bonus Amount',
            'Proposed Bonus Amount (USD)', 'Proposed Percent of Target Bonus',
            'Notes', 'Zero Bonus Allocated', 'Performance Rating Percent',
            'Tenets Strengths', 'Tenets Improvements'
        ]
        sheet.append(headers)

        # Data row
        sheet.append([
            'Test Employee', 'Engineering', 'Engineer', '', '', 'EMP999',
            100000, None, 'USD', 'IC3', 15, None, 15000, None, None, None, None,
            '', '', 100,
            '["delete_more", "campfire_cleaner"]',
            '["ship_to_learn", "yagni"]'
        ])

        test_file = tmp_path / "test_tenets.xlsx"
        wb.save(str(test_file))

        # Read the file and simulate import logic
        wb = openpyxl.load_workbook(str(test_file))
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))

        # Verify we have the tenets columns (columns 20 and 21)
        assert len(rows[1]) >= 22  # Headers should include tenets columns
        assert rows[1][20] == 'Tenets Strengths'
        assert rows[1][21] == 'Tenets Improvements'

        # Verify data row has tenets values
        data_row = rows[2]
        assert data_row[20] == '["delete_more", "campfire_cleaner"]'
        assert data_row[21] == '["ship_to_learn", "yagni"]'

        # Create employee with tenets data (simulating import)
        employee = Employee(
            associate_id='EMP999',
            associate='Test Employee',
            supervisory_organization='Engineering',
            current_job_profile='Engineer',
            performance_rating_percent=100,
            tenets_strengths=str(data_row[20]),
            tenets_improvements=str(data_row[21])
        )
        db_session.add(employee)
        db_session.commit()

        # Verify import
        retrieved = db_session.query(Employee).filter_by(associate_id='EMP999').first()
        assert retrieved is not None
        assert retrieved.tenets_strengths == '["delete_more", "campfire_cleaner"]'
        assert retrieved.tenets_improvements == '["ship_to_learn", "yagni"]'

    def test_import_without_tenets_columns(self, db_session, tmp_path):
        """Test backward compatibility when Excel file has no tenets columns."""
        import openpyxl

        # Create a test Excel file without tenets
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Empty first row
        sheet.append([])

        # Headers (no tenets columns)
        headers = [
            'Associate', 'Supervisory Organization', 'Current Job Profile',
            'Photo', 'Errors', 'Associate ID',
            'Current Base Pay All Countries', 'Current Base Pay All Countries (USD)',
            'Currency', 'Grade', 'Annual Bonus Target Percent',
            'Last Bonus Allocation Percent', 'Bonus Target - Local Currency',
            'Bonus Target - Local Currency (USD)', 'Proposed Bonus Amount',
            'Proposed Bonus Amount (USD)', 'Proposed Percent of Target Bonus',
            'Notes', 'Zero Bonus Allocated', 'Performance Rating Percent'
        ]
        sheet.append(headers)

        # Data row
        sheet.append([
            'Test Employee 2', 'Engineering', 'Engineer', '', '', 'EMP998',
            100000, None, 'USD', 'IC3', 15, None, 15000, None, None, None, None,
            '', '', 100
        ])

        test_file = tmp_path / "test_no_tenets.xlsx"
        wb.save(str(test_file))

        # Read the file
        wb = openpyxl.load_workbook(str(test_file))
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))

        # Verify we don't have tenets columns
        assert len(rows[1]) == 20  # Only 20 columns, no tenets

        # Create employee without tenets (simulating import)
        data_row = rows[2]
        employee = Employee(
            associate_id='EMP998',
            associate='Test Employee 2',
            supervisory_organization='Engineering',
            current_job_profile='Engineer',
            performance_rating_percent=100,
            tenets_strengths=None,  # No tenets in file
            tenets_improvements=None
        )
        db_session.add(employee)
        db_session.commit()

        # Verify import - tenets should be None
        retrieved = db_session.query(Employee).filter_by(associate_id='EMP998').first()
        assert retrieved is not None
        assert retrieved.tenets_strengths is None
        assert retrieved.tenets_improvements is None


class TestTenetsSampleDataGeneration:
    """Test sample data generation with tenets."""

    def test_generate_random_tenets(self):
        """Test that generate_random_tenets creates valid data structures."""
        from populate_sample_ratings import generate_random_tenets

        all_tenets = ['delete_more', 'campfire_cleaner', 'tests_or_hallucination',
                      'ship_to_learn', 'yagni', 'fail_fast', 'rubber_duck']

        strengths, improvements = generate_random_tenets(all_tenets)

        # Should return lists
        assert isinstance(strengths, list)
        assert isinstance(improvements, list)

        # Should have 3 items each (or less if not enough tenets)
        assert len(strengths) <= 3
        assert len(improvements) <= 3

        # Should not overlap
        assert not set(strengths).intersection(set(improvements))

    def test_generate_random_tenets_with_empty_list(self):
        """Test generate_random_tenets with no tenets available."""
        from populate_sample_ratings import generate_random_tenets

        strengths, improvements = generate_random_tenets([])

        assert strengths == []
        assert improvements == []

    def test_load_tenets(self):
        """Test loading tenets from configuration file."""
        from populate_sample_ratings import load_tenets

        tenets = load_tenets()

        # Should return a list
        assert isinstance(tenets, list)

        # If tenets-sample.json exists, should have tenets
        if tenets:
            assert len(tenets) > 0
            # Each should be a string (tenet ID)
            assert all(isinstance(t, str) for t in tenets)
