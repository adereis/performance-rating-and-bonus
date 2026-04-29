"""
Pytest configuration and fixtures for testing.
"""
import os

# CRITICAL: Set testing flag BEFORE importing app/models to prevent
# production database access. This protects against accidental writes
# if test code is imported outside of pytest (e.g., python -c "...").
os.environ['TESTING'] = 'true'

import pytest
import tempfile
from jinja2 import StrictUndefined
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from models import Base, Employee
from app import app as flask_app


def _default_bonus_cycle(target, args, kwargs):
    """Default test employees into the bonus cycle unless explicitly set."""
    if 'in_current_bonus_cycle' not in kwargs:
        target.in_current_bonus_cycle = True


event.listen(Employee, 'init', _default_bonus_cycle)


@pytest.fixture(scope='function')
def test_db():
    """Create a temporary test database for each test."""
    # Create a temporary database file
    db_fd, db_path = tempfile.mkstemp(suffix='.db')
    db_url = f'sqlite:///{db_path}'

    # Create engine and tables
    engine = create_engine(db_url, echo=False)
    Base.metadata.create_all(bind=engine)

    # Create session factory
    TestSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

    yield TestSessionLocal, db_path

    # Cleanup: dispose engine first to close all connections
    engine.dispose()
    os.close(db_fd)
    os.unlink(db_path)


@pytest.fixture(scope='function')
def db_session(test_db):
    """Provide a database session for a test."""
    SessionLocal, db_path = test_db
    session = SessionLocal()

    yield session

    session.close()


@pytest.fixture(scope='function')
def sample_employee_data():
    """Sample employee data for testing."""
    return {
        'associate_id': 'EMP001',
        'associate': 'John Doe',
        'supervisory_organization': 'Engineering',
        'current_job_profile': 'Senior Software Engineer',
        'current_base_pay_manager_currency': 120000.0,
        'currency': 'USD',
        'annual_bonus_target_percent': 15.0,
        'bonus_target_local_currency': 18000.0,  # Required for rating functionality
        'performance_rating_percent': None,
        'justification': '',
        'mentor': '',
        'mentees': '',
        # New fields from 2025 Workday format
        'country': 'United States',
        'management_level': 'IC 3',
        'last_perf_review_name': None,
        'last_perf_review_rating': None,
        # Talent data markers (required for calibration functionality)
        'talent_perf_what_original': 'Meets Expectations',
        'talent_perf_how_original': 'Meets Expectations',
    }


@pytest.fixture(scope='function')
def sample_employees():
    """Multiple sample employees for testing."""
    return [
        {
            'associate_id': 'EMP001',
            'associate': 'Alice Johnson',
            'supervisory_organization': 'Engineering',
            'current_job_profile': 'Senior Software Engineer',
            'current_base_pay_manager_currency': 130000.0,
            'currency': 'USD',
            'bonus_target_local_currency': 19500.0,  # Required for rating functionality
            'performance_rating_percent': 120.0,
            'justification': 'Excellent performance',
            'mentor': 'Bob Smith',
            'mentees': 'Charlie Brown',
            'talent_perf_what_original': 'Meets Expectations',  # Required for calibration
        },
        {
            'associate_id': 'EMP002',
            'associate': 'Bob Smith',
            'supervisory_organization': 'Engineering',
            'current_job_profile': 'Staff Software Engineer',
            'current_base_pay_manager_currency': 160000.0,
            'currency': 'USD',
            'bonus_target_local_currency': 24000.0,  # Required for rating functionality
            'performance_rating_percent': 140.0,
            'justification': 'Outstanding contributions',
            'mentor': '',
            'mentees': 'Alice Johnson, Charlie Brown',
        },
        {
            'associate_id': 'EMP003',
            'associate': 'Charlie Brown',
            'supervisory_organization': 'Engineering',
            'current_job_profile': 'Software Engineer',
            'current_base_pay_manager_currency': 95000.0,
            'currency': 'USD',
            'bonus_target_local_currency': 14250.0,  # Required for rating functionality
            'performance_rating_percent': 85.0,
            'justification': 'Good progress, needs more experience',
            'mentor': 'Alice Johnson',
            'mentees': '',
        },
        {
            'associate_id': 'EMP004',
            'associate': 'Diana Prince',
            'supervisory_organization': 'Product',
            'current_job_profile': 'Product Manager',
            'current_base_pay_manager_currency': 125000.0,
            'currency': 'USD',
            'bonus_target_local_currency': 18750.0,  # Required for rating functionality
            'performance_rating_percent': None,
            'justification': '',
            'mentor': '',
            'mentees': '',
            }
    ]


@pytest.fixture(scope='function')
def app(test_db):
    """Create Flask app configured for testing."""
    SessionLocal, db_path = test_db

    # Configure app for testing
    flask_app.config['TESTING'] = True
    flask_app.config['DATABASE_URL'] = f'sqlite:///{db_path}'

    # Catch missing template variables (e.g., forgetting to pass a var
    # to render_template). Without this, Jinja2 silently treats undefined
    # variables as falsy, hiding regressions like missing promotion_ready.
    original_undefined = flask_app.jinja_env.undefined
    flask_app.jinja_env.undefined = StrictUndefined

    # Override the get_db function to use test database
    import models
    original_get_db = models.get_db

    def test_get_db():
        return SessionLocal()

    models.get_db = test_get_db

    # Also patch in app module
    import app as app_module
    app_module.get_db = test_get_db

    yield flask_app

    # Restore original
    flask_app.jinja_env.undefined = original_undefined
    models.get_db = original_get_db
    app_module.get_db = original_get_db


@pytest.fixture(scope='function')
def client(app):
    """Create Flask test client."""
    return app.test_client()


@pytest.fixture(scope='function')
def populated_db(db_session, sample_employees):
    """Database session populated with sample employees."""
    for emp_data in sample_employees:
        employee = Employee(**emp_data)
        db_session.add(employee)

    db_session.commit()

    return db_session


@pytest.fixture(scope='function')
def populated_db_with_tenets(db_session):
    """Database with employees that have tenets data populated.

    This fixture exercises template code paths that require tenets,
    catching issues like missing template filters.

    Uses fictitious pun names from generate-sample-xlsx.py.
    """
    import json
    employees = [
        Employee(
            associate_id='EMP001',
            associate='Paige Duty',
            supervisory_organization='Engineering (Della Gate)',
            current_job_profile='Senior Software Developer',
            current_base_pay_manager_currency=150000.0,
            currency='USD',
            bonus_target_local_currency=22500.0,
            performance_rating_percent=120.0,
            justification='Excellent performance on key projects',
            tenets_strengths=json.dumps(['tenet1', 'tenet2', 'tenet3']),
            tenets_improvements=json.dumps(['tenet4', 'tenet5']),
            talent_perf_what_original='Surpasses Expectations',
            talent_perf_how_original='Meets Expectations',
        ),
        Employee(
            associate_id='EMP002',
            associate='Justin Time',
            supervisory_organization='Engineering (Della Gate)',
            current_job_profile='Software Developer',
            current_base_pay_manager_currency=115000.0,
            currency='USD',
            bonus_target_local_currency=17250.0,
            performance_rating_percent=100.0,
            justification='Solid contributor, delivers on schedule',
            tenets_strengths=json.dumps(['tenet1', 'tenet2', 'tenet3']),
            tenets_improvements=json.dumps(['tenet3', 'tenet4']),
        ),
        Employee(
            associate_id='EMP003',
            associate='Devin Null',
            supervisory_organization='Engineering (Della Gate)',
            current_job_profile='Software Developer',
            current_base_pay_manager_currency=112000.0,
            currency='USD',
            bonus_target_local_currency=16800.0,
            # No rating yet - tests incomplete state
        ),
    ]
    for emp in employees:
        db_session.add(emp)
    db_session.commit()
    return db_session


@pytest.fixture(scope='function')
def sample_tenets(app, monkeypatch):
    """
    Configure app to use sample tenets for testing.

    This fixture monkeypatches load_tenets_config to return sample tenets,
    keeping test fixtures out of production code.
    """
    import json
    sample_file = 'samples/tenets-sample.json'
    with open(sample_file, 'r') as f:
        sample_config = json.load(f)
    sample_map = {t['id']: t['name'] for t in sample_config.get('tenets', [])}

    import app as app_module
    monkeypatch.setattr(app_module, 'load_tenets_config',
                        lambda: (sample_config, sample_map))
    return sample_config


@pytest.fixture(scope='function')
def talent_xlsx_file():
    """Create a temporary talent calibration XLSX file for testing."""
    from openpyxl import Workbook

    # Create workbook with talent headers (matching real Workday export structure)
    wb = Workbook()
    ws = wb.active
    ws.title = "Talent Calibration"

    # Talent files have metadata rows before headers (like real Workday exports)
    # Headers at row 5 with 'Worker' instead of 'Associate'
    ws['A1'] = 'Talent Calibration Report'
    ws['A2'] = 'Generated: 2025-01-15'
    ws['A3'] = ''
    ws['A4'] = ''

    # Headers at row 5
    headers = [
        'Associate ID', 'Worker', 'Supervisory Organization',
        'Current Job Profile', 'Performance: What', 'Performance: How',
        'Future Talent: Growth Agility', 'Future Talent: Change Agility',
        'Movement Readiness'
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=header)

    # Sample data rows
    data = [
        ['T001', 'Test Employee One', 'Engineering (Manager One)', 'Senior Engineer',
         'Surpasses Expectations', 'Meets Expectations',
         'Always/Most of the Time', 'Always/Most of the Time', 'Ready Now'],
        ['T002', 'Test Employee Two', 'Engineering (Manager One)', 'Engineer',
         'Meets Expectations', 'Meets Expectations',
         'Sometimes', 'Always/Most of the Time', 'Ready in 1-2 Years'],
    ]

    for row_idx, row_data in enumerate(data, 6):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save to temporary file
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(path)

    yield path

    # Cleanup
    os.unlink(path)
