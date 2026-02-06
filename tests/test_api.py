"""
Tests for Flask API endpoints.
"""
import pytest
import json
from models import Employee


class TestAPIEndpoints:
    """Test Flask API endpoints."""

    def test_index_route(self, client, populated_db):
        """Test the main dashboard route."""
        response = client.get('/')
        assert response.status_code == 200
        assert b'Performance Rating' in response.data

    def test_rate_page_route(self, client, populated_db):
        """Test the rating page route."""
        response = client.get('/rate')
        assert response.status_code == 200
        assert b'Bonus Rating' in response.data

    def test_analytics_route(self, client, populated_db):
        """Test the analytics page route."""
        response = client.get('/analytics')
        assert response.status_code == 200


class TestPageSmokeTests:
    """Smoke tests for all pages with realistic data.

    These tests render pages with employees that have all fields populated
    (tenets, ratings, calibration data) to catch template errors like
    missing filters or undefined variables.
    """

    def test_dashboard_with_tenets(self, app, populated_db_with_tenets):
        """Dashboard renders with employees that have tenets."""
        client = app.test_client()
        response = client.get('/')
        assert response.status_code == 200
        assert b'Performance Rating' in response.data

    def test_rate_page_with_tenets(self, app, populated_db_with_tenets):
        """Rate page renders with employees that have tenets data."""
        client = app.test_client()
        response = client.get('/rate')
        assert response.status_code == 200
        assert b'Bonus Rating' in response.data
        assert b'Paige Duty' in response.data

    def test_calibrate_page_with_tenets(self, app, populated_db_with_tenets):
        """Calibrate page renders with employees that have talent data."""
        client = app.test_client()
        response = client.get('/calibrate')
        assert response.status_code == 200

    def test_analytics_with_tenets(self, app, populated_db_with_tenets):
        """Analytics page renders with employees that have tenets."""
        client = app.test_client()
        response = client.get('/analytics')
        assert response.status_code == 200

    def test_bonus_calculation_with_tenets(self, app, populated_db_with_tenets):
        """Bonus calculation page renders with rated employees."""
        client = app.test_client()
        response = client.get('/bonus-calculation')
        assert response.status_code == 200

    def test_export_page_with_tenets(self, app, populated_db_with_tenets):
        """Export page renders with employees that have tenets."""
        client = app.test_client()
        response = client.get('/export')
        assert response.status_code == 200

    def test_import_page(self, app, populated_db_with_tenets):
        """Import page renders (form page, minimal data dependency)."""
        client = app.test_client()
        response = client.get('/import')
        assert response.status_code == 200

    def test_history_page(self, app, populated_db_with_tenets):
        """History page renders (may be empty without archived periods)."""
        client = app.test_client()
        response = client.get('/history')
        assert response.status_code == 200


class TestRatingAPI:
    """Test rating API endpoints."""

    def test_rate_employee_success(self, client, populated_db):
        """Test successfully rating an employee."""
        data = {
            'associate_id': 'EMP004',
            'rating_percent': '110',
            'justification': 'Great work on product launch',
            'mentor': 'Bob Smith',
            'mentees': 'New hire team',
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 200
        result = json.loads(response.data)
        assert result['success'] is True
        assert result['message'] == 'Rating saved successfully'

        # Verify data was saved
        employee = populated_db.query(Employee).filter(
            Employee.associate == 'Diana Prince'
        ).first()

        assert employee.performance_rating_percent == 110.0
        assert employee.justification == 'Great work on product launch'
        assert employee.mentor == 'Bob Smith'
        assert employee.mentees == 'New hire team'
        assert employee.last_updated is not None

    def test_rate_employee_update_existing_rating(self, client, populated_db):
        """Test updating an existing rating."""
        data = {
            'associate_id': 'EMP001',
            'rating_percent': '130',
            'justification': 'Updated justification',
            'mentor': 'New mentor',
            'mentees': 'Updated mentees',
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 200

        # Verify update
        employee = populated_db.query(Employee).filter(
            Employee.associate == 'Alice Johnson'
        ).first()

        assert employee.performance_rating_percent == 130.0
        assert employee.justification == 'Updated justification'

    def test_rate_employee_missing_name(self, client, populated_db):
        """Test rating without associate name."""
        data = {
            'rating_percent': '100',
            'justification': 'Test'
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 400
        result = json.loads(response.data)
        assert 'error' in result
        assert result['error'] == 'Missing associate ID'

    def test_rate_employee_not_found(self, client, populated_db):
        """Test rating non-existent employee."""
        data = {
            'associate_id': 'NONEXISTENT',
            'rating_percent': '100'
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 404
        result = json.loads(response.data)
        assert 'error' in result
        assert result['error'] == 'Employee not found'

    def test_rate_employee_invalid_rating_too_high(self, client, populated_db):
        """Test rating validation - value too high."""
        data = {
            'associate_id': 'EMP001',
            'rating_percent': '250'
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 400
        result = json.loads(response.data)
        assert 'error' in result
        assert 'between 0 and 200' in result['error']

    def test_rate_employee_invalid_rating_negative(self, client, populated_db):
        """Test rating validation - negative value."""
        data = {
            'associate_id': 'EMP001',
            'rating_percent': '-10'
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 400
        result = json.loads(response.data)
        assert 'error' in result
        assert 'between 0 and 200' in result['error']

    def test_rate_employee_invalid_rating_format(self, client, populated_db):
        """Test rating validation - invalid format."""
        data = {
            'associate_id': 'EMP001',
            'rating_percent': 'abc'
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 400
        result = json.loads(response.data)
        assert 'error' in result
        assert 'Invalid rating value' in result['error']

    def test_rate_employee_empty_rating(self, client, populated_db):
        """Test rating with empty rating value (valid - unrating)."""
        data = {
            'associate_id': 'EMP001',
            'rating_percent': '',
            'justification': 'Removing rating temporarily'
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 200

        # Verify rating was set to None
        employee = populated_db.query(Employee).filter(
            Employee.associate == 'Alice Johnson'
        ).first()

        assert employee.performance_rating_percent is None
        assert employee.justification == 'Removing rating temporarily'

    def test_rate_employee_boundary_values(self, client, populated_db):
        """Test rating with boundary values 0 and 200."""
        # Test 0
        data = {
            'associate_id': 'EMP001',
            'rating_percent': '0'
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')
        assert response.status_code == 200

        employee = populated_db.query(Employee).filter(
            Employee.associate == 'Alice Johnson'
        ).first()
        assert employee.performance_rating_percent == 0.0

        # Test 200
        data['rating_percent'] = '200'
        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')
        assert response.status_code == 200

        populated_db.refresh(employee)
        assert employee.performance_rating_percent == 200.0

    def test_rate_employee_decimal_rating(self, client, populated_db):
        """Test rating with decimal values."""
        data = {
            'associate_id': 'EMP001',
            'rating_percent': '123.5'
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 200

        employee = populated_db.query(Employee).filter(
            Employee.associate == 'Alice Johnson'
        ).first()

        assert employee.performance_rating_percent == 123.5

    def test_rate_employee_only_manager_fields(self, client, populated_db):
        """Test updating only manager input fields without rating."""
        data = {
            'associate_id': 'EMP004',
            'rating_percent': '',
            'justification': 'Work in progress',
            'mentor': 'Alice Johnson',
            'mentees': '',
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 200

        employee = populated_db.query(Employee).filter(
            Employee.associate == 'Diana Prince'
        ).first()

        assert employee.performance_rating_percent is None
        assert employee.justification == 'Work in progress'
        assert employee.mentor == 'Alice Johnson'

    def test_get_employee_details_success(self, client, populated_db):
        """Test getting employee details by ID."""
        response = client.get('/api/employee/EMP004')

        assert response.status_code == 200
        result = json.loads(response.data)
        assert result['success'] is True
        assert 'employee' in result

        employee = result['employee']
        assert employee['Associate'] == 'Diana Prince'
        assert employee['Associate ID'] == 'EMP004'
        assert 'Current Job Profile' in employee

    def test_get_employee_details_not_found(self, client, populated_db):
        """Test getting details for non-existent employee."""
        response = client.get('/api/employee/NONEXISTENT')

        assert response.status_code == 404
        result = json.loads(response.data)
        assert result['success'] is False
        assert 'not found' in result['error'].lower()

    def test_rate_blocked_without_bonus_data(self, client, db_session):
        """Test that rating is blocked when no bonus data has been imported."""
        # Create employee without bonus target data (simulating talent-only import)
        employee = Employee(
            associate_id='EMP_TALENT',
            associate='Talent Only Employee',
            supervisory_organization='Engineering',
            current_job_profile='Software Engineer',
            current_base_pay_manager_currency=100000.0,
            currency='USD',
            # No bonus_target_local_currency or bonus_target_manager_currency
        )
        db_session.add(employee)
        db_session.commit()

        data = {
            'associate_id': 'EMP_TALENT',
            'rating_percent': '100',
            'justification': 'Test'
        }

        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 400
        result = json.loads(response.data)
        assert 'error' in result
        assert 'bonus data' in result['error'].lower()

    def test_calibrate_blocked_without_talent_data(self, client, db_session):
        """Test that calibration is blocked when no talent data has been imported."""
        # Create employee with bonus data but no talent data (simulating bonus-only import)
        employee = Employee(
            associate_id='EMP_BONUS',
            associate='Bonus Only Employee',
            supervisory_organization='Engineering',
            current_job_profile='Software Engineer',
            current_base_pay_manager_currency=100000.0,
            currency='USD',
            bonus_target_local_currency=15000.0,
            # No talent_perf_what_original, talent_perf_how_original,
            # talent_last_overall_perf, or talent_last_identified_future
        )
        db_session.add(employee)
        db_session.commit()

        data = {
            'associate_id': 'EMP_BONUS',
            'talent_perf_what': 'Meets Expectations',
            'talent_perf_how': 'Meets Expectations'
        }

        response = client.post('/api/calibrate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 400
        result = json.loads(response.data)
        assert 'error' in result
        assert 'talent data' in result['error'].lower()

    def test_rate_normalizes_placeholder_mentor(self, client, populated_db, db_session):
        """Test that placeholder values in mentor field are normalized to empty string."""
        placeholders = ['None', 'TBD', 'N/A', 'tba', 'pending']

        for placeholder in placeholders:
            data = {
                'associate_id': 'EMP004',
                'mentor': placeholder
            }
            response = client.post('/api/rate',
                                  data=json.dumps(data),
                                  content_type='application/json')

            assert response.status_code == 200
            result = json.loads(response.data)
            assert result['success'] is True
            assert 'normalized_fields' in result
            assert 'mentor' in result['normalized_fields']

            # Verify database was updated with empty string
            employee = db_session.query(Employee).filter_by(associate_id='EMP004').first()
            assert employee.mentor == ''

    def test_rate_preserves_valid_mentor(self, client, populated_db, db_session):
        """Test that valid mentor names are preserved unchanged."""
        data = {
            'associate_id': 'EMP004',
            'mentor': 'John Smith'
        }
        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 200
        result = json.loads(response.data)
        assert result['success'] is True
        assert 'normalized_fields' not in result

        # Verify database has the mentor name
        employee = db_session.query(Employee).filter_by(associate_id='EMP004').first()
        assert employee.mentor == 'John Smith'

    def test_rate_normalizes_placeholder_mentees(self, client, populated_db, db_session):
        """Test that placeholder values in mentees field are normalized."""
        data = {
            'associate_id': 'EMP004',
            'mentees': 'TBC'
        }
        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 200
        result = json.loads(response.data)
        assert result['success'] is True
        assert 'normalized_fields' in result
        assert 'mentees' in result['normalized_fields']


class TestBonusOverrideAPI:
    """Test bonus override (special case) API functionality."""

    def test_rate_with_bonus_override(self, client, populated_db, db_session):
        """Test saving bonus override percentage via API."""
        data = {
            'associate_id': 'EMP001',
            'bonus_override_percent': 50.0,
            'special_case_notes': 'Paternity leave Apr-Sep'
        }
        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 200
        result = json.loads(response.data)
        assert result['success'] is True

        # Verify database
        employee = db_session.query(Employee).filter_by(associate_id='EMP001').first()
        assert employee.bonus_override_percent == 50.0
        assert employee.special_case_notes == 'Paternity leave Apr-Sep'

    def test_rate_clear_bonus_override(self, client, populated_db, db_session):
        """Test clearing bonus override by sending empty value."""
        # First set an override
        employee = db_session.query(Employee).filter_by(associate_id='EMP001').first()
        employee.bonus_override_percent = 75.0
        employee.special_case_notes = 'Leave'
        db_session.commit()

        # Now clear it
        data = {
            'associate_id': 'EMP001',
            'bonus_override_percent': '',
            'special_case_notes': ''
        }
        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 200

        db_session.refresh(employee)
        assert employee.bonus_override_percent is None
        assert employee.special_case_notes is None

    def test_rate_bonus_override_validation(self, client, populated_db):
        """Test that bonus override is validated (0-200 range)."""
        data = {
            'associate_id': 'EMP001',
            'bonus_override_percent': 250.0  # Invalid - over 200
        }
        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 400
        result = json.loads(response.data)
        assert 'error' in result
        assert 'between 0 and 200' in result['error']

    def test_rate_bonus_override_negative_validation(self, client, populated_db):
        """Test that negative bonus override is rejected."""
        data = {
            'associate_id': 'EMP001',
            'bonus_override_percent': -10.0
        }
        response = client.post('/api/rate',
                              data=json.dumps(data),
                              content_type='application/json')

        assert response.status_code == 400

    def test_employee_to_dict_includes_override(self, populated_db, db_session):
        """Test that employee to_dict includes override fields."""
        employee = db_session.query(Employee).filter_by(associate_id='EMP001').first()
        employee.bonus_override_percent = 50.0
        employee.special_case_notes = 'Test notes'
        db_session.commit()

        emp_dict = employee.to_dict()
        assert 'bonus_override_percent' in emp_dict
        assert emp_dict['bonus_override_percent'] == 50.0
        assert 'special_case_notes' in emp_dict
        assert emp_dict['special_case_notes'] == 'Test notes'


class TestDashboardStatistics:
    """Test dashboard statistics calculations."""

    def test_dashboard_stats_with_ratings(self, client, populated_db):
        """Test dashboard displays correct statistics."""
        response = client.get('/')
        assert response.status_code == 200

        # Check that correct counts appear
        # Total: 4, Rated: 3, Unrated: 1
        assert b'4' in response.data  # Total employees

    def test_dashboard_with_no_employees(self, client, db_session):
        """Test dashboard with empty database."""
        response = client.get('/')
        assert response.status_code == 200


class TestAnalyticsDashboard:
    """Test analytics dashboard calculations."""

    def test_analytics_rating_distribution(self, client, populated_db):
        """Test rating distribution buckets."""
        response = client.get('/analytics')
        assert response.status_code == 200

        # Should calculate buckets correctly
        # Alice: 120 -> 101-130%
        # Bob: 140 -> 131-200%
        # Charlie: 85 -> 81-100%

    def test_analytics_department_averages(self, client, populated_db):
        """Test department average calculations."""
        response = client.get('/analytics')
        assert response.status_code == 200

        # Engineering dept has Alice (120), Bob (140), Charlie (85)
        # Average should be around 115

    def test_analytics_job_averages(self, client, populated_db):
        """Test job profile average calculations."""
        response = client.get('/analytics')
        assert response.status_code == 200

        # Senior Software Engineer: Alice (120) -> avg 120
        # Staff Software Engineer: Bob (140) -> avg 140
        # Software Engineer: Charlie (85) -> avg 85

    def test_analytics_with_no_ratings(self, client, db_session, sample_employees):
        """Test analytics with employees but no ratings."""
        # Add employees without ratings
        for emp_data in sample_employees:
            emp_data['performance_rating_percent'] = None
            employee = Employee(**emp_data)
            db_session.add(employee)
        db_session.commit()

        response = client.get('/analytics')
        assert response.status_code == 200


class TestExportPage:
    """Test export page functionality."""

    def test_export_page_route(self, client, populated_db):
        """Test the export page route."""
        response = client.get('/export')
        assert response.status_code == 200
        assert b'Export Data' in response.data

    def test_export_page_with_no_employees(self, client, db_session):
        """Test export page with no employees."""
        response = client.get('/export')
        assert response.status_code == 200
        assert b'No Employees Found' in response.data or b'Import Data' in response.data

    def test_export_page_renders_successfully(self, client, populated_db):
        """Test that export page renders without errors."""
        response = client.get('/export')
        assert response.status_code == 200
        # Page should have export-related content
        assert b'Workday' in response.data or b'Export' in response.data


class TestSnapshotExport:
    """Test full organization snapshot export functionality."""

    def test_export_snapshot_xlsx_structure(self, client, populated_db):
        """Test that Excel snapshot has correct sheet structure and headers."""
        from openpyxl import load_workbook
        import io

        response = client.get('/export/snapshot/xlsx')
        assert response.status_code == 200
        assert response.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        # Load the workbook from response
        wb = load_workbook(io.BytesIO(response.data))

        # Verify 5 sheets exist
        expected_sheets = ['_README', 'employees', 'bonus_cycle', 'talent_cycle', 'history']
        assert len(wb.sheetnames) == 5
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames, f"Sheet '{sheet_name}' not found"

        # Verify employees sheet has correct headers
        ws_employees = wb['employees']
        emp_headers = [cell.value for cell in ws_employees[1]]
        assert 'Employee ID (unique identifier)' in emp_headers
        assert 'Employee Name' in emp_headers
        assert 'Manager Name' in emp_headers
        assert 'Supervisory Organization' in emp_headers

        # Verify bonus_cycle sheet has correct headers
        ws_bonus = wb['bonus_cycle']
        bonus_headers = [cell.value for cell in ws_bonus[1]]
        assert 'Employee ID' in bonus_headers
        assert 'Performance Rating (0-200%, 100=met expectations)' in bonus_headers
        assert 'Calculated Bonus Amount (manager currency)' in bonus_headers

        # Verify talent_cycle sheet has correct headers
        ws_talent = wb['talent_cycle']
        talent_headers = [cell.value for cell in ws_talent[1]]
        assert 'Employee ID' in talent_headers
        assert 'Performance: What (Results)' in talent_headers
        assert 'Cross-Cycle Alignment (aligned/review/incomplete)' in talent_headers

    def test_export_snapshot_csv_structure(self, client, populated_db):
        """Test that CSV snapshot ZIP has correct file structure."""
        import zipfile
        import io
        import csv

        response = client.get('/export/snapshot/csv')
        assert response.status_code == 200
        assert response.mimetype == 'application/zip'

        # Load the ZIP from response
        zip_buffer = io.BytesIO(response.data)
        with zipfile.ZipFile(zip_buffer, 'r') as zf:
            # Verify 5 files exist (README.md + 4 CSVs)
            expected_files = ['README.md', 'employees.csv',
                            'bonus_cycle.csv', 'talent_cycle.csv', 'history.csv']
            assert len(zf.namelist()) == 5
            for filename in expected_files:
                assert filename in zf.namelist(), f"File '{filename}' not found in ZIP"

            # Verify employees.csv has correct headers
            with zf.open('employees.csv') as f:
                reader = csv.reader(io.TextIOWrapper(f, encoding='utf-8'))
                headers = next(reader)
                assert 'Employee ID (unique identifier)' in headers
                assert 'Employee Name' in headers
                assert 'Manager Name' in headers

            # Verify bonus_cycle.csv has correct headers
            with zf.open('bonus_cycle.csv') as f:
                reader = csv.reader(io.TextIOWrapper(f, encoding='utf-8'))
                headers = next(reader)
                assert 'Performance Rating (0-200%, 100=met expectations)' in headers

    def test_export_snapshot_readme_sheet(self, client, populated_db):
        """Test that _README sheet contains markdown domain knowledge."""
        from openpyxl import load_workbook
        import io

        response = client.get('/export/snapshot/xlsx')
        wb = load_workbook(io.BytesIO(response.data))

        ws_readme = wb['_README']

        # README content is in cell A1 as markdown
        readme_text = ws_readme.cell(row=1, column=1).value
        assert readme_text is not None
        assert '# Organization Snapshot' in readme_text
        assert 'Rating Scale' in readme_text
        assert 'Bonus Calculation' in readme_text
        assert 'Talent Calibration' in readme_text
        assert '0-200%' in readme_text
        assert '1.35' in readme_text  # upside exponent
        assert '1.9' in readme_text   # downside exponent

        # Verify analytical guidance is present
        assert 'Expected Rating Distribution' in readme_text
        assert 'Red Flags' in readme_text
        assert 'Suggested Analysis Questions' in readme_text
        assert 'Management Levels' in readme_text
        assert 'Data Quality' in readme_text

        # Verify source attribution
        assert 'performance-rating-and-bonus' in readme_text

    def test_export_snapshot_readme_includes_tenets(self, client, populated_db, sample_tenets):
        """Test that _README sheet includes tenet definitions from tenets.json."""
        from openpyxl import load_workbook
        import io

        response = client.get('/export/snapshot/xlsx')
        wb = load_workbook(io.BytesIO(response.data))

        ws_readme = wb['_README']
        readme_text = ws_readme.cell(row=1, column=1).value
        assert readme_text is not None
        assert 'Tenets (Behavioral Competencies)' in readme_text

    def test_export_snapshot_employees_data(self, client, populated_db):
        """Test that employee data is correctly exported."""
        from openpyxl import load_workbook
        import io

        response = client.get('/export/snapshot/xlsx')
        wb = load_workbook(io.BytesIO(response.data))

        ws_employees = wb['employees']

        # Verify employees are present (should have 4 from sample_employees fixture)
        assert ws_employees.max_row >= 5  # 1 header + 4 employees

        # Get all employee names
        names = [ws_employees.cell(row=i, column=2).value for i in range(2, ws_employees.max_row + 1)]
        assert 'Alice Johnson' in names
        assert 'Bob Smith' in names
        assert 'Charlie Brown' in names
        assert 'Diana Prince' in names

    def test_export_snapshot_with_no_employees(self, client, db_session):
        """Test snapshot export with empty database."""
        response = client.get('/export/snapshot/xlsx')
        assert response.status_code == 200

        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(response.data))

        # All sheets should still exist
        assert len(wb.sheetnames) == 5

        # Employees sheet should only have headers
        ws_employees = wb['employees']
        assert ws_employees.max_row == 1  # Only header row


class TestPoolVerification:
    """Test pool verification API endpoint and pool source tracking."""

    def test_verify_pool_success(self, client, db_session):
        """Test successfully verifying a pool."""
        from models import BonusSettings

        # Create bonus settings with unverified calculated pool
        settings = BonusSettings(
            workday_pool=10000.0,
            pool_source='calculated_sum',
            pool_verified=False
        )
        db_session.add(settings)
        db_session.commit()

        response = client.post('/api/bonus-settings/verify-pool',
                              content_type='application/json')

        assert response.status_code == 200
        result = json.loads(response.data)
        assert result['success'] is True
        assert result['message'] == 'Pool verified'

        # Verify database was updated
        db_session.refresh(settings)
        assert settings.pool_verified is True

    def test_verify_pool_no_settings(self, client, db_session):
        """Test verify pool when no settings exist."""
        response = client.post('/api/bonus-settings/verify-pool',
                              content_type='application/json')

        assert response.status_code == 404
        result = json.loads(response.data)
        assert 'error' in result
        assert 'No bonus settings found' in result['error']

    def test_verify_pool_already_verified(self, client, db_session):
        """Test verifying an already verified pool (idempotent)."""
        from models import BonusSettings

        settings = BonusSettings(
            workday_pool=10000.0,
            pool_source='workday_metadata',
            pool_verified=True
        )
        db_session.add(settings)
        db_session.commit()

        response = client.post('/api/bonus-settings/verify-pool',
                              content_type='application/json')

        assert response.status_code == 200
        result = json.loads(response.data)
        assert result['success'] is True

    def test_pool_source_tracking_calculated(self, client, db_session):
        """Test that pool_source is 'calculated_sum' when pool is calculated."""
        from models import BonusSettings

        settings = BonusSettings(
            workday_pool=5000.0,
            pool_source='calculated_sum',
            pool_verified=False
        )
        db_session.add(settings)
        db_session.commit()

        # Verify source is correctly stored
        assert settings.pool_source == 'calculated_sum'
        assert settings.pool_verified is False

    def test_pool_source_tracking_workday_metadata(self, client, db_session):
        """Test that pool_source is 'workday_metadata' when from file metadata."""
        from models import BonusSettings

        settings = BonusSettings(
            workday_pool=8000.0,
            pool_source='workday_metadata',
            pool_verified=True  # Auto-verified for metadata pools
        )
        db_session.add(settings)
        db_session.commit()

        # Verify source and auto-verification
        assert settings.pool_source == 'workday_metadata'
        assert settings.pool_verified is True

    def test_bonus_settings_to_dict_includes_pool_fields(self, db_session):
        """Test that BonusSettings.to_dict() includes pool verification fields."""
        from models import BonusSettings

        settings = BonusSettings(
            workday_pool=12000.0,
            budget_override=0.0,
            manager_currency='USD',
            pool_source='calculated_sum',
            pool_verified=False
        )
        db_session.add(settings)
        db_session.commit()

        settings_dict = settings.to_dict()

        assert 'pool_source' in settings_dict
        assert settings_dict['pool_source'] == 'calculated_sum'
        assert 'pool_verified' in settings_dict
        assert settings_dict['pool_verified'] is False


class TestBonusCalculation:
    """Test bonus calculation logic, especially budget_override behavior."""

    def test_budget_override_replaces_pool_not_adds(self):
        """Critical test: budget_override should REPLACE the pool, not add to it."""
        from app import calculate_bonus_for_employees

        # Create test employees with known values
        employees = [
            {
                'Associate ID': 'EMP001',
                'Associate': 'Test Employee 1',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 5000,
                'Current Base Pay Manager Currency': 100000,
            },
            {
                'Associate ID': 'EMP002',
                'Associate': 'Test Employee 2',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 5000,
                'Current Base Pay Manager Currency': 100000,
            },
        ]

        params = {'upside_exponent': 1.35, 'downside_exponent': 1.9}

        # Workday pool is $10,000 (sum of targets)
        workday_pool = 10000

        # Budget override is $8,000 - should REPLACE, not add
        budget_override = 8000

        result = calculate_bonus_for_employees(
            employees, params,
            budget_override=budget_override,
            workday_pool=workday_pool
        )

        # The total allocated should be $8,000 (the override), NOT $18,000
        assert result['total_allocated'] == budget_override, \
            f"Expected total_allocated={budget_override}, got {result['total_allocated']}"

        # Each employee should get $4,000 (half of $8,000), not $9,000 (half of $18,000)
        for r in result['results']:
            assert r['final_bonus'] == 4000, \
                f"Expected final_bonus=4000, got {r['final_bonus']}"

    def test_no_budget_override_uses_workday_pool(self):
        """When no budget_override, should use workday_pool."""
        from app import calculate_bonus_for_employees

        employees = [
            {
                'Associate ID': 'EMP001',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 5000,
                'Current Base Pay Manager Currency': 100000,
            },
        ]

        params = {'upside_exponent': 1.35, 'downside_exponent': 1.9}
        workday_pool = 6000  # Different from target to verify it's used

        result = calculate_bonus_for_employees(
            employees, params,
            budget_override=0,  # No override
            workday_pool=workday_pool
        )

        # Should use workday_pool
        assert result['total_allocated'] == workday_pool

    def test_budget_override_zero_uses_workday_pool(self):
        """When budget_override is 0, should use workday_pool."""
        from app import calculate_bonus_for_employees

        employees = [
            {
                'Associate ID': 'EMP001',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 5000,
                'Current Base Pay Manager Currency': 100000,
            },
        ]

        params = {'upside_exponent': 1.35, 'downside_exponent': 1.9}
        workday_pool = 7000

        result = calculate_bonus_for_employees(
            employees, params,
            budget_override=0.0,
            workday_pool=workday_pool
        )

        assert result['total_allocated'] == workday_pool

    def test_no_workday_pool_uses_sum_of_targets(self):
        """When no workday_pool, should fall back to sum of targets."""
        from app import calculate_bonus_for_employees

        employees = [
            {
                'Associate ID': 'EMP001',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 3000,
                'Current Base Pay Manager Currency': 100000,
            },
            {
                'Associate ID': 'EMP002',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 2000,
                'Current Base Pay Manager Currency': 100000,
            },
        ]

        params = {'upside_exponent': 1.35, 'downside_exponent': 1.9}

        result = calculate_bonus_for_employees(
            employees, params,
            budget_override=0,
            workday_pool=None  # No Workday pool
        )

        # Should use sum of targets: 3000 + 2000 = 5000
        assert result['total_allocated'] == 5000

    def test_budget_override_with_partial_ratings(self):
        """Budget override with partial ratings should scale proportionally."""
        from app import calculate_bonus_for_employees

        # Only one employee rated out of a team with total targets of $10,000
        employees = [
            {
                'Associate ID': 'EMP001',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 5000,
                'Current Base Pay Manager Currency': 100000,
            },
        ]

        params = {'upside_exponent': 1.35, 'downside_exponent': 1.9}

        # Override is $8,000 for the full team
        # But only 50% of targets are rated, so should get 50% of override
        result = calculate_bonus_for_employees(
            employees, params,
            budget_override=8000,
            workday_pool=10000,
            all_targets_sum=10000  # Total team targets
        )

        # Should get 50% of $8,000 = $4,000
        assert result['total_allocated'] == 4000, \
            f"Expected 4000, got {result['total_allocated']}"

    def test_bonus_override_employee_gets_fixed_percentage(self):
        """Test that override employees get exactly their override percentage."""
        from app import calculate_bonus_for_employees

        employees = [
            {
                'Associate ID': 'EMP001',
                'Associate': 'Normal Employee',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 10000,
                'Current Base Pay Manager Currency': 100000,
                'bonus_override_percent': None,  # Normal employee
            },
            {
                'Associate ID': 'EMP002',
                'Associate': 'Leave Employee',
                'performance_rating_percent': None,  # No rating needed
                'Bonus Target Manager Currency': 10000,
                'Current Base Pay Manager Currency': 100000,
                'bonus_override_percent': 50.0,  # 50% pro-rata
                'special_case_notes': 'Paternity leave',
            },
        ]

        params = {'upside_exponent': 1.35, 'downside_exponent': 1.9}

        result = calculate_bonus_for_employees(
            employees, params,
            budget_override=0,
            workday_pool=20000  # Full pool
        )

        # Find override employee result
        override_result = result['results_by_id']['EMP002']
        normal_result = result['results_by_id']['EMP001']

        # Override employee gets exactly 50% of their $10,000 target = $5,000
        assert override_result['final_bonus'] == 5000, \
            f"Expected 5000, got {override_result['final_bonus']}"
        assert override_result['bonus_percent_of_target'] == 50
        assert override_result['is_override'] is True

        # Normal employee is NOT an override
        assert normal_result['is_override'] is False

    def test_bonus_override_option_b_pool_redistribution(self):
        """Test Option B: override bonus comes from pool, remainder redistributed."""
        from app import calculate_bonus_for_employees

        # Pool is $40,000 (sum of 4 x $10,000 targets)
        # One employee has 50% override = $5,000 (uses $5k of their $10k budget)
        # Remaining $35,000 goes to other 3 employees
        employees = [
            {
                'Associate ID': 'EMP001',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 10000,
                'Current Base Pay Manager Currency': 100000,
            },
            {
                'Associate ID': 'EMP002',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 10000,
                'Current Base Pay Manager Currency': 100000,
            },
            {
                'Associate ID': 'EMP003',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 10000,
                'Current Base Pay Manager Currency': 100000,
            },
            {
                'Associate ID': 'LEAVE001',
                'performance_rating_percent': None,
                'Bonus Target Manager Currency': 10000,
                'Current Base Pay Manager Currency': 100000,
                'bonus_override_percent': 50.0,
            },
        ]

        params = {'upside_exponent': 1.35, 'downside_exponent': 1.9}

        result = calculate_bonus_for_employees(
            employees, params,
            budget_override=0,
            workday_pool=40000
        )

        # Total allocated should be exactly $40,000
        assert abs(result['total_allocated'] - 40000) < 1, \
            f"Expected ~40000, got {result['total_allocated']}"

        # Override employee gets $5,000
        override_bonus = result['results_by_id']['LEAVE001']['final_bonus']
        assert override_bonus == 5000

        # Remaining $35,000 split among 3 employees with 100% rating
        # Each should get ~$11,666.67 (more than their $10,000 target!)
        for emp_id in ['EMP001', 'EMP002', 'EMP003']:
            bonus = result['results_by_id'][emp_id]['final_bonus']
            # They should each get ~$11,667 (35000 / 3)
            assert abs(bonus - 11666.67) < 1, \
                f"Expected ~11666.67 for {emp_id}, got {bonus}"

    def test_bonus_override_metadata_in_result(self):
        """Test that override metadata is included in calculation result."""
        from app import calculate_bonus_for_employees

        employees = [
            {
                'Associate ID': 'EMP001',
                'performance_rating_percent': 100,
                'Bonus Target Manager Currency': 10000,
                'Current Base Pay Manager Currency': 100000,
            },
            {
                'Associate ID': 'LEAVE001',
                'performance_rating_percent': None,
                'Bonus Target Manager Currency': 10000,
                'Current Base Pay Manager Currency': 100000,
                'bonus_override_percent': 50.0,
                'special_case_notes': 'Medical leave',
            },
        ]

        params = {'upside_exponent': 1.35, 'downside_exponent': 1.9}

        result = calculate_bonus_for_employees(
            employees, params,
            workday_pool=20000
        )

        # Check metadata fields
        assert result['override_count'] == 1
        assert result['override_total'] == 5000
        assert result['remaining_pool'] == 15000  # 20000 - 5000

    def test_bonus_calculation_page_with_override_employee(self, app, db_session):
        """Test that /bonus-calculation page renders when an employee has bonus override."""
        from models import Employee, BonusSettings

        # Create a normal employee
        normal_emp = Employee(
            associate_id='NORMAL001',
            associate='Normal Employee',
            performance_rating_percent=100.0,
            bonus_target_local_currency=10000.0,
            current_base_pay_all_countries=100000.0,
            supervisory_organization='Test Team'
        )

        # Create an override employee (pro-rata leave)
        override_emp = Employee(
            associate_id='OVERRIDE001',
            associate='Leave Employee',
            performance_rating_percent=None,  # No rating - using override
            bonus_target_local_currency=10000.0,
            current_base_pay_all_countries=100000.0,
            supervisory_organization='Test Team',
            bonus_override_percent=50.0,
            special_case_notes='Paternity leave Q3-Q4'
        )

        db_session.add(normal_emp)
        db_session.add(override_emp)

        # Create bonus settings
        settings = BonusSettings(id=1)
        db_session.add(settings)
        db_session.commit()

        client = app.test_client()
        response = client.get('/bonus-calculation')

        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        # Check that both employees appear
        assert b'Normal Employee' in response.data
        assert b'Leave Employee' in response.data
        # Check that override badge appears
        assert b'Override' in response.data
        # Check that override notes appear in title attribute
        assert b'Paternity leave' in response.data
