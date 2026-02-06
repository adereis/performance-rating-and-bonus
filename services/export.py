"""Export helper functions: data preparation, tenet resolution, XLSX writing.

Pure functions with no Flask or database dependencies. These operate on
employee dicts (from to_dict()) and prepare data for CSV/XLSX export.
Route handlers remain in app.py as thin wrappers.
"""
import json

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from services.employee_utils import (
    get_currency_format,
    get_rating_category,
    is_manager,
    parse_manager_name_from_org,
    CURRENCY_FORMATS,
)


# Standard header styling for export XLSX sheets
HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')


def resolve_tenets_text(emp, field_prefix, tenets_map, separator='; '):
    """Resolve tenet IDs to human-readable names for export.

    Tenet fields store JSON arrays of tenet IDs (e.g. '["t1","t2"]').
    This resolves them to display names using the tenets_map lookup.

    Args:
        emp: Employee dict (from to_dict())
        field_prefix: Field name prefix, e.g. 'tenets' or 'talent_tenets'.
            Looks up '{field_prefix}_strengths' and '{field_prefix}_improvements'.
        tenets_map: Dict mapping tenet ID -> tenet name
        separator: Join separator ('; ' for snapshot exports, ', ' for legacy)

    Returns:
        Tuple of (strengths_text, improvements_text) as joined strings
    """
    strengths_text = ''
    improvements_text = ''

    try:
        strengths_field = f'{field_prefix}_strengths'
        improvements_field = f'{field_prefix}_improvements'

        if emp.get(strengths_field):
            strength_ids = (
                json.loads(emp[strengths_field])
                if isinstance(emp[strengths_field], str)
                else emp[strengths_field]
            )
            names = [tenets_map.get(tid, tid) for tid in strength_ids if tid in tenets_map]
            strengths_text = separator.join(names)

        if emp.get(improvements_field):
            improvement_ids = (
                json.loads(emp[improvements_field])
                if isinstance(emp[improvements_field], str)
                else emp[improvements_field]
            )
            names = [tenets_map.get(tid, tid) for tid in improvement_ids if tid in tenets_map]
            improvements_text = separator.join(names)
    except Exception:
        pass

    return strengths_text, improvements_text


def write_xlsx_sheet(ws, headers, rows, header_font=None, header_fill=None):
    """Write headers and data rows to an openpyxl worksheet.

    Applies header styling (font + fill) and auto-adjusts column widths.

    Args:
        ws: openpyxl Worksheet object
        headers: List of header strings
        rows: List of lists, each inner list is one data row (same order as headers)
        header_font: Font for headers (defaults to HEADER_FONT)
        header_fill: PatternFill for headers (defaults to HEADER_FILL)
    """
    if header_font is None:
        header_font = HEADER_FONT
    if header_fill is None:
        header_fill = HEADER_FILL

    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
        if max_length > 0:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


# ---------------------------------------------------------------------------
# Snapshot export data preparation
# ---------------------------------------------------------------------------

# Header definitions shared between XLSX and CSV snapshot exports.
# Each constant is a list of descriptive column names.

SNAPSHOT_EMPLOYEE_HEADERS = [
    'Employee ID (unique identifier)',
    'Employee Name',
    'Manager Name',
    'Supervisory Organization',
    'Job Title',
    'Management Level (IC/Manager/Director)',
    'Grade',
    'Country',
    'Region',
    'Currency (local)',
    'Hire Date',
    'Length of Service',
    'Time in Current Role',
    'Annual Base Pay (local currency)',
    'Annual Base Pay (manager currency)',
    'Bonus Target % of Base Pay',
    'Bonus Target Amount (local currency)',
    'Bonus Target Amount (manager currency)',
    'Is Manager (has direct reports)',
]

SNAPSHOT_BONUS_HEADERS = [
    'Employee ID',
    'Employee Name',
    'Performance Rating (0-200%, 100=met expectations)',
    'Rating Category (High/Solid/Below)',
    'Calculated Bonus Amount (manager currency)',
    'Bonus % of Target',
    'Justification',
    'Strength Tenets',
    'Improvement Tenets',
    'Mentor',
    'Mentees',
    'Last Updated',
]

SNAPSHOT_TALENT_HEADERS = [
    'Employee ID',
    'Employee Name',
    'Performance: What (Results)',
    'Performance: How (Behaviors)',
    'Overall Performance (derived)',
    'Previous Overall Performance',
    'Growth Agility',
    'Change Agility',
    'Identified as Future Talent',
    'Previous Future Talent Status',
    'Movement Readiness',
    'Previous Movement Readiness',
    'Proposed Talent Actions',
    'Talent Strength Tenets',
    'Talent Improvement Tenets',
    'Talent Mentor',
    'Talent Mentees',
    'Promotion: Proposed Job Profile',
    'Promotion: Business Need',
    'Promotion: Expanded Role Scope',
    'Promotion: Associate Readiness',
    'Cross-Cycle Alignment (aligned/review/incomplete)',
    'Last Updated',
]

SNAPSHOT_HISTORY_HEADERS = [
    'Period ID',
    'Period Name',
    'Cycle Type (bonus/talent)',
    'Archived Date',
    'Employee ID',
    'Employee Name (at snapshot)',
    'Supervisory Org (at snapshot)',
    'Job Profile (at snapshot)',
    'Performance Rating (0-200%)',
    'Final Bonus Allocation',
    'Bonus Target (at snapshot)',
    'Justification',
    'Strength Tenets',
    'Improvement Tenets',
    'Mentors',
    'Mentees',
    'Talent: Overall Performance',
    'Talent: What',
    'Talent: How',
    'Talent: Growth Agility',
    'Talent: Change Agility',
    'Talent: Movement Readiness',
]


def _format_future_talent(value):
    """Format boolean future talent field for export display."""
    if value is True:
        return 'Yes'
    if value is False:
        return 'No'
    return ''


def prepare_snapshot_employee_row(emp):
    """Build a single employee row for snapshot exports.

    Args:
        emp: Employee dict (from to_dict())

    Returns:
        List of values matching SNAPSHOT_EMPLOYEE_HEADERS order
    """
    return [
        emp.get('Associate ID', ''),
        emp.get('Associate', ''),
        parse_manager_name_from_org(emp.get('Supervisory Organization', '')),
        emp.get('Supervisory Organization', ''),
        emp.get('Current Job Profile', ''),
        emp.get('management_level', ''),
        emp.get('Grade', ''),
        emp.get('country', ''),
        emp.get('region', ''),
        emp.get('Currency', ''),
        emp.get('hire_date', ''),
        emp.get('length_of_service', ''),
        emp.get('time_in_job_profile', ''),
        emp.get('Current Base Pay All Countries', ''),
        emp.get('Current Base Pay Manager Currency', ''),
        emp.get('Annual Bonus Target Percent', ''),
        emp.get('Bonus Target - Local Currency', ''),
        emp.get('Bonus Target Manager Currency', ''),
        is_manager(emp),
    ]


def prepare_snapshot_bonus_row(emp, tenets_map, bonus_results_by_id):
    """Build a single bonus_cycle row for snapshot exports.

    Args:
        emp: Employee dict (from to_dict())
        tenets_map: Dict mapping tenet ID -> tenet name
        bonus_results_by_id: Dict mapping associate_id -> bonus calculation result

    Returns:
        List of values matching SNAPSHOT_BONUS_HEADERS order
    """
    emp_id = emp.get('Associate ID', '')
    bonus_result = bonus_results_by_id.get(emp_id, {})

    strengths_text, improvements_text = resolve_tenets_text(
        emp, 'tenets', tenets_map
    )

    return [
        emp_id,
        emp.get('Associate', ''),
        emp.get('performance_rating_percent', ''),
        get_rating_category(emp.get('performance_rating_percent')),
        bonus_result.get('final_bonus', ''),
        bonus_result.get('percent_of_target', ''),
        emp.get('justification', ''),
        strengths_text,
        improvements_text,
        emp.get('mentor', ''),
        emp.get('mentees', ''),
        emp.get('last_updated', ''),
    ]


def prepare_snapshot_talent_row(emp, tenets_map, get_cross_cycle_alignment):
    """Build a single talent_cycle row for snapshot exports.

    Args:
        emp: Employee dict (from to_dict())
        tenets_map: Dict mapping tenet ID -> tenet name
        get_cross_cycle_alignment: Function from models.py that computes alignment

    Returns:
        List of values matching SNAPSHOT_TALENT_HEADERS order
    """
    talent_strengths_text, talent_improvements_text = resolve_tenets_text(
        emp, 'talent_tenets', tenets_map
    )

    cross_align = get_cross_cycle_alignment(
        emp.get('performance_rating_percent'),
        emp.get('talent_overall_perf')
    )

    return [
        emp.get('Associate ID', ''),
        emp.get('Associate', ''),
        emp.get('talent_perf_what', ''),
        emp.get('talent_perf_how', ''),
        emp.get('talent_overall_perf', ''),
        emp.get('talent_last_overall_perf', ''),
        emp.get('talent_growth_agility', ''),
        emp.get('talent_change_agility', ''),
        _format_future_talent(emp.get('talent_identified_future')),
        _format_future_talent(emp.get('talent_last_identified_future')),
        emp.get('talent_movement_readiness', ''),
        emp.get('talent_last_movement_readiness', ''),
        emp.get('talent_proposed_actions', ''),
        talent_strengths_text,
        talent_improvements_text,
        emp.get('talent_mentor', ''),
        emp.get('talent_mentees', ''),
        emp.get('talent_promo_job_profile', ''),
        emp.get('talent_promo_business_need', ''),
        emp.get('talent_promo_role_scope', ''),
        emp.get('talent_promo_readiness', ''),
        cross_align,
        emp.get('talent_last_updated', ''),
    ]


def prepare_snapshot_history_row(snap):
    """Build a single history row for snapshot exports.

    Args:
        snap: History snapshot dict (from get_all_history_snapshots())

    Returns:
        List of values matching SNAPSHOT_HISTORY_HEADERS order
    """
    return [
        snap.get('period_id', ''),
        snap.get('period_name', ''),
        snap.get('cycle_type', ''),
        snap.get('archived_at', ''),
        snap.get('associate_id', ''),
        snap.get('snapshot_name', ''),
        snap.get('snapshot_org', ''),
        snap.get('snapshot_job_profile', ''),
        snap.get('performance_rating', ''),
        snap.get('bonus_allocation', ''),
        snap.get('snapshot_bonus_target_manager_currency', ''),
        snap.get('justification', ''),
        snap.get('tenets_strengths', ''),
        snap.get('tenets_improvements', ''),
        snap.get('mentors', ''),
        snap.get('mentees', ''),
        snap.get('snapshot_talent_overall_perf', ''),
        snap.get('snapshot_talent_perf_what', ''),
        snap.get('snapshot_talent_perf_how', ''),
        snap.get('snapshot_talent_growth_agility', ''),
        snap.get('snapshot_talent_change_agility', ''),
        snap.get('snapshot_talent_movement_readiness', ''),
    ]


def prepare_snapshot_data(all_employees, tenets_map, bonus_results_by_id,
                          history_data, get_cross_cycle_alignment):
    """Prepare all four data sheets for snapshot export (XLSX or CSV).

    Sorts employees by name and builds row lists for each sheet.

    Args:
        all_employees: List of employee dicts
        tenets_map: Dict mapping tenet ID -> tenet name
        bonus_results_by_id: Dict mapping associate_id -> bonus calculation result
        history_data: List of history snapshot dicts
        get_cross_cycle_alignment: Function from models.py

    Returns:
        Dict with keys 'employees', 'bonus_cycle', 'talent_cycle', 'history',
        each containing a list of row lists (values only, no headers).
    """
    sorted_employees = sorted(all_employees, key=lambda x: x.get('Associate', ''))

    employee_rows = [
        prepare_snapshot_employee_row(emp)
        for emp in sorted_employees
    ]

    bonus_rows = [
        prepare_snapshot_bonus_row(emp, tenets_map, bonus_results_by_id)
        for emp in sorted_employees
    ]

    talent_rows = [
        prepare_snapshot_talent_row(emp, tenets_map, get_cross_cycle_alignment)
        for emp in sorted_employees
    ]

    history_rows = [
        prepare_snapshot_history_row(snap)
        for snap in history_data
    ]

    return {
        'employees': employee_rows,
        'bonus_cycle': bonus_rows,
        'talent_cycle': talent_rows,
        'history': history_rows,
    }


def build_context_markdown(tenets_config, demo_mode=False):
    """Build markdown README content for AI consumption.

    Generates human-readable prose explaining the domain knowledge, rating
    philosophy, bonus calculation algorithm, and tenet definitions. Designed
    to be consumed by AI tools (NotebookLM, Claude, etc.) alongside the
    data sheets.

    Args:
        tenets_config: Dict from tenets.json
        demo_mode: If True, adds demo mode warning at top

    Returns:
        Markdown string optimized for AI analysis
    """
    lines = []

    # Demo mode warning
    if demo_mode:
        lines.append('> **\u26a0\ufe0f DEMO MODE - FICTITIOUS DATA ONLY**')
        lines.append('> This export contains synthetic sample data for demonstration purposes.')
        lines.append('> Do not use for actual HR decisions.')
        lines.append('')
        lines.append('---')
        lines.append('')

    # Header with attribution
    lines.append('# Organization Snapshot')
    lines.append('')
    lines.append('> This data was exported from the Performance Rating Tool.')
    lines.append('> Source: https://github.com/adereis/performance-rating-and-bonus')
    lines.append('')
    lines.append('---')
    lines.append('')

    # How to use
    lines.append('## How to Use This Data')
    lines.append('')
    lines.append('**This is an auto-generated, read-only export.** The data in the accompanying')
    lines.append('sheets was entered by managers through a web application and exported for analysis.')
    lines.append('')
    lines.append('**Your role:** Analyze and summarize this data. Do not suggest improvements to')
    lines.append('how the data was entered\u2014the entry process is already complete.')
    lines.append('')
    lines.append('This README provides context to help you interpret the data sheets correctly.')
    lines.append('')

    # Rating Scale
    lines.append('## Rating Scale (0-200%)')
    lines.append('')
    lines.append('| Rating | Meaning |')
    lines.append('|--------|---------|')
    lines.append('| 0-60% | Significant performance concerns |')
    lines.append('| 60-90% | Needs improvement |')
    lines.append('| **90-110%** | **Met expectations** (most employees) |')
    lines.append('| 110-130% | Exceeded expectations |')
    lines.append('| 130-200% | Exceptional performance |')
    lines.append('')
    lines.append('**100% is the baseline** - a solid performer who met all expectations.')
    lines.append('')

    # Expected Distributions
    lines.append('### Expected Rating Distribution')
    lines.append('')
    lines.append('A healthy team shows a bell curve centered around 100% with slight right skew:')
    lines.append('')
    lines.append('- **~60-70%** of employees at 90-110% (met expectations)')
    lines.append('- **~15-25%** at 110-130% (exceeded expectations)')
    lines.append('- **~5-10%** at 130%+ (exceptional)')
    lines.append('- **~5-10%** below 90% (needs improvement)')
    lines.append('')
    lines.append('**Warning signs:**')
    lines.append('- If >50% are 120%+: Ratings may lack differentiation (grade inflation)')
    lines.append('- If everyone is 95-105%: Manager may be avoiding differentiation')
    lines.append('- If bimodal (two clusters): May indicate team culture issues or inconsistent standards')
    lines.append('')

    # Bonus Calculation
    lines.append('## Bonus Calculation (Three-Step Process)')
    lines.append('')
    lines.append("### Step 1: Start with Each Person's Target")
    lines.append('')
    lines.append("Every employee has a bonus target from Workday (typically a percentage of salary).")
    lines.append("This is their baseline - what they'd receive if everyone performed at 100%.")
    lines.append('')

    lines.append('### Step 2: Adjust for Performance')
    lines.append('')
    lines.append("Each person's share is modified based on their performance rating:")
    lines.append('')
    lines.append('- **Rating above 100%**: Exponential boost (e.g., 120% rating \u2192 128% of target)')
    lines.append('- **Rating at 100%**: Close to target amount')
    lines.append('- **Rating below 100%**: Exponential penalty (e.g., 80% rating \u2192 65% of target)')
    lines.append('')
    lines.append('**Formula:**')
    lines.append('- If rating \u2265 100%: `performance_multiplier = (rating/100)^1.35`')
    lines.append('- If rating < 100%: `performance_multiplier = (rating/100)^1.9`')
    lines.append('')
    lines.append('**Example calculations:**')
    lines.append('')
    lines.append('| Rating | Multiplier | Effect |')
    lines.append('|--------|------------|--------|')
    lines.append('| 140% | 1.40^1.35 = 1.58x | Exceptional boost |')
    lines.append('| 120% | 1.20^1.35 = 1.28x | Strong reward |')
    lines.append('| 100% | 1.00x | Baseline |')
    lines.append('| 80% | 0.80^1.9 = 0.65x | Significant penalty |')
    lines.append('| 60% | 0.60^1.9 = 0.39x | Severe penalty |')
    lines.append('')

    lines.append('### Step 3: Normalize to Budget')
    lines.append('')
    lines.append("After calculating everyone's \"raw shares\" based on performance, all bonuses")
    lines.append('scale proportionally so the total exactly matches the bonus pool.')
    lines.append('')
    lines.append('**Final Bonus = Bonus Target \u00d7 Performance Multiplier \u00d7 Normalization Factor**')
    lines.append('')

    # Normalization Scenarios
    lines.append('### Normalization Scenarios')
    lines.append('')
    lines.append('The normalization factor adjusts based on team composition:')
    lines.append('')
    lines.append('| Scenario | Norm Factor | Effect |')
    lines.append('|----------|-------------|--------|')
    lines.append('| **Balanced team** | ~0.90 | High performers compress budget; 100% performer gets ~90% of target |')
    lines.append('| **All average team** | 1.0 | Everyone at 100% rating gets exactly their target |')
    lines.append('| **Low performing team** | ~1.5 | Budget surplus redistributed; everyone gets more than raw calculation |')
    lines.append('')

    # Currency Handling
    lines.append('### Currency Handling')
    lines.append('')
    lines.append("- **Domestic employees**: Uses \"Bonus Target - Local Currency\" (same as manager currency)")
    lines.append('- **International employees**: Uses converted "Bonus Target Manager Currency" column')
    lines.append("- All calculations use the manager's currency")
    lines.append('')

    # Talent Calibration
    lines.append('## Talent Calibration')
    lines.append('')
    lines.append('### Overall Performance Derivation')
    lines.append('')
    lines.append('Overall Performance is derived from two dimensions: **What** (results) and **How** (behaviors).')
    lines.append('')
    lines.append('| What | How | Overall Performance |')
    lines.append('|------|-----|---------------------|')
    lines.append('| Surpasses | Surpasses | High Impact Performer |')
    lines.append('| Surpasses | Meets | High Impact Performer |')
    lines.append('| Meets | Surpasses | High Impact Performer |')
    lines.append('| Meets | Meets | Successful Performer |')
    lines.append('| Surpasses | Meets Some | Successful Performer |')
    lines.append('| Meets Some | Surpasses | Successful Performer |')
    lines.append('| Meets | Meets Some | Evolving Performer |')
    lines.append('| Meets Some | Meets | Evolving Performer |')
    lines.append('| Meets Some | Meets Some | Evolving Performer |')
    lines.append('| Any | Does Not Meet | Low Performer |')
    lines.append('| Does Not Meet | Any | Low Performer |')
    lines.append('')

    lines.append('### Future Talent Criteria')
    lines.append('')
    lines.append('An employee is identified as **Future Talent** if BOTH:')
    lines.append('- **Growth Agility** contains "Always"')
    lines.append('- **Change Agility** contains "Always"')
    lines.append('')

    lines.append('### Movement Readiness')
    lines.append('')
    lines.append('Indicates promotion readiness:')
    lines.append('- **Ready Now**: Prepared for immediate promotion')
    lines.append('- **Ready in 1-2 Years**: On track for future advancement')
    lines.append('- **Not Ready**: Needs development in current role')
    lines.append('')

    # Cross-Cycle Alignment
    lines.append('## Cross-Cycle Alignment')
    lines.append('')
    lines.append('Bonus ratings should align with talent calibration results:')
    lines.append('')
    lines.append('| Calibration Category | Expected Rating Range | Alignment |')
    lines.append('|---------------------|----------------------|-----------|')
    lines.append('| High Impact Performer | 120-200% | aligned |')
    lines.append('| Successful Performer | 90-119% | aligned |')
    lines.append('| Evolving Performer | 70-89% | aligned |')
    lines.append('| Low Performer | 0-69% | aligned |')
    lines.append('')
    lines.append('**Alignment values:**')
    lines.append('- **aligned**: Rating matches talent calibration - no action needed')
    lines.append('- **review**: Rating and calibration disagree - investigate the mismatch')
    lines.append('- **incomplete**: Missing either bonus rating or talent calibration')
    lines.append('')

    # Management Levels
    lines.append('## Management Levels')
    lines.append('')
    lines.append('| Level | Description |')
    lines.append('|-------|-------------|')
    lines.append('| IC (Individual Contributor) | No direct reports, technical/specialist track |')
    lines.append('| IC 1-2 | Early career |')
    lines.append('| IC 3-4 | Mid-level |')
    lines.append('| IC 5+ | Senior/staff |')
    lines.append('| Manager | First-line manager with direct reports |')
    lines.append('| Senior Manager | Manages managers or large teams |')
    lines.append('| Director | Manages multiple teams or functions |')
    lines.append('| VP/Executive | Senior leadership, organizational strategy |')
    lines.append('')

    # Tenets
    lines.append('## Tenets (Behavioral Competencies)')
    lines.append('')
    lines.append('Employees are assessed on behavioral tenets:')
    lines.append('- **3 strengths** per employee')
    lines.append('- **2-3 improvement areas** per employee')
    lines.append('')

    if tenets_config and 'tenets' in tenets_config:
        # Group by category
        categories = {}
        for tenet in tenets_config['tenets']:
            if tenet.get('active', True):
                cat = tenet.get('category', 'Other')
                if cat not in categories:
                    categories[cat] = []
                categories[cat].append(tenet)

        for category, tenets in categories.items():
            lines.append(f'### {category}')
            lines.append('')
            for tenet in tenets:
                lines.append(f'**{tenet.get("name", "")}**')
                lines.append(f'{tenet.get("description", "")}')
                lines.append('')

    # Data Sheets
    lines.append('## Data Sheets')
    lines.append('')
    lines.append('| Sheet | Description |')
    lines.append('|-------|-------------|')
    lines.append('| employees | Core identity, compensation, manager info |')
    lines.append('| bonus_cycle | Performance ratings, justifications, calculated bonuses |')
    lines.append('| talent_cycle | Calibration data, agility, movement readiness, promotions |')
    lines.append('| history | Historical rating snapshots from previous cycles |')
    lines.append('')

    # Suggested Analysis Questions
    lines.append('## Suggested Analysis Questions')
    lines.append('')
    lines.append('- **Distribution**: What is the rating distribution? Does it match the expected curve?')
    lines.append('- **By level**: Do ratings correlate with job level? (seniors should trend higher)')
    lines.append('- **By tenure**: Are new hires rated differently than tenured employees?')
    lines.append('- **Alignment**: How many employees have "review" cross-cycle alignment?')
    lines.append('- **Outliers**: Who are the top/bottom performers? Is justification adequate?')
    lines.append('- **Trends**: Compare to history: are ratings improving, declining, or stable?')
    lines.append('- **Tenets**: Which tenets appear most as strengths? As improvements?')
    lines.append('- **Mentorship**: Is mentorship well-distributed? Do high performers mentor?')
    lines.append('- **Future Talent**: What % identified as future talent? Is pipeline healthy?')
    lines.append('- **Movement Readiness**: Who is "Ready Now" for promotion? Any blockers?')
    lines.append('')

    # Red Flags
    lines.append('## Red Flags to Investigate')
    lines.append('')
    lines.append('| Pattern | Concern |')
    lines.append('|---------|---------|')
    lines.append('| High rating + Low Performer calibration | Possible over-rating, or calibration data outdated |')
    lines.append('| Low rating + High Impact calibration | Possible under-rating, or calibration data outdated |')
    lines.append('| Long tenure + consistent low ratings | May indicate development stagnation or role mismatch |')
    lines.append('| New hire + very high rating | Verify not just "honeymoon effect" - ensure evidence-based |')
    lines.append('| Manager with low-rated direct reports | May indicate leadership development need |')
    lines.append('| Same improvement tenets across team | May indicate team-wide skill gap or manager bias |')
    lines.append('| No justification for outlier ratings | Ratings <80% or >130% should have clear justification |')
    lines.append('')

    # Data Quality
    lines.append('## Data Quality Notes')
    lines.append('')
    lines.append('- **Missing ratings**: Employees in employees sheet but not in bonus_cycle may be unrated')
    lines.append('- **Missing calibration**: Employees without talent_cycle data have incomplete picture')
    lines.append('- **Empty justifications**: Non-100% ratings should have justification - empty is a gap')
    lines.append('- **Incomplete tenets**: Should have 3 strengths and 2-3 improvements per employee')
    lines.append('')

    # History Interpretation
    lines.append('## History Sheet Interpretation')
    lines.append('')
    lines.append('The history sheet contains archived snapshots from previous rating/calibration cycles:')
    lines.append('')
    lines.append('- **Period ID**: Unique identifier for the archived cycle')
    lines.append('- **Cycle Type**: "bonus" = performance rating cycle, "talent" = calibration cycle')
    lines.append('- **Snapshot data**: Employee data at time of archive (may differ from current)')
    lines.append('')
    lines.append('Use history for:')
    lines.append('- Trend analysis: Compare current vs historical ratings for same employee')
    lines.append('- Mobility tracking: Track job profile changes, org moves between periods')
    lines.append('')

    return '\n'.join(lines)
