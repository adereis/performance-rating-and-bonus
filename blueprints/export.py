"""Export blueprint: Workday bonus/talent exports and the full org snapshot.

Routes: /export, /export/csv, /export/xlsx, /export/talent, /export/talent/csv,
/export/snapshot/xlsx, /export/snapshot/csv.

Moved verbatim from app.py (docs/REFACTOR_APP_SPLIT.md, Phase 2). The demo
flag is read as demo_mode.DEMO_MODE and get_db via the models module (both
resolved at call time, and importing neither app nor creating a cycle that
would break `python app.py`, where app runs as __main__).
"""
import io
import csv
import json

from flask import (
    Blueprint, render_template, request, send_file, make_response, Response,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import demo_mode
import models
from models import Period

from services.db_helpers import (
    get_filter_params, get_all_employees, apply_employee_filters,
    get_bonus_settings, load_tenets_config, get_all_history_snapshots,
)
from services.employee_utils import is_employee_calibrated, is_employee_rated
from services.bonus import calculate_bonus_for_employees
from services.export import (
    build_context_markdown, prepare_snapshot_data, write_xlsx_sheet,
    SNAPSHOT_EMPLOYEE_HEADERS, SNAPSHOT_BONUS_HEADERS,
    SNAPSHOT_TALENT_HEADERS, SNAPSHOT_HISTORY_HEADERS,
)

export_bp = Blueprint('export', __name__)


@export_bp.route('/export')
def export_page():
    """Export page for Workday bonus and talent data."""
    # Get filter params from URL
    filter_params = get_filter_params()

    # Get all employees (bonus cycle only)
    all_employees = get_all_employees(bonus_cycle_only=True)

    # Apply filters
    team_data, filter_info = apply_employee_filters(all_employees, filter_params)

    # Detect which data types are available
    # Include override employees — they have a fixed bonus % even without a performance rating
    rated_employees = [
        emp for emp in team_data
        if emp.get('performance_rating_percent') is not None or emp.get('bonus_override_percent') is not None
    ]
    calibrated_employees = [emp for emp in team_data if is_employee_calibrated(emp)]

    # has_bonus_data: true only if rated employees have actual bonus target data from Workday
    # (not just ratings without bonus targets)
    rated_with_bonus_targets = [
        emp for emp in rated_employees
        if emp.get('Bonus Target Manager Currency') or emp.get('Bonus Target - Local Currency')
    ]
    has_bonus_data = len(rated_with_bonus_targets) > 0
    has_talent_data = len(calibrated_employees) > 0

    # Determine default mode: prefer bonus if available (ratings + targets), else talent
    export_mode = request.args.get('mode', 'bonus' if has_bonus_data else 'talent')

    # If no data at all, show empty state
    if not has_bonus_data and not has_talent_data:
        return render_template('export.html',
                             export_data=[],
                             talent_export_data=[],
                             has_data=False,
                             has_bonus_data=False,
                             has_talent_data=False,
                             export_mode='bonus',
                             bonus_pending_sync_count=0,
                             total_employees=len(team_data),
                             rated_count=len(rated_employees),
                             calibrated_count=len(calibrated_employees),
                             total_calibrated=0,
                             history_period_count=0,
                             filter_info=filter_info)

    # Get bonus calculation settings
    params = {
        'upside_exponent': float(request.args.get('upside_exponent', 1.35)),
        'downside_exponent': float(request.args.get('downside_exponent', 1.9))
    }

    # Get bonus settings (pool and override)
    settings = get_bonus_settings()
    budget_override = settings.budget_override if settings else 0.0
    workday_pool = settings.workday_pool if settings else None

    # Calculate sum of ALL employee bonus targets (for proportional pool calculation)
    # Must use all_employees (not team_data) so filtered-out employees are included
    all_targets_sum = 0
    for emp in all_employees:
        bonus_target = emp.get('Bonus Target Manager Currency') or emp.get('Bonus Target - Local Currency')
        if bonus_target:
            try:
                all_targets_sum += float(bonus_target)
            except (ValueError, TypeError):
                pass

    # Calculate bonuses for all rated employees
    bonus_calc = calculate_bonus_for_employees(rated_employees, params, budget_override, workday_pool, all_targets_sum)

    # Load tenets configuration
    _, tenets_map = load_tenets_config()

    # Helper functions for modification detection
    def is_field_modified(emp, field_name):
        """Check if a field was modified from its original imported value.

        Conservative: only returns True if we have an original to compare against.
        Used for Workday Content fields (justification, proposed_actions, promo fields).
        """
        current = emp.get(field_name) or ''
        original = emp.get(f'{field_name}_original')
        # Can only detect modification if we tracked the original value at import
        if original is None:
            return False
        return current != original

    def needs_sync_to_workday(emp, field_name):
        """Check if a Tool Additions field needs to be synced to Workday.

        More aggressive: returns True if content exists but wasn't in original import.
        Used for tool-generated fields (tenets, mentor, mentees).
        """
        current = emp.get(field_name) or ''
        original = emp.get(f'{field_name}_original')
        # If we have content but no original, it's a new addition needing sync
        if original is None:
            return bool(current)
        # If we have an original, compare
        return current != (original or '')

    # Format export data
    export_data = []
    bonus_pending_sync_count = 0
    for result in bonus_calc['results']:
        employee = result['employee']

        # Parse tenets
        strengths = []
        improvements = []
        try:
            if employee.get('tenets_strengths'):
                strength_ids = json.loads(employee['tenets_strengths']) if isinstance(employee['tenets_strengths'], str) else employee['tenets_strengths']
                strengths = [tenets_map.get(tid, tid) for tid in strength_ids if tid in tenets_map]
            if employee.get('tenets_improvements'):
                improvement_ids = json.loads(employee['tenets_improvements']) if isinstance(employee['tenets_improvements'], str) else employee['tenets_improvements']
                improvements = [tenets_map.get(tid, tid) for tid in improvement_ids if tid in tenets_map]
        except Exception as e:
            print(f"Error parsing tenets for {employee.get('Associate')}: {e}")

        # Build structured description text (human-readable and machine-parseable)
        description_lines = []

        # Performance rating
        if employee.get('performance_rating_percent'):
            description_lines.append(f"Performance Rating: {employee['performance_rating_percent']}%")

        # Justification
        if employee.get('justification'):
            description_lines.append(f"Justification: {employee['justification']}")

        # Mentor/Mentee
        if employee.get('mentor'):
            description_lines.append(f"Mentor: {employee['mentor']}")
        if employee.get('mentees'):
            description_lines.append(f"Mentees: {employee['mentees']}")

        # Tenets
        if strengths:
            description_lines.append(f"Strengths: {', '.join(strengths)}")
        if improvements:
            description_lines.append(f"Areas for Improvement: {', '.join(improvements)}")

        description_text = '\n'.join(description_lines)

        # Calculate bonus percent of target
        bonus_percent_of_target = result['bonus_percent_of_target']

        # Format tenets as text for display
        strengths_text = ', '.join(strengths) if strengths else ''
        improvements_text = ', '.join(improvements) if improvements else ''

        # Build structured description text using format_notes_field
        from notes_parser import format_notes_field
        description_text = format_notes_field(
            performance_rating=employee.get('performance_rating_percent'),
            justification=employee.get('justification'),
            mentor=employee.get('mentor'),
            mentees=employee.get('mentees'),
            tenets_strengths=strengths_text,
            tenets_improvements=improvements_text,
            bonus_override_percent=employee.get('bonus_override_percent'),
            special_case_notes=employee.get('special_case_notes'),
        )

        # Calculate bonus percent of target
        bonus_percent_of_target = result['bonus_percent_of_target']

        # Build tool additions text for copying (formatted for Workday paste)
        tool_additions_parts = []
        if employee.get('performance_rating_percent') is not None:
            tool_additions_parts.append(f"[Performance Rating: {employee['performance_rating_percent']:g}%]")
        # Special case override (pro-rata leave, etc.)
        if employee.get('bonus_override_percent') is not None:
            override_fmt = f"{employee['bonus_override_percent']:g}"
            if employee.get('special_case_notes'):
                tool_additions_parts.append(f"[Override: {override_fmt}%, {employee['special_case_notes']}]")
            else:
                tool_additions_parts.append(f"[Override: {override_fmt}%]")
        if strengths_text:
            tool_additions_parts.append(f"[Strengths: {strengths_text}]")
        if improvements_text:
            tool_additions_parts.append(f"[Improvements: {improvements_text}]")
        if employee.get('mentor'):
            tool_additions_parts.append(f"[Mentor: {employee['mentor']}]")
        if employee.get('mentees'):
            # Normalize to semicolon-separated for consistency
            normalized_mentees = '; '.join(m.strip() for m in employee['mentees'].replace(';', ',').split(',') if m.strip())
            tool_additions_parts.append(f"[Mentees: {normalized_mentees}]")
        # Justification at the end with section header (allows multi-line)
        if employee.get('justification'):
            tool_additions_parts.append('')  # Blank line separator
            tool_additions_parts.append('Justification:')
            tool_additions_parts.append(employee['justification'])

        # Compare calculated bonus to Workday's proposed bonus
        workday_proposed_bonus = employee.get('Proposed Percent of Target Bonus')
        bonus_allocation_differs = False
        bonus_delta = None  # Delta: calculated - workday (positive = tool gives more)
        if bonus_percent_of_target is not None:
            if workday_proposed_bonus is not None:
                # Compare as integers - both should be whole numbers
                bonus_allocation_differs = int(bonus_percent_of_target) != round(workday_proposed_bonus)
                bonus_delta = int(bonus_percent_of_target) - round(workday_proposed_bonus)
            else:
                # Workday has no value yet - needs sync if tool calculated something
                bonus_allocation_differs = True

        # Check modification status for each field
        rating_modified = needs_sync_to_workday(employee, 'performance_rating_percent')
        justification_modified = is_field_modified(employee, 'justification')
        mentor_modified = needs_sync_to_workday(employee, 'mentor')
        mentees_modified = needs_sync_to_workday(employee, 'mentees')
        tenets_strengths_modified = needs_sync_to_workday(employee, 'tenets_strengths')
        tenets_improvements_modified = needs_sync_to_workday(employee, 'tenets_improvements')

        # Override uses `is not None` (not bool) because 0.0 is a valid override value
        override_current = employee.get('bonus_override_percent')
        override_original = employee.get('bonus_override_percent_original')
        override_modified = (
            override_current is not None and override_original is None
        ) or (
            override_current != override_original
        )

        # Combined flags - all tracked fields are in Tool Additions (bracketed)
        tool_additions_modified = (
            rating_modified or
            justification_modified or
            mentor_modified or
            mentees_modified or
            tenets_strengths_modified or
            tenets_improvements_modified or
            override_modified
        )

        # Combined "needs sync" flag - tool additions OR bonus allocation differs
        needs_sync = tool_additions_modified or bonus_allocation_differs

        if needs_sync:
            bonus_pending_sync_count += 1

        export_data.append({
            'employee': employee,
            'bonus_percent': bonus_percent_of_target,  # Already integer from calculation
            'description': description_text,
            'final_bonus': result['final_bonus'],
            'rating': result['rating'],
            # Tool Additions (bracketed metadata - parsed on re-import)
            'tool_additions_text': '\n'.join(tool_additions_parts),
            'tool_additions_modified': tool_additions_modified,
            # Bonus allocation comparison (calculated vs Workday)
            'bonus_allocation_differs': bonus_allocation_differs,
            'workday_proposed_bonus': round(workday_proposed_bonus, 1) if workday_proposed_bonus is not None else None,
            'bonus_delta': bonus_delta,  # Calculated - Workday (positive = tool gives more)
            # Per-field modification tracking
            'rating_modified': rating_modified,
            'justification_modified': justification_modified,
            'mentor_modified': mentor_modified,
            'mentees_modified': mentees_modified,
            'tenets_strengths_modified': tenets_strengths_modified,
            'tenets_improvements_modified': tenets_improvements_modified,
            'needs_sync': needs_sync,
            # Current field values for display
            'mentor': employee.get('mentor') or '',
            'mentees': employee.get('mentees') or '',
            'tenets_strengths': strengths_text,
            'tenets_improvements': improvements_text,
        })

    # Sort by employee name
    export_data.sort(key=lambda x: x['employee']['Associate'])

    # Build talent export data
    talent_export_data = []
    for emp in calibrated_employees:
        # Parse talent tenets
        talent_strengths = []
        talent_improvements = []
        try:
            if emp.get('talent_tenets_strengths'):
                strength_ids = json.loads(emp['talent_tenets_strengths']) if isinstance(emp['talent_tenets_strengths'], str) else emp['talent_tenets_strengths']
                talent_strengths = [tenets_map.get(tid, tid) for tid in strength_ids if tid in tenets_map]
            if emp.get('talent_tenets_improvements'):
                improvement_ids = json.loads(emp['talent_tenets_improvements']) if isinstance(emp['talent_tenets_improvements'], str) else emp['talent_tenets_improvements']
                talent_improvements = [tenets_map.get(tid, tid) for tid in improvement_ids if tid in tenets_map]
        except Exception as e:
            print(f"Error parsing talent tenets for {emp.get('Associate')}: {e}")

        # Build proposed actions text (includes embedded tenets for Workday)
        proposed_actions_parts = []
        if emp.get('talent_proposed_actions'):
            proposed_actions_parts.append(emp['talent_proposed_actions'])
        if talent_strengths:
            proposed_actions_parts.append(f"[Strengths: {', '.join(talent_strengths)}]")
        if talent_improvements:
            proposed_actions_parts.append(f"[Improvements: {', '.join(talent_improvements)}]")
        proposed_actions_text = ' '.join(proposed_actions_parts)

        talent_export_data.append({
            'employee': emp,
            'overall_perf': emp.get('talent_overall_perf', ''),
            'perf_what': emp.get('talent_perf_what', ''),
            'perf_how': emp.get('talent_perf_how', ''),
            'growth_agility': emp.get('talent_growth_agility', ''),
            'change_agility': emp.get('talent_change_agility', ''),
            'identified_future': emp.get('talent_identified_future', False),
            'movement_readiness': emp.get('talent_movement_readiness', ''),
            'proposed_actions': proposed_actions_text,
            'promo_job_profile': emp.get('talent_promo_job_profile', ''),
        })

    # Sort talent export data by employee name
    talent_export_data.sort(key=lambda x: x['employee']['Associate'])

    # Count employees needing sync
    pending_sync_count = sum(1 for item in talent_export_data if item.get('needs_sync'))

    # Get history period count for snapshot stats
    db = models.get_db()
    try:
        history_period_count = db.query(Period).count()
    finally:
        db.close()

    return render_template('export.html',
                         export_data=export_data,
                         talent_export_data=talent_export_data,
                         has_data=True,
                         has_bonus_data=has_bonus_data,
                         has_talent_data=has_talent_data,
                         export_mode=export_mode,
                         total_employees=len(team_data),
                         rated_count=len(rated_employees),
                         calibrated_count=len(calibrated_employees),
                         total_calibrated=len(talent_export_data),
                         bonus_pending_sync_count=bonus_pending_sync_count,
                         pending_sync_count=pending_sync_count,
                         history_period_count=history_period_count,
                         filter_info=filter_info)


@export_bp.route('/export/csv')
def export_csv():
    """Export employee data as CSV (same content as Excel)."""
    # Get filter params from URL
    filter_params = get_filter_params()

    # Get all employees (bonus cycle only)
    all_employees = get_all_employees(bonus_cycle_only=True)

    # Apply filters
    team_data, filter_info = apply_employee_filters(all_employees, filter_params)

    # Load tenets for description
    _, tenets_map = load_tenets_config()

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Add demo mode warning header if in demo mode
    if demo_mode.DEMO_MODE:
        writer.writerow(['*** DEMO MODE - FICTITIOUS DATA ONLY ***'])
        writer.writerow(['This export contains sample data for demonstration purposes.'])
        writer.writerow(['Do NOT use this data for any real business decisions.'])
        writer.writerow([])

    # Write header (matching to_dict() keys for data access)
    writer.writerow([
        'Associate ID',
        'Associate',
        'Supervisory Organization',
        'Current Job Profile',
        'Photo',
        'Errors',
        'Current Base Pay All Countries',
        'Current Base Pay Manager Currency',
        'Currency',
        'Grade',
        'Annual Bonus Target Percent',
        'Last Bonus Allocation Percent',
        'Bonus Target - Local Currency',
        'Bonus Target Manager Currency',
        'Proposed Bonus Amount',
        'Proposed Bonus Amount Manager Currency',
        'Proposed Percent of Target Bonus',
        'Notes',
        'Zero Bonus Allocated',
        'Performance Rating Percent',
        'Justification',
        'Mentor',
        'Mentees',
        'Tenets Strengths',
        'Tenets Improvements',
        'Description'
    ])

    # Write data rows
    for employee in team_data:
        # Parse tenets
        strengths_text = ''
        improvements_text = ''
        description_parts = []

        try:
            if employee.get('tenets_strengths'):
                strength_ids = json.loads(employee['tenets_strengths']) if isinstance(employee['tenets_strengths'], str) else employee['tenets_strengths']
                strengths = [tenets_map.get(tid, tid) for tid in strength_ids if tid in tenets_map]
                strengths_text = ', '.join(strengths)

            if employee.get('tenets_improvements'):
                improvement_ids = json.loads(employee['tenets_improvements']) if isinstance(employee['tenets_improvements'], str) else employee['tenets_improvements']
                improvements = [tenets_map.get(tid, tid) for tid in improvement_ids if tid in tenets_map]
                improvements_text = ', '.join(improvements)
        except Exception as e:
            print(f"Error parsing tenets: {e}")

        # Build description
        if employee.get('performance_rating_percent'):
            description_parts.append(f"Performance Rating: {employee['performance_rating_percent']}%")
        if employee.get('justification'):
            description_parts.append(f"Justification: {employee['justification']}")
        if employee.get('mentor'):
            description_parts.append(f"Mentor: {employee['mentor']}")
        if employee.get('mentees'):
            description_parts.append(f"Mentees: {employee['mentees']}")
        if strengths_text:
            description_parts.append(f"Strengths: {strengths_text}")
        if improvements_text:
            description_parts.append(f"Areas for Improvement: {improvements_text}")

        description = '\n'.join(description_parts)

        writer.writerow([
            employee.get('Associate ID', ''),
            employee.get('Associate', ''),
            employee.get('Supervisory Organization', ''),
            employee.get('Current Job Profile', ''),
            employee.get('Photo', ''),
            employee.get('Errors', ''),
            employee.get('Current Base Pay All Countries', ''),
            employee.get('Current Base Pay Manager Currency', ''),
            employee.get('Currency', ''),
            employee.get('Grade', ''),
            employee.get('Annual Bonus Target Percent', ''),
            employee.get('Last Bonus Allocation Percent', ''),
            employee.get('Bonus Target - Local Currency', ''),
            employee.get('Bonus Target Manager Currency', ''),
            employee.get('Proposed Bonus Amount', ''),
            employee.get('Proposed Bonus Amount Manager Currency', ''),
            employee.get('Proposed Percent of Target Bonus', ''),
            employee.get('Notes', ''),
            employee.get('Zero Bonus Allocated', ''),
            employee.get('performance_rating_percent', ''),
            employee.get('justification', ''),
            employee.get('mentor', ''),
            employee.get('mentees', ''),
            strengths_text,
            improvements_text,
            description
        ])

    # Create response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=performance_export.csv'
    return response


@export_bp.route('/export/xlsx')
def export_xlsx():
    """Export employee data as Excel file with all fields."""
    # Get filter params from URL
    filter_params = get_filter_params()

    # Get all employees (bonus cycle only)
    all_employees = get_all_employees(bonus_cycle_only=True)

    # Apply filters
    team_data, filter_info = apply_employee_filters(all_employees, filter_params)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Data"

    # Add demo mode warning if in demo mode
    demo_row_offset = 0
    if demo_mode.DEMO_MODE:
        demo_warning_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        demo_warning_font = Font(bold=True, color='FFFFFF', size=14)

        ws.merge_cells('A1:Z1')
        demo_cell = ws.cell(row=1, column=1, value='*** DEMO MODE - FICTITIOUS DATA ONLY - DO NOT USE FOR BUSINESS DECISIONS ***')
        demo_cell.fill = demo_warning_fill
        demo_cell.font = demo_warning_font
        demo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        demo_row_offset = 2  # Skip 2 rows for demo header

    # Define headers (matching to_dict() keys for data access)
    headers = [
        'Associate ID',
        'Associate',
        'Supervisory Organization',
        'Current Job Profile',
        'Photo',
        'Errors',
        'Current Base Pay All Countries',
        'Current Base Pay Manager Currency',
        'Currency',
        'Grade',
        'Annual Bonus Target Percent',
        'Last Bonus Allocation Percent',
        'Bonus Target - Local Currency',
        'Bonus Target Manager Currency',
        'Proposed Bonus Amount',
        'Proposed Bonus Amount Manager Currency',
        'Proposed Percent of Target Bonus',
        'Notes',
        'Zero Bonus Allocated',
        # Our custom fields
        'Performance Rating Percent',
        'Justification',
        'Mentor',
        'Mentees',
        'Tenets Strengths',
        'Tenets Improvements',
        'Description'  # Combined description field
    ]

    # Style header row
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    header_row = 1 + demo_row_offset
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Load tenets for description
    _, tenets_map = load_tenets_config()

    # Write data rows
    data_start_row = 2 + demo_row_offset
    for row_num, employee in enumerate(team_data, data_start_row):
        # Parse tenets
        strengths_text = ''
        improvements_text = ''
        description_parts = []

        try:
            if employee.get('tenets_strengths'):
                strength_ids = json.loads(employee['tenets_strengths']) if isinstance(employee['tenets_strengths'], str) else employee['tenets_strengths']
                strengths = [tenets_map.get(tid, tid) for tid in strength_ids if tid in tenets_map]
                strengths_text = ', '.join(strengths)

            if employee.get('tenets_improvements'):
                improvement_ids = json.loads(employee['tenets_improvements']) if isinstance(employee['tenets_improvements'], str) else employee['tenets_improvements']
                improvements = [tenets_map.get(tid, tid) for tid in improvement_ids if tid in tenets_map]
                improvements_text = ', '.join(improvements)
        except Exception as e:
            print(f"Error parsing tenets: {e}")

        # Build description
        if employee.get('performance_rating_percent'):
            description_parts.append(f"Performance Rating: {employee['performance_rating_percent']}%")
        if employee.get('justification'):
            description_parts.append(f"Justification: {employee['justification']}")
        if employee.get('mentor'):
            description_parts.append(f"Mentor: {employee['mentor']}")
        if employee.get('mentees'):
            description_parts.append(f"Mentees: {employee['mentees']}")
        if strengths_text:
            description_parts.append(f"Strengths: {strengths_text}")
        if improvements_text:
            description_parts.append(f"Areas for Improvement: {improvements_text}")

        description = '\n'.join(description_parts)

        row_data = [
            employee.get('Associate ID', ''),
            employee.get('Associate', ''),
            employee.get('Supervisory Organization', ''),
            employee.get('Current Job Profile', ''),
            employee.get('Photo', ''),
            employee.get('Errors', ''),
            employee.get('Current Base Pay All Countries', ''),
            employee.get('Current Base Pay Manager Currency', ''),
            employee.get('Currency', ''),
            employee.get('Grade', ''),
            employee.get('Annual Bonus Target Percent', ''),
            employee.get('Last Bonus Allocation Percent', ''),
            employee.get('Bonus Target - Local Currency', ''),
            employee.get('Bonus Target Manager Currency', ''),
            employee.get('Proposed Bonus Amount', ''),
            employee.get('Proposed Bonus Amount Manager Currency', ''),
            employee.get('Proposed Percent of Target Bonus', ''),
            employee.get('Notes', ''),
            employee.get('Zero Bonus Allocated', ''),
            # Our custom fields
            employee.get('performance_rating_percent', ''),
            employee.get('justification', ''),
            employee.get('mentor', ''),
            employee.get('mentees', ''),
            strengths_text,
            improvements_text,
            description
        ]

        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths (skip merged cells which don't have column_letter)
    from openpyxl.utils import get_column_letter
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if cell.value and hasattr(cell, 'column_letter'):
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        if max_length > 0:
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='performance_export.xlsx'
    )


@export_bp.route('/export/talent')
def export_talent_xlsx():
    """Export talent calibration data as Excel file.

    Embeds tenets in the Proposed Actions field using a parseable format
    so that Workday becomes the source of truth and re-imports preserve data.

    Format: [Strengths: Tenet1, Tenet2] [Improvements: Tenet3]
    """
    # Get filter params from URL
    filter_params = get_filter_params()

    # Get all employees
    all_employees = get_all_employees()

    # Apply filters
    team_data, filter_info = apply_employee_filters(all_employees, filter_params)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Talent Calibration"

    # Add demo mode warning if in demo mode
    demo_row_offset = 0
    if demo_mode.DEMO_MODE:
        demo_warning_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        demo_warning_font = Font(bold=True, color='FFFFFF', size=14)

        ws.merge_cells('A1:Z1')
        demo_cell = ws.cell(row=1, column=1, value='*** DEMO MODE - FICTITIOUS DATA ONLY - DO NOT USE FOR BUSINESS DECISIONS ***')
        demo_cell.fill = demo_warning_fill
        demo_cell.font = demo_warning_font
        demo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        demo_row_offset = 2

    # Define headers matching Workday talent calibration format
    headers = [
        'Associate ID',
        'Worker',  # Workday uses 'Worker' not 'Associate'
        'Supervisory Organization',
        'Current Job Profile',
        'Performance: What',
        'Performance: How',
        'Overall Performance',
        'Future Talent: Growth Agility',
        'Future Talent: Change Agility',
        'Identified Future Talent',
        'Movement Readiness',
        'Proposed Talent Actions',  # Contains embedded tenets (matches import column map)
        'Promotions: Proposed Job Profile & Code',
        'Promotions: Business Need',
        'Promotions: Expanded Role Scope',
        'Promotions: Associate Readiness',
    ]

    # Style header row
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    header_row = 1 + demo_row_offset
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Load tenets for name lookup
    _, tenets_map = load_tenets_config()

    # Write data rows
    data_start_row = 2 + demo_row_offset
    for row_num, employee in enumerate(team_data, data_start_row):
        # Build Proposed Actions with embedded tenets
        proposed_actions = employee.get('talent_proposed_actions') or ''

        # Parse and format tenets
        strengths_text = ''
        improvements_text = ''

        try:
            if employee.get('talent_tenets_strengths'):
                strength_ids = json.loads(employee['talent_tenets_strengths']) if isinstance(employee['talent_tenets_strengths'], str) else employee['talent_tenets_strengths']
                strengths = [tenets_map.get(tid, tid) for tid in strength_ids if tid in tenets_map]
                if strengths:
                    strengths_text = '; '.join(strengths)  # Use semicolon (tenet names may contain commas)

            if employee.get('talent_tenets_improvements'):
                improvement_ids = json.loads(employee['talent_tenets_improvements']) if isinstance(employee['talent_tenets_improvements'], str) else employee['talent_tenets_improvements']
                improvements = [tenets_map.get(tid, tid) for tid in improvement_ids if tid in tenets_map]
                if improvements:
                    improvements_text = '; '.join(improvements)  # Use semicolon
        except Exception as e:
            print(f"Error parsing talent tenets: {e}")

        # Embed tenets and mentor/mentees in Proposed Actions using parseable format
        metadata_markers = []
        if strengths_text:
            metadata_markers.append(f"[Strengths: {strengths_text}]")
        if improvements_text:
            metadata_markers.append(f"[Improvements: {improvements_text}]")

        # Embed mentor/mentees
        talent_mentor = employee.get('talent_mentor', '')
        talent_mentees = employee.get('talent_mentees', '')
        if talent_mentor:
            metadata_markers.append(f"[Mentor: {talent_mentor}]")
        if talent_mentees:
            # Normalize to semicolon-separated for consistency with tenets
            normalized_mentees = '; '.join(m.strip() for m in talent_mentees.replace(';', ',').split(',') if m.strip())
            metadata_markers.append(f"[Mentees: {normalized_mentees}]")

        if metadata_markers:
            # Append to proposed actions with separator
            if proposed_actions:
                proposed_actions = proposed_actions.rstrip() + '\n\n' + ' '.join(metadata_markers)
            else:
                proposed_actions = ' '.join(metadata_markers)

        row_data = [
            employee.get('Associate ID', ''),
            employee.get('Associate', ''),  # Export as Worker column
            employee.get('Supervisory Organization', ''),
            employee.get('Current Job Profile', ''),
            employee.get('talent_perf_what', ''),
            employee.get('talent_perf_how', ''),
            employee.get('talent_overall_perf', ''),
            employee.get('talent_growth_agility', ''),
            employee.get('talent_change_agility', ''),
            'Yes' if employee.get('talent_identified_future') else 'No' if employee.get('talent_identified_future') is False else '',
            employee.get('talent_movement_readiness', ''),
            proposed_actions,
            employee.get('talent_promo_job_profile', ''),
            employee.get('talent_promo_business_need', ''),
            employee.get('talent_promo_role_scope', ''),
            employee.get('talent_promo_readiness', ''),
        ]

        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if cell.value and hasattr(cell, 'column_letter'):
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        if max_length > 0:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='talent_calibration_export.xlsx'
    )


@export_bp.route('/export/talent/csv')
def export_talent_csv():
    """Export talent calibration data as CSV."""
    # Get filter params from URL
    filter_params = get_filter_params()

    # Get all employees
    all_employees = get_all_employees()

    # Apply filters
    team_data, filter_info = apply_employee_filters(all_employees, filter_params)

    # Load tenets for name lookup
    _, tenets_map = load_tenets_config()

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Add demo mode warning header if in demo mode
    if demo_mode.DEMO_MODE:
        writer.writerow(['*** DEMO MODE - FICTITIOUS DATA ONLY ***'])
        writer.writerow(['This export contains sample data for demonstration purposes.'])
        writer.writerow(['Do NOT use this data for any real business decisions.'])
        writer.writerow([])

    # Write header (matching Workday talent calibration format)
    writer.writerow([
        'Associate ID',
        'Worker',
        'Supervisory Organization',
        'Current Job Profile',
        'Performance: What',
        'Performance: How',
        'Overall Performance',
        'Future Talent: Growth Agility',
        'Future Talent: Change Agility',
        'Identified Future Talent',
        'Movement Readiness',
        'Proposed Talent Actions',
        'Promotions: Proposed Job Profile & Code',
        'Promotions: Business Need',
        'Promotions: Expanded Role Scope',
        'Promotions: Associate Readiness',
    ])

    # Write data rows
    for employee in team_data:
        # Build Proposed Actions with embedded tenets
        proposed_actions = employee.get('talent_proposed_actions') or ''

        # Parse and format tenets
        strengths_text = ''
        improvements_text = ''

        try:
            if employee.get('talent_tenets_strengths'):
                strength_ids = json.loads(employee['talent_tenets_strengths']) if isinstance(employee['talent_tenets_strengths'], str) else employee['talent_tenets_strengths']
                strengths = [tenets_map.get(tid, tid) for tid in strength_ids if tid in tenets_map]
                if strengths:
                    strengths_text = '; '.join(strengths)

            if employee.get('talent_tenets_improvements'):
                improvement_ids = json.loads(employee['talent_tenets_improvements']) if isinstance(employee['talent_tenets_improvements'], str) else employee['talent_tenets_improvements']
                improvements = [tenets_map.get(tid, tid) for tid in improvement_ids if tid in tenets_map]
                if improvements:
                    improvements_text = '; '.join(improvements)
        except Exception as e:
            print(f"Error parsing talent tenets: {e}")

        # Embed tenets and mentor/mentees in Proposed Actions
        metadata_markers = []
        if strengths_text:
            metadata_markers.append(f"[Strengths: {strengths_text}]")
        if improvements_text:
            metadata_markers.append(f"[Improvements: {improvements_text}]")

        # Embed mentor/mentees
        talent_mentor = employee.get('talent_mentor', '')
        talent_mentees = employee.get('talent_mentees', '')
        if talent_mentor:
            metadata_markers.append(f"[Mentor: {talent_mentor}]")
        if talent_mentees:
            # Normalize to semicolon-separated for consistency with tenets
            normalized_mentees = '; '.join(m.strip() for m in talent_mentees.replace(';', ',').split(',') if m.strip())
            metadata_markers.append(f"[Mentees: {normalized_mentees}]")

        if metadata_markers:
            if proposed_actions:
                proposed_actions = proposed_actions.rstrip() + ' ' + ' '.join(metadata_markers)
            else:
                proposed_actions = ' '.join(metadata_markers)

        writer.writerow([
            employee.get('Associate ID', ''),
            employee.get('Associate', ''),
            employee.get('Supervisory Organization', ''),
            employee.get('Current Job Profile', ''),
            employee.get('talent_perf_what', ''),
            employee.get('talent_perf_how', ''),
            employee.get('talent_overall_perf', ''),
            employee.get('talent_growth_agility', ''),
            employee.get('talent_change_agility', ''),
            'Yes' if employee.get('talent_identified_future') else 'No' if employee.get('talent_identified_future') is False else '',
            employee.get('talent_movement_readiness', ''),
            proposed_actions,
            employee.get('talent_promo_job_profile', ''),
            employee.get('talent_promo_business_need', ''),
            employee.get('talent_promo_role_scope', ''),
            employee.get('talent_promo_readiness', ''),
        ])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': 'attachment; filename=talent_calibration_export.csv'
        }
    )


@export_bp.route('/export/snapshot/xlsx')
def export_snapshot_xlsx():
    """Export full multi-tab Excel snapshot with complete organizational data.

    Creates a workbook with 5 sheets:
    - _README: Domain knowledge, rating philosophy, tenets (markdown for AI)
    - employees: Core identity, compensation, manager info
    - bonus_cycle: Performance ratings, justifications, calculated bonuses
    - talent_cycle: Calibration data, agility, movement, promotions
    - history: Historical rating snapshots by period
    """
    from models import get_cross_cycle_alignment

    # Get all employees
    all_employees = get_all_employees()

    # Load tenets config
    tenets_config, tenets_map = load_tenets_config()

    # Calculate bonuses for rated employees
    rated_employees = [e for e in all_employees if is_employee_rated(e)]
    params = {'upside_exponent': 1.35, 'downside_exponent': 1.9}

    bonus_results = {}
    if rated_employees:
        bonus_calc = calculate_bonus_for_employees(rated_employees, params)
        bonus_results = bonus_calc.get('results_by_id', {})

    # Get history snapshots
    history_data = get_all_history_snapshots()

    # Build all sheet rows via the shared snapshot builders (services/export.py)
    sheet_data = prepare_snapshot_data(
        all_employees, tenets_map, bonus_results, history_data,
        get_cross_cycle_alignment
    )

    # Create workbook
    wb = Workbook()

    # Sheet 1: _README (markdown content for AI consumption)
    ws_readme = wb.active
    ws_readme.title = "_README"
    readme_content = build_context_markdown(tenets_config, demo_mode=demo_mode.DEMO_MODE)
    cell = ws_readme.cell(row=1, column=1, value=readme_content)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws_readme.column_dimensions['A'].width = 100

    # Data sheets: headers + rows written via shared helper
    write_xlsx_sheet(wb.create_sheet("employees"),
                     SNAPSHOT_EMPLOYEE_HEADERS, sheet_data['employees'])
    write_xlsx_sheet(wb.create_sheet("bonus_cycle"),
                     SNAPSHOT_BONUS_HEADERS, sheet_data['bonus_cycle'])
    write_xlsx_sheet(wb.create_sheet("talent_cycle"),
                     SNAPSHOT_TALENT_HEADERS, sheet_data['talent_cycle'])
    write_xlsx_sheet(wb.create_sheet("history"),
                     SNAPSHOT_HISTORY_HEADERS, sheet_data['history'])

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='organization_snapshot.xlsx'
    )


@export_bp.route('/export/snapshot/csv')
def export_snapshot_csv():
    """Export full ZIP with CSV files and README containing complete organizational data.

    Creates a ZIP archive with 5 files:
    - README.md: Domain knowledge, rating philosophy, tenets (for AI consumption)
    - employees.csv: Core identity and compensation
    - bonus_cycle.csv: Performance ratings and bonuses
    - talent_cycle.csv: Calibration and development data
    - history.csv: Historical snapshots
    """
    from models import get_cross_cycle_alignment
    import zipfile

    # Get all employees
    all_employees = get_all_employees()

    # Load tenets config
    tenets_config, tenets_map = load_tenets_config()

    # Calculate bonuses for rated employees
    rated_employees = [e for e in all_employees if is_employee_rated(e)]
    params = {'upside_exponent': 1.35, 'downside_exponent': 1.9}

    bonus_results = {}
    if rated_employees:
        bonus_calc = calculate_bonus_for_employees(rated_employees, params)
        bonus_results = bonus_calc.get('results_by_id', {})

    # Get history snapshots
    history_data = get_all_history_snapshots()

    # Build all sheet rows via the shared snapshot builders (services/export.py)
    sheet_data = prepare_snapshot_data(
        all_employees, tenets_map, bonus_results, history_data,
        get_cross_cycle_alignment
    )

    def _csv_bytes(headers, rows):
        """Serialize one sheet (header + value rows) to CSV text."""
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        return buf.getvalue()

    # Create ZIP in memory
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # README.md: Human-readable context for AI consumption
        readme_content = build_context_markdown(tenets_config, demo_mode=demo_mode.DEMO_MODE)
        zip_file.writestr('README.md', readme_content)

        zip_file.writestr('employees.csv',
                          _csv_bytes(SNAPSHOT_EMPLOYEE_HEADERS, sheet_data['employees']))
        zip_file.writestr('bonus_cycle.csv',
                          _csv_bytes(SNAPSHOT_BONUS_HEADERS, sheet_data['bonus_cycle']))
        zip_file.writestr('talent_cycle.csv',
                          _csv_bytes(SNAPSHOT_TALENT_HEADERS, sheet_data['talent_cycle']))
        zip_file.writestr('history.csv',
                          _csv_bytes(SNAPSHOT_HISTORY_HEADERS, sheet_data['history']))

    zip_buffer.seek(0)

    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name='organization_snapshot.zip'
    )

